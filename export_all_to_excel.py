#!/usr/bin/env python3
"""
石竹烯-CIRI 项目全结果汇总导出为 Excel
整合 GAT 预测结果、训练指标、节点特征、边信息到一个工作簿
所有统计值均从数据文件动态计算，避免硬编码
"""

import json
import pickle
from pathlib import Path

import numpy as np
import pandas as pd
import torch
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


def load_config(config_path: str = "config.yaml") -> dict:
    with open(config_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def load_training_log(log_dir: Path) -> dict:
    """从日志文件提取最佳模型性能指标"""
    log_file = log_dir / "training_log.txt"
    metrics = {
        "best_epoch": None,
        "val_roc_auc": None,
        "val_f1": None,
    }
    test_metrics = {}

    if not log_file.exists():
        return metrics, test_metrics

    with open(log_file, "r", encoding="utf-8") as f:
        for line in f:
            # 解析最佳 epoch
            if "新最佳模型" in line and "Val ROC-AUC=" in line:
                # 格式: Epoch 023: 新最佳模型，Val ROC-AUC=0.8145, Val F1=0.6275
                try:
                    parts = line.split("Epoch ")[1].split(":")
                    metrics["best_epoch"] = int(parts[0].strip())
                    val_part = line.split("Val ROC-AUC=")[1].split(",")[0]
                    metrics["val_roc_auc"] = float(val_part)
                    f1_part = line.split("Val F1=")[1].strip()
                    metrics["val_f1"] = float(f1_part)
                except Exception:
                    pass
            # 解析测试集指标
            if "测试集指标:" in line:
                # 下一行开始是指标
                continue
            if line.strip().startswith("  ") and ":" in line:
                try:
                    key_val = line.strip().split(":")
                    key = key_val[0].strip()
                    val = float(key_val[1].strip())
                    test_metrics[key] = val
                except Exception:
                    pass

    return metrics, test_metrics


def load_model_info(model_dir: Path, config: dict) -> dict:
    """加载模型检查点信息"""
    best_model_path = model_dir / config["output"]["best_model"]
    info = {
        "total_params": 0,
        "best_epoch": None,
        "val_roc_auc": None,
        "val_f1": None,
    }
    if best_model_path.exists():
        try:
            checkpoint = torch.load(best_model_path, map_location="cpu", weights_only=False)
            info["best_epoch"] = checkpoint.get("epoch")
            info["val_roc_auc"] = checkpoint.get("val_roc_auc")
            info["val_f1"] = checkpoint.get("val_f1")
        except Exception:
            pass
    return info


def create_comprehensive_excel(output_path: str = "./石竹烯-CIRI_全结果汇总.xlsx"):
    """创建综合 Excel 汇总报告（全部动态计算）"""

    # ------------------------------------------------------------------
    # 动态加载所有数据
    # ------------------------------------------------------------------
    config = load_config("config.yaml")
    data_cfg = config["data"]
    model_cfg = config["model"]
    train_cfg = config["training"]
    out_cfg = config["output"]

    processed_dir = Path(data_cfg["processed_dir"])
    results_dir = Path(out_cfg["results_dir"])
    model_dir = Path(out_cfg["model_dir"])
    logs_dir = Path(out_cfg["logs_dir"])

    # 节点特征
    node_df = pd.read_csv(processed_dir / "node_features.csv")
    gene_pool_size = len(node_df)
    in_channels = len([c for c in node_df.columns if c != "GeneSymbol"])

    # 标签
    labels_df = pd.read_csv(processed_dir / "labels.csv")
    n_pos = (labels_df["Label"] == 1).sum()
    n_neg = (labels_df["Label"] == 0).sum()
    n_unk = (labels_df["Label"] == -1).sum()

    # 边
    edges_df = pd.read_csv(processed_dir / "edge_index.csv")
    n_edges = len(edges_df)

    # 模型信息
    model_info = load_model_info(model_dir, config)
    # 计算参数量
    from gat_model import GAT
    dummy_model = GAT(
        in_channels=model_cfg["in_channels"],
        hidden_channels=model_cfg["hidden_channels"],
        out_channels=model_cfg["out_channels"],
        num_heads=model_cfg["num_heads"],
        num_classes=model_cfg["num_classes"],
        dropout=model_cfg["dropout"],
        attention_dropout=model_cfg["attention_dropout"],
        use_edge_attr=model_cfg.get("use_edge_attr", True),
        use_batch_norm=model_cfg["use_batch_norm"],
    )
    total_params = sum(p.numel() for p in dummy_model.parameters())

    # 训练日志
    log_metrics, test_metrics = load_training_log(logs_dir)

    # 使用日志中的指标，如果日志缺失则使用模型检查点
    best_epoch = log_metrics.get("best_epoch") or model_info.get("best_epoch") or "N/A"
    val_roc_auc = log_metrics.get("val_roc_auc") or model_info.get("val_roc_auc") or 0.0
    val_f1 = log_metrics.get("val_f1") or model_info.get("val_f1") or 0.0

    # ------------------------------------------------------------------
    # 创建 Excel
    # ------------------------------------------------------------------
    wb = Workbook()
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ========== Sheet 1: 项目概览 ==========
    ws_overview = wb.active
    ws_overview.title = "项目概览"

    overview_data = [
        ["石竹烯-CIRI 项目全结果汇总", ""],
        ["", ""],
        ["生成时间", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")],
        ["", ""],
        ["数据集统计", ""],
        ["基因池总数", gene_pool_size],
        ["阳性标签（已验证靶点）", int(n_pos)],
        ["阴性标签（非靶点）", int(n_neg)],
        ["未知标签（待预测）", int(n_unk)],
        ["STRING 边数", n_edges],
        ["", ""],
        ["模型架构", ""],
        ["输入维度", in_channels],
        ["隐藏层1", f"{model_cfg['hidden_channels']} channels × {model_cfg['num_heads']} heads"],
        ["隐藏层2", f"{model_cfg['out_channels']} channels × {model_cfg['num_heads']} heads"],
        ["输出维度", model_cfg["num_classes"]],
        ["总参数量", total_params],
        ["", ""],
        ["训练配置", ""],
        ["学习率", train_cfg["learning_rate"]],
        ["Weight Decay", train_cfg["weight_decay"]],
        ["Dropout", model_cfg["dropout"]],
        ["早停 Patience", train_cfg["patience"]],
        ["铜死亡先验权重", train_cfg["cupro_weight"]],
        ["", ""],
        ["最佳模型性能", ""],
        ["最佳 Epoch", best_epoch],
        ["验证集 ROC-AUC", val_roc_auc],
        ["验证集 F1", val_f1],
    ]

    # 动态添加测试集指标
    if test_metrics:
        for key, val in test_metrics.items():
            overview_data.append([f"测试集 {key}", val])

    for r_idx, row in enumerate(overview_data, 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_overview.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True, size=16, color="FFFFFF")
                cell.fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            elif value in ["数据集统计", "模型架构", "训练配置", "最佳模型性能"]:
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            elif c_idx == 1 and r_idx > 1:
                cell.font = Font(bold=True)
            if isinstance(value, float):
                cell.number_format = "0.0000"

    ws_overview.column_dimensions["A"].width = 25
    ws_overview.column_dimensions["B"].width = 20
    ws_overview.merge_cells("A1:B1")

    # ========== Sheet 2: Top50 候选靶点 ==========
    ws50 = wb.create_sheet("Top50候选靶点")
    df50 = pd.read_csv(results_dir / "top_targets_50.csv")
    df50.columns = ["基因符号", "靶点概率", "排名", "度中心性", "PageRank", "聚类系数", "铜死亡基因", "炎症相关"]

    for r_idx, row in enumerate(dataframe_to_rows(df50, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws50.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            else:
                if c_idx == 7 and value == 1:
                    cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                if c_idx == 8 and value == 1:
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                if c_idx == 2 and isinstance(value, float):
                    cell.number_format = "0.0000"

    for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws50.column_dimensions[col].width = 15

    # ========== Sheet 3: Top100 候选靶点 ==========
    ws100 = wb.create_sheet("Top100候选靶点")
    df100 = pd.read_csv(results_dir / "top_targets_100.csv")
    df100.columns = ["基因符号", "靶点概率", "排名", "度中心性", "PageRank", "聚类系数", "铜死亡基因", "炎症相关"]

    for r_idx, row in enumerate(dataframe_to_rows(df100, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws100.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            else:
                if c_idx == 7 and value == 1:
                    cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                if c_idx == 8 and value == 1:
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                if c_idx == 2 and isinstance(value, float):
                    cell.number_format = "0.0000"

    for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws100.column_dimensions[col].width = 15

    # ========== Sheet 4: Top200 候选靶点 ==========
    ws200 = wb.create_sheet("Top200候选靶点")
    df200 = pd.read_csv(results_dir / "top_targets_200.csv")
    df200.columns = ["基因符号", "靶点概率", "排名", "度中心性", "PageRank", "聚类系数", "铜死亡基因", "炎症相关"]

    for r_idx, row in enumerate(dataframe_to_rows(df200, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws200.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            else:
                if c_idx == 7 and value == 1:
                    cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                if c_idx == 8 and value == 1:
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                if c_idx == 2 and isinstance(value, float):
                    cell.number_format = "0.0000"

    for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws200.column_dimensions[col].width = 15

    # ========== Sheet 5: 铜死亡基因专项 ==========
    ws_cupro = wb.create_sheet("铜死亡基因专项")
    df_all = pd.read_csv(results_dir / "all_unknown_predictions.csv")
    df_all.columns = ["基因符号", "靶点概率", "排名", "度中心性", "PageRank", "聚类系数", "铜死亡基因", "炎症相关"]
    cupro_df = df_all[df_all["铜死亡基因"] == 1].sort_values("靶点概率", ascending=False)

    for r_idx, row in enumerate(dataframe_to_rows(cupro_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_cupro.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            else:
                if c_idx == 2 and isinstance(value, float):
                    cell.number_format = "0.0000"

    for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws_cupro.column_dimensions[col].width = 15

    # ========== Sheet 6: 全部未知节点预测 ==========
    ws_all = wb.create_sheet("全部未知节点预测")
    for r_idx, row in enumerate(dataframe_to_rows(df_all, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_all.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            else:
                if c_idx == 7 and value == 1:
                    cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                if c_idx == 8 and value == 1:
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                if c_idx == 2 and isinstance(value, float):
                    cell.number_format = "0.0000"

    for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws_all.column_dimensions[col].width = 15

    # ========== Sheet 7: 节点特征矩阵 ==========
    ws_features = wb.create_sheet("节点特征矩阵")
    features_df = pd.read_csv(processed_dir / "node_features.csv")

    for r_idx, row in enumerate(dataframe_to_rows(features_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_features.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            if isinstance(value, float):
                cell.number_format = "0.0000"

    for c_idx, col in enumerate(features_df.columns, 1):
        col_letter = get_column_letter(c_idx)
        ws_features.column_dimensions[col_letter].width = 12

    # ========== Sheet 8: 边索引 ==========
    ws_edges = wb.create_sheet("边索引")

    for r_idx, row in enumerate(dataframe_to_rows(edges_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_edges.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            if c_idx == 3 and isinstance(value, float):
                cell.number_format = "0.0000"

    ws_edges.column_dimensions["A"].width = 15
    ws_edges.column_dimensions["B"].width = 15
    ws_edges.column_dimensions["C"].width = 12
    ws_edges.column_dimensions["D"].width = 12

    # ========== Sheet 9: 标签信息 ==========
    ws_labels = wb.create_sheet("标签信息")
    labels_df.columns = ["基因符号", "标签"]
    labels_df["标签说明"] = labels_df["标签"].map({1: "阳性靶点", 0: "阴性非靶点", -1: "未知待预测"})

    for r_idx, row in enumerate(dataframe_to_rows(labels_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_labels.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            else:
                if c_idx == 2:
                    if value == 1:
                        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                    elif value == 0:
                        cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                    elif value == -1:
                        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    ws_labels.column_dimensions["A"].width = 15
    ws_labels.column_dimensions["B"].width = 12
    ws_labels.column_dimensions["C"].width = 15

    # 保存
    wb.save(output_path)
    print(f"综合 Excel 报告已生成: {output_path}")
    print(f"包含工作表: {wb.sheetnames}")


if __name__ == "__main__":
    create_comprehensive_excel()
