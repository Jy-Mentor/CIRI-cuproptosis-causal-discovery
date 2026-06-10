import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# 设置路径
work_dir = "C:/Users/Jy-Mentor-7/Desktop/你信我 这不是重蹈覆辙"
result_dir = os.path.join(work_dir, "String_Network_Systematic_Analysis")
output_file = os.path.join(work_dir, "String_Network_Systematic_Analysis_Summary.xlsx")

# 创建workbook
wb = Workbook()
wb.remove(wb.active)

# 定义样式
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
title_font = Font(bold=True, size=14, color="1F4E78")
subtitle_font = Font(bold=True, size=11)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ==================== Sheet 1: 分析概述 ====================
ws_summary = wb.create_sheet("分析概述")

summary_data = [
    ["String PPI网络系统性生物信息学分析报告"],
    [""],
    ["网络预处理"],
    ["原始节点数", 142],
    ["原始边数", 1944],
    ["孤立节点数", 7],
    ["孤立节点列表", "BST1, CNR2, MGAT1, RENBP, SERPINB10, STK4, TCN2"],
    ["预处理后节点数", 135],
    ["预处理后边数", 1944],
    [""],
    ["K-core分解结果"],
    ["K>=3核心节点数", 124],
    ["K>=3核心边数", 1924],
    ["K-core分布", "K=2: 11, K=4: 13, K=6: 15, K=8: 10, K=10: 7, ... K=38: 21"],
    [""],
    ["中心性指标说明"],
    ["DC", "度中心性 (Degree Centrality) - 连接数"],
    ["BC", "介数中心性 (Betweenness Centrality) - 桥接作用"],
    ["CC", "接近中心性 (Closeness Centrality) - 到达效率"],
    ["EC", "特征向量中心性 (Eigenvector Centrality) - 邻居重要性"],
    ["SC", "子图中心性 (Subgraph Centrality) - 需cytoHubba"],
    ["NC", "网络中心性 (Network Centrality) - 需cytoHubba"],
    ["LAC", "局部平均连通性 (Local Average Connectivity) - 需cytoHubba"],
    ["MCC", "最大团中心性 (Maximum Clique Centrality) - 需cytoHubba"],
    [""],
    ["RRA综合评分"],
    ["方法", "Robust Rank Aggregation (稳健排名聚合)"],
    ["整合指标", "DC, BC, CC, EC (R部分) + SC, NC, LAC, MCC (Cytoscape部分)"],
]

for row_idx, row_data in enumerate(summary_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.font = Font(bold=True, size=16, color="1F4E78")
            cell.alignment = Alignment(horizontal="center")
        elif value in ["网络预处理", "K-core分解结果", "中心性指标说明", "RRA综合评分"]:
            cell.font = subtitle_font
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

ws_summary.column_dimensions['A'].width = 25
ws_summary.column_dimensions['B'].width = 50

# ==================== Sheet 2: K-core分解结果 ====================
ws_kcore = wb.create_sheet("K-core分解结果")
kcore_df = pd.read_csv(os.path.join(result_dir, "02_kcore_decomposition.txt"), sep="\t")

ws_kcore['A1'] = "K-core分解结果"
ws_kcore['A1'].font = title_font

for r_idx, row in enumerate(dataframe_to_rows(kcore_df, index=False, header=True), 2):
    for c_idx, value in enumerate(row, 1):
        cell = ws_kcore.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        if r_idx == 2:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

ws_kcore.column_dimensions['A'].width = 15
ws_kcore.column_dimensions['B'].width = 15

# ==================== Sheet 3: 中心性指标汇总 ====================
ws_centrality = wb.create_sheet("中心性指标汇总")
centrality_df = pd.read_csv(os.path.join(result_dir, "03_centrality_measures.txt"), sep="\t")

ws_centrality['A1'] = "网络中心性指标汇总 (igraph计算)"
ws_centrality['A1'].font = title_font

for r_idx, row in enumerate(dataframe_to_rows(centrality_df, index=False, header=True), 2):
    for c_idx, value in enumerate(row, 1):
        cell = ws_centrality.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        if r_idx == 2:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_centrality.column_dimensions[col].width = 15

# ==================== Sheet 4: Top 20 Hub节点 (RRA) ====================
ws_top20 = wb.create_sheet("Top20 Hub节点 (RRA)")
top20_df = pd.read_csv(os.path.join(result_dir, "04_top20_hub_nodes_rra.txt"), sep="\t")

ws_top20['A1'] = "Top 20 Hub节点 (RRA综合评分)"
ws_top20['A1'].font = title_font

# 选择关键列
top20_display = top20_df[['Node', 'DC', 'BC', 'CC', 'EC', 'K_core', 'RRA_Score']]

for r_idx, row in enumerate(dataframe_to_rows(top20_display, index=False, header=True), 2):
    for c_idx, value in enumerate(row, 1):
        cell = ws_top20.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        if r_idx == 2:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
    ws_top20.column_dimensions[col].width = 15

# ==================== Sheet 5: 孤立节点列表 ====================
ws_isolated = wb.create_sheet("孤立节点列表")
isolated_df = pd.read_csv(os.path.join(result_dir, "01_isolated_nodes_removed.txt"), sep="\t")

ws_isolated['A1'] = "删除的孤立节点 (度为0)"
ws_isolated['A1'].font = title_font

ws_isolated['A2'] = "以下节点在网络中没有连接边，已删除:"

for r_idx, row in enumerate(dataframe_to_rows(isolated_df, index=False, header=True), 3):
    for c_idx, value in enumerate(row, 1):
        cell = ws_isolated.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        if r_idx == 3:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

ws_isolated.column_dimensions['A'].width = 20

# ==================== Sheet 6: Cytoscape操作指南 ====================
ws_guide = wb.create_sheet("Cytoscape操作指南")

guide_text = """
【MCODE聚类分析】
1. 打开Cytoscape软件
2. 导入网络文件: 05_cytoscape_edge_list.txt
   - File -> Import -> Network from File
   - 选择05_cytoscape_edge_list.txt
   - Source Column: source, Target Column: target

3. 导入节点属性: 05_cytoscape_node_attributes.txt
   - File -> Import -> Table from File
   - 选择05_cytoscape_node_attributes.txt
   - Key Column for Network: 选择Node列

4. 运行MCODE插件:
   - Apps -> MCODE
   - 参数设置:
     * Network Scoring: Degree Cutoff = 2, K-Core = 2
     * Cluster Finding: Node Score Cutoff = 0.2, Haircut = TRUE
   - 点击'Analyze Cluster'

5. 导出MCODE结果:
   - MCODE -> View/Export Clusters
   - 保存为CSV或Cytoscape会话文件

【cytoHubba中心性分析】
1. 确保网络已导入
2. 运行cytoHubba:
   - Apps -> cytoHubba
   - 选择要计算的指标:
     * SC (Subgraph Centrality)
     * NC (Network Centrality)
     * LAC (Local Average Connectivity)
     * MCC (Maximum Clique Centrality)
   - 选择网络 -> 选择节点 -> 点击'Calculate'

3. 导出cytoHubba结果:
   - cytoHubba -> Export Results
   - 保存为TXT或CSV格式

【综合分析建议】
1. 将cytoHubba结果与R计算的中心性指标合并
2. 使用RRA方法整合所有8种中心性指标:
   - DC (igraph) - 度中心性
   - BC (igraph) - 介数中心性
   - CC (igraph) - 接近中心性
   - EC (igraph) - 特征向量中心性
   - SC (cytoHubba) - 子图中心性
   - NC (cytoHubba) - 网络中心性
   - LAC (cytoHubba) - 局部平均连通性
   - MCC (cytoHubba) - 最大团中心性
3. 最终确定Top 20 Hub基因

【完整分析流程】
步骤1: 网络预处理 - 删除孤立节点 (已完成)
步骤2: K-core分解分析 - 提取g_k3子图 (已完成)
步骤3: 中心性指标计算 - R部分4种指标 (已完成)
步骤4: MCODE聚类分析 - 需在Cytoscape中执行
步骤5: cytoHubba中心性 - 计算剩余4种指标
步骤6: RRA综合评分 - 整合所有8种指标
"""

ws_guide['A1'] = "Cytoscape网络分析操作指南"
ws_guide['A1'].font = title_font

row = 3
for line in guide_text.strip().split('\n'):
    ws_guide.cell(row=row, column=1, value=line)
    if line.startswith('【') or line.startswith('步骤'):
        ws_guide.cell(row=row, column=1).font = subtitle_font
    row += 1

ws_guide.column_dimensions['A'].width = 80

# 保存文件
wb.save(output_file)
print(f"系统性生物信息学分析汇总Excel已生成: {output_file}")
print("包含以下sheet:")
for sheet in wb.sheetnames:
    print(f"  - {sheet}")
