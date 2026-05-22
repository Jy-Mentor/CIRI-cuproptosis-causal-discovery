import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import gzip
import tarfile
from scipy.io import mmread
import scanpy as sc
import anndata

SEED = 42
np.random.seed(SEED)

OUTPUT_DIR = r"C:\Users\Jy-Mentor-7\Desktop\你信我 这不是重蹈覆辙\CIRI-cuproptosis-causal-discovery\results\L1_phenotype_anchoring"

BULK_SIG_GENES = {
    'Cdkn2a': '上调', 'Cox17': '下调', 'Dbt': '上调', 'Dld': '下调',
    'Gls': '下调', 'Nfe2l2': '上调', 'Pdha1': '下调', 'Pdhb': '下调',
    'Slc31a1': '上调', 'Slc31a2': '上调', 'Atox1': '上调', 'Mt2': '上调', 'Cp': '上调'
}
BULK_NSIG_GENES = {
    'Atp7a': '下调', 'Atp7b': '下调', 'Lias': '下调', 'Fdx1': '上调',
    'Mtf1': '下调', 'Lipt2': '上调', 'Dlst': '下调', 'Gcsh': '下调',
    'Cox11': '下调', 'Sco1': '上调', 'Alb': '下调', 'Sod1': '下调',
    'Sod3': '下调', 'Commd1': '上调', 'Slc11a2': '下调', 'Steap3': '上调'
}
ALL_CHECK_GENES = list(BULK_SIG_GENES.keys()) + list(BULK_NSIG_GENES.keys())

CUPROPTOSIS_CORE = ['Fdx1', 'Lias', 'Lipt1', 'Lipt2', 'Dld', 'Dlat', 'Dlst', 'Pdha1', 'Pdhb', 'Dbt',
                     'Mtf1', 'Nfe2l2', 'Nlrp3', 'Gls', 'Cdkn2a', 'Cox17', 'Atp7a', 'Atp7b',
                     'Slc31a1', 'Gcsh']
CUPROPTOSIS_CU_HOMEOSTASIS = ['Slc31a2', 'Slc11a2', 'Steap3', 'Atox1', 'Ccs', 'Cox11',
                               'Sco1', 'Sco2', 'Mt1', 'Mt2', 'Alb', 'Cp', 'Sod1', 'Sod3', 'Commd1']

SAMPLES = [
    {"id": "sham1", "condition": "Sham", "files": ["GSM5319987_sham1"]},
    {"id": "sham2", "condition": "Sham", "files": ["GSM5319988_sham2"]},
    {"id": "sham3", "condition": "Sham", "files": ["GSM5319989_sham3"]},
    {"id": "mcao1", "condition": "MCAO", "files": ["GSM5319990_MCAO1"]},
    {"id": "mcao2", "condition": "MCAO", "files": ["GSM5319991_MCAO2"]},
    {"id": "mcao3", "condition": "MCAO", "files": ["GSM5319992_MCAO3"]},
]
MARKERS = {
    "Microglia": ["Ptprc", "Aif1", "Cx3cr1", "Tmem119", "P2ry12", "C1qa", "C1qb"],
    "Neuron": ["Snap25", "Syt1", "Nefl", "Rbfox3", "Syn1"],
    "Astrocyte": ["Gfap", "Aqp4", "Slc1a3", "Aldh1l1"],
    "Endothelial": ["Pecam1", "Vwf", "Cldn5", "Cdh5"],
    "Oligodendrocyte": ["Mbp", "Plp1", "Mog", "Mag"],
    "OPC": ["Pdgfra", "Vcan", "Cspg4", "Olig1", "Olig2"],
}
DATA_DIR = r"D:\反向网络药理学\L1 数据集\RNA-seq"
RAW_TAR = os.path.join(DATA_DIR, "GSE174574_RAW.tar")
EXTRACT_DIR = os.path.join(DATA_DIR, "GSE174574_extracted")

header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill('solid', fgColor='4472C4')
title_font = Font(name='Arial', bold=True, size=14, color='1F3864')
subtitle_font = Font(name='Arial', bold=True, size=12, color='2E75B6')
sig_fill = PatternFill('solid', fgColor='C6EFCE')
nsig_fill = PatternFill('solid', fgColor='FFC7CE')
warn_fill = PatternFill('solid', fgColor='FFEB9C')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def apply_header_style(ws, row=1, max_col=10):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

def apply_data_borders(ws, max_row, max_col):
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = thin_border

def apply_auto_width(ws, max_col, min_width=10, max_width=35):
    for col in range(1, max_col + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = max(min_width, min(max_width, max_len + 4))

# ============================================================
# BULK EXCEL
# ============================================================
bulk_degs = pd.read_csv(os.path.join(OUTPUT_DIR, "GSE61616_GEO2R_DEGs.csv"))
bulk_cupro = pd.read_csv(os.path.join(OUTPUT_DIR, "GSE61616_cuproptosis_genes.csv"))

wb_bulk = Workbook()

# Sheet 1: 所有差异基因
ws1 = wb_bulk.active
ws1.title = "All_DEGs"
ws1['A1'] = 'GSE61616 Bulk RNA-seq 全部差异基因结果 (GEO2R标准流程)'
ws1['A1'].font = title_font
ws1.merge_cells('A1:H1')

headers = ['Gene Symbol', 'log2FC', 'AveExpr', 't', 'P.Value', 'adj.P.Val', 'B', '显著性 (adj.P<0.05)']
for i, h in enumerate(headers, 1):
    ws1.cell(row=3, column=i, value=h)
apply_header_style(ws1, row=3, max_col=8)

for idx, row_data in bulk_degs.iterrows():
    r = idx + 4
    ws1.cell(row=r, column=1, value=bulk_degs.index[idx])
    ws1.cell(row=r, column=2, value=round(row_data['logFC'], 4))
    ws1.cell(row=r, column=3, value=round(row_data['AveExpr'], 4))
    ws1.cell(row=r, column=4, value=round(row_data['t'], 4))
    ws1.cell(row=r, column=5, value=row_data['P.Value'])
    ws1.cell(row=r, column=6, value=row_data['adj.P.Val'])
    ws1.cell(row=r, column=7, value=round(row_data['B'], 4))
    sig = "显著" if row_data['adj.P.Val'] < 0.05 else "不显著"
    ws1.cell(row=r, column=8, value=sig)
    if row_data['adj.P.Val'] < 0.05:
        for c in range(1, 9):
            ws1.cell(row=r, column=c).fill = sig_fill
    else:
        for c in range(1, 9):
            ws1.cell(row=r, column=c).fill = nsig_fill

apply_data_borders(ws1, len(bulk_degs) + 3, 8)
apply_auto_width(ws1, 8)

# Sheet 2: 铜死亡/铜稳态基因汇总
ws2 = wb_bulk.create_sheet("Cuproptosis_Genes")
ws2['A1'] = 'GSE61616 铜死亡+铜稳态基因验证结果 (35基因)'
ws2['A1'].font = title_font
ws2.merge_cells('A1:K1')

headers2 = ['基因类别', 'Gene Symbol', 'log2FC', 'P.Value', 'adj.P.Val', '方向',
            '显著 (adj.P<0.05)', '检测状态', 'scRNA-seq 细胞特异性', '方向一致性', '备注']
for i, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=i, value=h)
apply_header_style(ws2, row=3, max_col=11)

r = 4
for gene in CUPROPTOSIS_CORE + CUPROPTOSIS_CU_HOMEOSTASIS:
    cat = "铜死亡核心" if gene in CUPROPTOSIS_CORE else "铜稳态"
    matched = bulk_cupro[bulk_cupro['Gene'].str.lower() == gene.lower()]
    if len(matched) > 0:
        m = matched.iloc[0]
        ws2.cell(row=r, column=1, value=cat)
        ws2.cell(row=r, column=2, value=gene)
        ws2.cell(row=r, column=3, value=round(m['logFC'], 4))
        ws2.cell(row=r, column=4, value=m['P.Value'])
        ws2.cell(row=r, column=5, value=m['adj.P.Val'])
        ws2.cell(row=r, column=6, value=m['Direction'])
        ws2.cell(row=r, column=7, value="显著" if m['adj.P.Val'] < 0.05 else "不显著")
        ws2.cell(row=r, column=8, value="检出")
        if m['adj.P.Val'] < 0.05:
            for c in range(1, 12):
                ws2.cell(row=r, column=c).fill = sig_fill
        else:
            for c in range(1, 12):
                ws2.cell(row=r, column=c).fill = nsig_fill
    else:
        ws2.cell(row=r, column=1, value=cat)
        ws2.cell(row=r, column=2, value=gene)
        ws2.cell(row=r, column=8, value="未检出")
        for c in range(1, 12):
            ws2.cell(row=r, column=c).fill = warn_fill
    r += 1

apply_data_borders(ws2, r - 1, 11)
apply_auto_width(ws2, 11)

# Sheet 3: 统计摘要
ws3 = wb_bulk.create_sheet("Summary")
ws3['A1'] = 'GSE61616 分析统计摘要'
ws3['A1'].font = title_font
ws3.merge_cells('A1:B1')
stats = [
    ("数据集", "GSE61616 (GPL1355 Rat Affymetrix)"),
    ("样本数", "15 (Sham=5, IR=10)"),
    ("芯片平台", "GPL1355-10794"),
    ("探针总数", "31,099"),
    ("映射后基因数", "15,248"),
    ("差异基因总数 (P<0.05)", str(len(bulk_degs[bulk_degs['adj.P.Val'] < 0.05]))),
    ("铜死亡核心基因", "20"),
    ("铜稳态基因", "15"),
    ("铜死亡显著基因 (adj.P<0.05)", str(len(bulk_cupro[bulk_cupro['adj.P.Val'] < 0.05]))),
    ("铜死亡检出率", f"{len(bulk_cupro)}/35 ({len(bulk_cupro)/35*100:.1f}%)"),
    ("极显著基因 (adj.P<0.001)", str(len(bulk_cupro[bulk_cupro['adj.P.Val'] < 0.001]))),
    ("分析标准", "GEO2R (GEOquery + limma)"),
    ("探针映射策略", "最高表达探针"),
]
for i, (k, v) in enumerate(stats, 3):
    ws3.cell(row=i, column=1, value=k).font = Font(bold=True, name='Arial')
    ws3.cell(row=i, column=2, value=v)

ws3.column_dimensions['A'].width = 35
ws3.column_dimensions['B'].width = 50

bulk_path = os.path.join(OUTPUT_DIR, "L1_Bulk_GSE61616_Summary.xlsx")
wb_bulk.save(bulk_path)
print(f"Bulk Excel saved: {bulk_path}")

# ============================================================
# scRNA-seq EXCEL
# ============================================================
sc_all = pd.read_csv(os.path.join(OUTPUT_DIR, "GSE174574_all_DEGs.csv"))
sc_cupro = pd.read_csv(os.path.join(OUTPUT_DIR, "GSE174574_cuproptosis_DEGs.csv"))
sc_sig = pd.read_csv(os.path.join(OUTPUT_DIR, "GSE174574_sig_cuproptosis_DEGs.csv"))

# Re-run scRNA-seq for cross-validation
os.makedirs(EXTRACT_DIR, exist_ok=True)
with tarfile.open(RAW_TAR, "r") as tar:
    tar.extractall(EXTRACT_DIR)

adatas = []
for sample in SAMPLES:
    prefix = sample["files"][0]
    matrix_path = os.path.join(EXTRACT_DIR, f"{prefix}_matrix.mtx.gz")
    genes_path = os.path.join(EXTRACT_DIR, f"{prefix}_genes.tsv.gz")
    barcodes_path = os.path.join(EXTRACT_DIR, f"{prefix}_barcodes.tsv.gz")
    if all(os.path.exists(p) for p in [matrix_path, genes_path, barcodes_path]):
        with gzip.open(genes_path, 'rt') as f:
            genes = [line.strip().split('\t')[1] for line in f]
        with gzip.open(barcodes_path, 'rt') as f:
            barcodes = [line.strip() for line in f]
        with gzip.open(matrix_path, 'rb') as f:
            matrix = mmread(f).tocsr()
        adata = sc.AnnData(matrix.T)
        adata.var_names = genes
        adata.obs_names = [f"{bc}_{prefix}" for bc in barcodes]
        adata.obs["sample_id"] = sample["id"]
        adata.obs["condition"] = sample["condition"]
        adatas.append(adata)

for adata in adatas:
    adata.obs_names_make_unique()
    adata.var_names_make_unique()

adata = anndata.concat(adatas, axis=0, join="outer", label="batch", keys=[s["id"] for s in SAMPLES])
adata.obs_names_make_unique()
adata.var["mt"] = adata.var_names.str.startswith("mt-")
sc.pp.calculate_qc_metrics(adata, qc_vars=["mt"], inplace=True)
sc.pp.filter_cells(adata, min_genes=200)
sc.pp.filter_genes(adata, min_cells=3)
adata = adata[adata.obs.n_genes_by_counts < 5000, :]
adata = adata[adata.obs.pct_counts_mt < 20, :]
adata = adata[adata.obs.total_counts > 500, :]
sc.pp.normalize_total(adata, target_sum=1e4)
sc.pp.log1p(adata)
adata.raw = adata

for ctype, genes in MARKERS.items():
    present = [g for g in genes if g in adata.var_names]
    if present:
        sc.tl.score_genes(adata, gene_list=present, score_name=f"score_{ctype}", use_raw=False)

adata.obs["cell_type"] = "Unknown"
score_cols = [f"score_{ct}" for ct in MARKERS if f"score_{ct}" in adata.obs.columns]
if score_cols:
    best_scores = adata.obs[score_cols].idxmax(axis=1)
    adata.obs["cell_type"] = best_scores.str.replace("score_", "")

# Cross-validation
cross_results = []
present_genes = [g for g in ALL_CHECK_GENES if g in adata.var_names]
for ctype in adata.obs["cell_type"].unique():
    if ctype == "Unknown":
        continue
    subset = adata[adata.obs["cell_type"] == ctype]
    if subset.n_obs < 100:
        continue
    sc.tl.rank_genes_groups(subset, groupby="condition", reference="Sham", method="wilcoxon", use_raw=True)
    de_df = sc.get.rank_genes_groups_df(subset, group="MCAO")
    for gene in present_genes:
        row = de_df[de_df["names"] == gene]
        if len(row) > 0:
            row = row.iloc[0]
            log2fc = row["logfoldchanges"]
            pval = row["pvals_adj"]
            sig = pval < 0.05
            direction = "上调" if log2fc > 0 else "下调"
            bulk_dir = BULK_SIG_GENES.get(gene, BULK_NSIG_GENES.get(gene, "?"))
            cross_results.append({
                "Gene": gene, "CellType": ctype, "nCells": subset.n_obs,
                "log2FC": round(log2fc, 4), "adj.P": pval,
                "scRNA_Sig": "显著" if sig else "不显著",
                "scRNA_Dir": direction,
                "Bulk_Dir": bulk_dir,
                "Bulk_Sig": "显著" if gene in BULK_SIG_GENES else "不显著",
                "Consistent": "一致" if (sig and direction == bulk_dir) else ("N/A" if not sig else "矛盾")
            })

cross_df = pd.DataFrame(cross_results)
cross_df.to_csv(os.path.join(OUTPUT_DIR, "cross_validation_scRNA_bulk.csv"), index=False)

wb_sc = Workbook()

# Sheet 1: 所有差异基因
ws1 = wb_sc.active
ws1.title = "All_DEGs"
ws1['A1'] = 'GSE174574 scRNA-seq 全部差异基因 (Wilcoxon, MCAO vs Sham)'
ws1['A1'].font = title_font
ws1.merge_cells('A1:G1')
headers = ['Gene', 'log2FC', 'P.Value', 'adj.P.Val', '显著性', '上调/下调']
for i, h in enumerate(headers, 1):
    ws1.cell(row=3, column=i, value=h)
apply_header_style(ws1, row=3, max_col=6)
for idx, row_data in sc_all.head(5000).iterrows():
    r = idx + 4
    ws1.cell(row=r, column=1, value=row_data['names'])
    ws1.cell(row=r, column=2, value=round(row_data['logfoldchanges'], 4))
    ws1.cell(row=r, column=3, value=row_data['pvals'])
    ws1.cell(row=r, column=4, value=row_data['pvals_adj'])
    sig = "显著" if row_data['pvals_adj'] < 0.05 else "不显著"
    ws1.cell(row=r, column=5, value=sig)
    ws1.cell(row=r, column=6, value="上调" if row_data['logfoldchanges'] > 0 else "下调")
    if row_data['pvals_adj'] < 0.05:
        for c in range(1, 7):
            ws1.cell(row=r, column=c).fill = sig_fill
    else:
        for c in range(1, 7):
            ws1.cell(row=r, column=c).fill = nsig_fill
apply_data_borders(ws1, min(len(sc_all), 5000) + 3, 6)
apply_auto_width(ws1, 6)

# Sheet 2: 铜死亡基因
ws2 = wb_sc.create_sheet("Cuproptosis_Genes")
ws2['A1'] = 'GSE174574 铜死亡+铜稳态基因差异表达 (35基因)'
ws2['A1'].font = title_font
ws2.merge_cells('A1:J1')
headers2 = ['类别', 'Gene', 'log2FC', 'P.Value', 'adj.P.Val', '显著性', '上调/下调', 'Bulk方向', 'scRNA vs Bulk']
for i, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=i, value=h)
apply_header_style(ws2, row=3, max_col=9)
all_genes = CUPROPTOSIS_CORE + CUPROPTOSIS_CU_HOMEOSTASIS
for i, gene in enumerate(all_genes):
    r_idx = i + 4
    cat = "铜死亡核心" if gene in CUPROPTOSIS_CORE else "铜稳态"
    matched = sc_cupro[sc_cupro['names'].str.lower() == gene.lower()]
    bulk_dir = BULK_SIG_GENES.get(gene, BULK_NSIG_GENES.get(gene, "未检出"))
    if len(matched) > 0:
        m = matched.iloc[0]
        ws2.cell(row=r_idx, column=1, value=cat)
        ws2.cell(row=r_idx, column=2, value=gene)
        ws2.cell(row=r_idx, column=3, value=round(m['logfoldchanges'], 4))
        ws2.cell(row=r_idx, column=4, value=m['pvals'])
        ws2.cell(row=r_idx, column=5, value=m['pvals_adj'])
        sig = "显著" if m['pvals_adj'] < 0.05 else "不显著"
        ws2.cell(row=r_idx, column=6, value=sig)
        ws2.cell(row=r_idx, column=7, value="上调" if m['logfoldchanges'] > 0 else "下调")
        ws2.cell(row=r_idx, column=8, value=bulk_dir)
        consistency = "一致" if (m['pvals_adj'] < 0.05 and ((m['logfoldchanges'] > 0 and bulk_dir == '上调') or (m['logfoldchanges'] < 0 and bulk_dir == '下调'))) else "N/A"
        ws2.cell(row=r_idx, column=9, value=consistency)
        if m['pvals_adj'] < 0.05:
            for c in range(1, 10):
                ws2.cell(row=r_idx, column=c).fill = sig_fill
        else:
            for c in range(1, 10):
                ws2.cell(row=r_idx, column=c).fill = nsig_fill
    else:
        ws2.cell(row=r_idx, column=1, value=cat)
        ws2.cell(row=r_idx, column=2, value=gene)
        ws2.cell(row=r_idx, column=6, value="未检出")
        ws2.cell(row=r_idx, column=8, value=bulk_dir)
        for c in range(1, 10):
            ws2.cell(row=r_idx, column=c).fill = warn_fill

apply_data_borders(ws2, len(all_genes) + 3, 9)
apply_auto_width(ws2, 9)

# Sheet 3: 交叉验证
ws3 = wb_sc.create_sheet("Cross_Validation")
ws3['A1'] = 'scRNA-seq × Bulk 交叉验证 (细胞类型特异性)'
ws3['A1'].font = title_font
ws3.merge_cells('A1:J1')
cv_headers = ['Gene', 'Cell Type', 'n Cells', 'log2FC', 'adj.P', 'scRNA Sig', 'scRNA Dir', 'Bulk Dir', 'Bulk Sig', 'Consistency']
for i, h in enumerate(cv_headers, 1):
    ws3.cell(row=3, column=i, value=h)
apply_header_style(ws3, row=3, max_col=10)
for idx, row_data in cross_df.iterrows():
    r = idx + 4
    for c, col in enumerate(['Gene', 'CellType', 'nCells', 'log2FC', 'adj.P', 'scRNA_Sig', 'scRNA_Dir', 'Bulk_Dir', 'Bulk_Sig', 'Consistent'], 1):
        ws3.cell(row=r, column=c, value=row_data[col])
    if row_data['Consistent'] == '一致':
        for c in range(1, 11):
            ws3.cell(row=r, column=c).fill = sig_fill
    elif row_data['Consistent'] == '矛盾':
        for c in range(1, 11):
            ws3.cell(row=r, column=c).fill = nsig_fill

apply_data_borders(ws3, len(cross_df) + 3, 10)
apply_auto_width(ws3, 10)

# Sheet 4: 统计摘要
ws4 = wb_sc.create_sheet("Summary")
ws4['A1'] = 'GSE174574 scRNA-seq 分析统计摘要'
ws4['A1'].font = title_font
ws4.merge_cells('A1:B1')
cell_counts = adata.obs['cell_type'].value_counts()
stats = [
    ("数据集", "GSE174574 (24h MCAO vs Sham, scRNA-seq)"),
    ("总细胞数", f"{adata.n_obs:,}"),
    ("总基因数", f"{adata.n_vars:,}"),
    ("样本数", "6 (Sham=3, MCAO=3)"),
    ("质控标准", "min_genes=200, min_cells=3, mt<20%, n_genes<5000"),
    ("差异分析方法", "Wilcoxon (Scanpy)"),
    ("差异阈值", "|log2FC|>0.25, adj.P<0.05"),
    ("总差异基因", str(len(sc_all[sc_all['pvals_adj'] < 0.05]))),
    ("铜死亡核心基因", "20"),
    ("铜稳态基因", "15"),
    ("scRNA-seq 铜死亡差异基因", str(len(sc_sig))),
    ("", ""),
    ("细胞类型分布", ""),
]
for i, (k, v) in enumerate(stats, 3):
    ws4.cell(row=i, column=1, value=k).font = Font(bold=True, name='Arial')
    ws4.cell(row=i, column=2, value=v)

r = len(stats) + 3
for ct, count in cell_counts.items():
    ws4.cell(row=r, column=1, value=ct).font = Font(name='Arial')
    ws4.cell(row=r, column=2, value=count)
    ws4.cell(row=r, column=3, value=f"{count/adata.n_obs*100:.1f}%")
    r += 1

ws4.column_dimensions['A'].width = 35
ws4.column_dimensions['B'].width = 50
ws4.column_dimensions['C'].width = 15

sc_path = os.path.join(OUTPUT_DIR, "L1_scRNA_GSE174574_Summary.xlsx")
wb_sc.save(sc_path)
print(f"scRNA-seq Excel saved: {sc_path}")
print("All Excel files generated successfully.")
