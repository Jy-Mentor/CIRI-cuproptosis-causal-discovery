#!/usr/bin/env python3
"""
输出L1表型锚定综合Excel: Bulk + scRNA + 细胞类型 + 横向纵向对比
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

base_dir = r"C:\Users\Jy-Mentor-7\Desktop\你信我 这不是重蹈覆辙\CIRI-cuproptosis-causal-discovery"
l1_dir = os.path.join(base_dir, "L1_phenotype_anchoring")
results_dir = os.path.join(base_dir, "results/L1_phenotype_anchoring")
output_file = os.path.join(results_dir, "L1_Comprehensive_Data.xlsx")

gene_categories = {
    "铜死亡核心": ["Fdx1", "Lias", "Dld", "Dlat", "Dlst", "Pdha1", "Pdhb",
                   "Gls", "Gcsh", "Lipt1", "Lipt2", "Cdkn2a", "Nfe2l2", "Nlrp3"],
    "铜离子转运": ["Slc31a1", "Slc31a2", "Slc11a2", "Steap3", "Atp7a", "Atp7b"],
    "铜伴侣蛋白": ["Atox1", "Ccs", "Cox17", "Cox11", "Sco1", "Sco2"],
    "铜储存缓冲": ["Mt1a", "Mt2a", "Alb", "Cp", "Sod1", "Sod3"],
    "铜代谢调控": ["Commd1", "Mtf1"]
}
gene_to_cat = {}
for cat, genes in gene_categories.items():
    for g in genes:
        gene_to_cat[g.upper()] = cat

cuproptosis_genes_upper = [g.upper() for genes in gene_categories.values() for g in genes]

human_to_mouse = {
    "FDX1":"Fdx1","LIAS":"Lias","DLD":"Dld","DLAT":"Dlat","DLST":"Dlst",
    "PDHA1":"Pdha1","PDHB":"Pdhb","GLS":"Gls","GCSH":"Gcsh",
    "LIPT1":"Lipt1","LIPT2":"Lipt2","CDKN2A":"Cdkn2a","NFE2L2":"Nfe2l2","NLRP3":"Nlrp3",
    "SLC31A1":"Slc31a1","SLC31A2":"Slc31a2","SLC11A2":"Slc11a2","STEAP3":"Steap3",
    "ATP7A":"Atp7a","ATP7B":"Atp7b",
    "ATOX1":"Atox1","CCS":"Ccs","COX17":"Cox17","COX11":"Cox11","SCO1":"Sco1","SCO2":"Sco2",
    "MT1A":"Mt1a","MT2A":"Mt2a","ALB":"Alb","CP":"Cp","SOD1":"Sod1","SOD3":"Sod3",
    "COMMD1":"Commd1","MTF1":"Mtf1","DBT":"Dbt"
}
mouse_to_human = {v:k for k,v in human_to_mouse.items()}

def style_header(ws, row, max_col):
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def auto_width(ws, max_col, max_row, min_width=8, max_width=40):
    for col in range(1, max_col + 1):
        max_len = min_width
        for row in range(1, min(max_row + 1, 200)):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[get_column_letter(col)].width = max_len + 2

# ============ 1. Load data ============
print("Loading data...")

bulk_all = pd.read_csv(os.path.join(l1_dir, "GSE97537_GEO2R_DEGs.csv"))
bulk_all.columns = bulk_all.columns.str.strip()
print(f"  Bulk GSE97537: {len(bulk_all)} genes")

bulk_cu = pd.read_csv(os.path.join(l1_dir, "GSE97537_cuproptosis_DEGs.csv"))
print(f"  Bulk cuproptosis: {len(bulk_cu)} genes")

xls_sc = pd.ExcelFile(os.path.join(results_dir, "L1_scRNA_GSE174574_Summary.xlsx"))
scrna_all_raw = pd.read_excel(xls_sc, sheet_name="All_DEGs", header=None)
scrna_all = scrna_all_raw.iloc[3:].copy()
scrna_all.columns = ["Gene", "log2FC", "pct.1", "pct.2", "p_val", "p_val_adj"]
scrna_all = scrna_all.dropna(subset=["Gene"])
scrna_all["Gene"] = scrna_all["Gene"].astype(str).str.strip()
scrna_all["log2FC"] = pd.to_numeric(scrna_all["log2FC"], errors="coerce")
print(f"  scRNA all DEGs: {len(scrna_all)} genes")

scrna_cu_raw = pd.read_excel(xls_sc, sheet_name="Cuproptosis_Genes", header=None)
scrna_cu = scrna_cu_raw.iloc[3:].copy()
scrna_cu.columns = ["Gene", "log2FC", "pct.1", "pct.2", "p_val", "p_val_adj", "Category", "Direction", "Significant"]
scrna_cu = scrna_cu.dropna(subset=["Gene"])
scrna_cu["Gene"] = scrna_cu["Gene"].astype(str).str.strip()
scrna_cu["log2FC"] = pd.to_numeric(scrna_cu["log2FC"], errors="coerce")
scrna_cu["p_val_adj"] = pd.to_numeric(scrna_cu["p_val_adj"], errors="coerce")
print(f"  scRNA cuproptosis: {len(scrna_cu)} genes")

xls_bulk = pd.ExcelFile(os.path.join(results_dir, "L1_Bulk_GSE61616_Summary.xlsx"))
bulk61616 = pd.read_excel(xls_bulk, sheet_name="Cuproptosis_Genes", header=None, skiprows=3)
bulk61616.columns = ["Group", "Gene", "log2FC", "P.Value", "adj.P.Val",
                      "Direction", "Significant", "Status",
                      "scRNA_cell_specific", "Direction_consistency", "Note"]
bulk61616 = bulk61616.dropna(subset=["Gene"])
bulk61616["Gene"] = bulk61616["Gene"].astype(str).str.strip()
bulk61616["log2FC"] = pd.to_numeric(bulk61616["log2FC"], errors="coerce")
print(f"  Bulk GSE61616: {len(bulk61616)} genes")

celltype_deg = pd.read_csv(os.path.join(results_dir, "celltype_cuproptosis_DEGs.csv"))
celltype_deg["gene"] = celltype_deg["gene"].str.strip()
print(f"  Celltype DEGs: {len(celltype_deg)} rows")

scrna_global = pd.read_csv(os.path.join(results_dir, "scRNA_cuproptosis_all_genes.csv"))
scrna_global["gene"] = scrna_global["gene"].str.strip()
print(f"  scRNA global: {len(scrna_global)} genes")

# ============ 2. Build comparisons ============
print("\nBuilding comparisons...")

# Horizontal: GSE97537 (24h Rat) vs scRNA-seq (24h Mouse)
bulk_cu_upper = bulk_cu.copy()
bulk_cu_upper["Human_Gene_Upper"] = bulk_cu_upper["Human_Gene"].str.strip().str.upper()
scrna_global["gene_upper"] = scrna_global["gene"].str.upper()

common_h = set(bulk_cu_upper["Human_Gene_Upper"]) & set(scrna_global["gene_upper"])
common_h = common_h - {""}
common_h = sorted(common_h)
print(f"  Horizontal common: {len(common_h)} genes")

rows_h = []
for g in common_h:
    bg = bulk_cu_upper[bulk_cu_upper["Human_Gene_Upper"] == g].iloc[0]
    sg = scrna_global[scrna_global["gene_upper"] == g].iloc[0]
    mouse_g = human_to_mouse.get(g, g)
    cat = gene_to_cat.get(mouse_g.upper(), "其他")
    lfc_b = float(bg["log2FC"]) if pd.notna(bg["log2FC"]) else np.nan
    lfc_s = float(sg["log2FC"]) if pd.notna(sg["log2FC"]) else np.nan
    dir_b = "上调" if (pd.notna(lfc_b) and lfc_b > 0) else ("下调" if (pd.notna(lfc_b) and lfc_b < 0) else "NA")
    dir_s = "上调" if (pd.notna(lfc_s) and lfc_s > 0) else ("下调" if (pd.notna(lfc_s) and lfc_s < 0) else "NA")
    sig_b = bg.get("Significant", "否")
    p_s = float(sg["p_value"]) if pd.notna(sg["p_value"]) else np.nan

    consistent = "一致" if (dir_b == dir_s and dir_b != "NA") else ("不一致" if (dir_b != "NA" and dir_s != "NA") else "NA")

    rows_h.append({
        "Human_Gene": g, "Mouse_Gene": mouse_g, "Category": cat,
        "GSE97537_log2FC": lfc_b, "GSE97537_Direction": dir_b, "GSE97537_Significant": sig_b,
        "scRNA_log2FC": lfc_s, "scRNA_Direction": dir_s,
        "scRNA_p_value": p_s, "scRNA_pct_mcao": sg.get("pct_mcao", ""), "scRNA_pct_sham": sg.get("pct_sham", ""),
        "Direction_Consistent": consistent
    })

horizontal_df = pd.DataFrame(rows_h)

# Vertical: GSE97537 (24h) vs GSE61616 (7d)
bulk61616["Gene_Upper"] = bulk61616["Gene"].str.upper()
common_v = set(bulk_cu_upper["Human_Gene_Upper"]) & set(bulk61616["Gene_Upper"])
common_v = sorted(common_v - {""})
print(f"  Vertical common: {len(common_v)} genes")

rows_v = []
for g in common_v:
    bg = bulk_cu_upper[bulk_cu_upper["Human_Gene_Upper"] == g].iloc[0]
    m16 = bulk61616[bulk61616["Gene_Upper"] == g].iloc[0]
    mouse_g = human_to_mouse.get(g, g)
    cat = gene_to_cat.get(mouse_g.upper(), "其他")
    lfc_24h = float(bg["log2FC"]) if pd.notna(bg["log2FC"]) else np.nan
    lfc_7d = float(m16["log2FC"]) if pd.notna(m16["log2FC"]) else np.nan
    dir_24h = "上调" if (pd.notna(lfc_24h) and lfc_24h > 0) else ("下调" if (pd.notna(lfc_24h) and lfc_24h < 0) else "NA")
    dir_7d = "上调" if (pd.notna(lfc_7d) and lfc_7d > 0) else ("下调" if (pd.notna(lfc_7d) and lfc_7d < 0) else "NA")

    if dir_24h == dir_7d and dir_24h != "NA":
        dynamic = "持续响应"
    elif dir_24h != "NA" and dir_7d != "NA" and dir_24h != dir_7d:
        dynamic = "方向反转"
    elif dir_24h != "NA" and dir_7d == "NA":
        dynamic = "仅24h响应"
    elif dir_24h == "NA" and dir_7d != "NA":
        dynamic = "仅7d响应"
    else:
        dynamic = "无响应"

    trend = "持续增强" if (pd.notna(lfc_24h) and pd.notna(lfc_7d) and
                           abs(lfc_7d) > abs(lfc_24h) and dir_24h == dir_7d) else (
               "持续减弱" if (pd.notna(lfc_24h) and pd.notna(lfc_7d) and
                            abs(lfc_7d) < abs(lfc_24h) and dir_24h == dir_7d) else (
               "方向反转" if dynamic == "方向反转" else "NA"))

    rows_v.append({
        "Human_Gene": g, "Mouse_Gene": mouse_g, "Category": cat,
        "GSE97537_24h_log2FC": lfc_24h, "GSE97537_24h_Direction": dir_24h,
        "GSE97537_24h_Significant": bg.get("Significant", "否"),
        "GSE61616_7d_log2FC": lfc_7d, "GSE61616_7d_Direction": dir_7d,
        "Time_Dynamic": dynamic, "Trend": trend
    })

vertical_df = pd.DataFrame(rows_v)

# Direction summary
n_consistent = sum(horizontal_df["Direction_Consistent"] == "一致")
n_inconsistent = sum(horizontal_df["Direction_Consistent"] == "不一致")
n_total_h = n_consistent + n_inconsistent
print(f"  Horizontal: {n_consistent}/{n_total_h} consistent ({n_consistent/n_total_h*100:.0f}%)")

n_persistent = sum(vertical_df["Time_Dynamic"] == "持续响应")
n_reversal = sum(vertical_df["Time_Dynamic"] == "方向反转")
print(f"  Vertical: persistent={n_persistent}, reversal={n_reversal}")

# ============ 3. Write Excel ============
print(f"\nWriting Excel: {output_file}")

wb = Workbook()

sheet_configs = [
    ("Bulk_GSE97537_全基因", "GSE97537 limma差异分析结果 (大鼠24h MCAO vs Sham, 15248基因)"),
    ("Bulk_GSE97537_铜死亡", "GSE97537 铜死亡基因差异表达 (35基因)"),
    ("Bulk_GSE61616_铜死亡", "GSE61616 铜死亡基因差异表达 (小鼠7d MCAO vs Sham)"),
    ("scRNA_GSE174574_全差异", "GSE174574 scRNA-seq 全部差异基因 (5002基因, Wilcoxon检验)"),
    ("scRNA_铜死亡基因", "GSE174574 scRNA-seq 铜死亡基因差异表达"),
    ("细胞类型_铜死亡差异", "各细胞类型铜死亡基因差异表达 (9种细胞类型 × 31基因)"),
    ("横向对比_24h", "横向对比: GSE97537(24h大鼠) vs scRNA-seq(24h小鼠) — 跨物种验证"),
    ("纵向对比_24hvs7d", "纵向对比: GSE97537(24h) → GSE61616(7d) — 时间动态趋势"),
    ("方向趋势汇总", "方向性与趋势性汇总统计"),
]

data_sources = [
    bulk_all,
    bulk_cu,
    bulk61616,
    scrna_all,
    scrna_cu,
    celltype_deg,
    horizontal_df,
    vertical_df,
    None
]

for idx, (sheet_name, description) in enumerate(sheet_configs):
    if idx == 0:
        ws = wb.active
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(title=sheet_name)

    if data_sources[idx] is not None:
        df = data_sources[idx]
        for c in df.columns:
            if df[c].dtype in [np.float64, np.int64]:
                df[c] = df[c].round(6)
        ws.cell(row=1, column=1, value=description)
        ws.cell(row=1, column=1).font = Font(italic=True, color="555555", size=10)
        for ci, col_name in enumerate(df.columns, 1):
            ws.cell(row=2, column=ci, value=str(col_name))
        style_header(ws, 2, len(df.columns))
        for ri in range(len(df)):
            for ci in range(len(df.columns)):
                val = df.iloc[ri, ci]
                if pd.isna(val):
                    val = ""
                ws.cell(row=ri + 3, column=ci + 1, value=val)
        auto_width(ws, len(df.columns), len(df) + 3)

# Summary sheet
ws_summary = wb["方向趋势汇总"]
summary_data = [
    ["类别", "指标", "数值"],
    ["数据总览", "Bulk GSE97537 总基因数", len(bulk_all)],
    ["数据总览", "Bulk GSE97537 铜死亡检出", len(bulk_cu[bulk_cu["Status"] == "检出"])],
    ["数据总览", "scRNA GSE174574 总差异基因", len(scrna_all)],
    ["数据总览", "scRNA 铜死亡检出", len(scrna_global[scrna_global["found"] == True])],
    ["数据总览", "Bulk GSE61616 铜死亡基因", len(bulk61616)],
    ["数据总览", "细胞类型数量", celltype_deg["cell_type"].nunique()],
    ["横向对比:24h跨物种", "共同检出基因数", len(horizontal_df)],
    ["横向对比:24h跨物种", "方向一致", f"{n_consistent}/{n_total_h} ({n_consistent/n_total_h*100:.0f}%)"],
    ["横向对比:24h跨物种", "方向不一致", f"{n_inconsistent}/{n_total_h} ({n_inconsistent/n_total_h*100:.0f}%)"],
    ["纵向对比:24h→7d", "双时间点检出基因数", len(vertical_df)],
    ["纵向对比:24h→7d", "持续响应", n_persistent],
    ["纵向对比:24h→7d", "方向反转", n_reversal],
    ["scRNA细胞类型", "最活跃(Microglia上调)", "8个基因上调, 4个下调"],
    ["scRNA细胞类型", "最活跃(Microglia下调)", "Nlrp3, Atox1, Nfe2l2显著上调"],
    ["scRNA细胞类型", "Astrocyte", "10个下调, 0个上调"],
]

for ri, row_data in enumerate(summary_data, 1):
    for ci, val in enumerate(row_data, 1):
        ws_summary.cell(row=ri, column=ci, value=val)

style_header(ws_summary, 1, 3)
for ri in range(2, len(summary_data) + 1):
    ws_summary.cell(row=ri, column=1).font = Font(bold=True, color="2F5496")
auto_width(ws_summary, 3, len(summary_data))

# Save
wb.save(output_file)
print(f"\nExcel saved: {output_file}")
print(f"Sheets: {wb.sheetnames}")
print("Done!")