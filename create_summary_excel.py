# -*- coding: utf-8 -*-
"""
创建v8.0结果汇总Excel
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 创建新的工作簿
wb = Workbook()

# 删除默认sheet
wb.remove(wb.active)

# Sheet 1: 项目总览
ws1 = wb.create_sheet('项目总览')
ws1['A1'] = '多组学整合分析揭示BCP治疗CIRI的潜在靶点 - v8.0无偏管线'
ws1['A1'].font = Font(bold=True, size=16, color='1F4E78')
ws1.merge_cells('A1:F1')

ws1['A3'] = '项目信息'
ws1['A3'].font = Font(bold=True, size=12)
ws1['A4'] = '版本'
ws1['B4'] = 'v8.0 (无偏泛靶点发现)'
ws1['A5'] = '日期'
ws1['B5'] = '2026-05-12'
ws1['A6'] = '核心创新'
ws1['B6'] = '删除铜死亡/BCP先验，纯数据驱动排名'
ws1['A7'] = '种子池大小'
ws1['B7'] = '2029基因'
ws1['A8'] = 'PPI网络'
ws1['B8'] = '9831节点, 66045边'

ws1['A10'] = 'Stage8权重 (v8无偏)'
ws1['A10'].font = Font(bold=True)
ws1['A11'] = 'W_GRN'
ws1['B11'] = '0.35 (GRN扰动)'
ws1['A12'] = 'W_ML'
ws1['B12'] = '0.131 (ML重要性，含样本量惩罚)'
ws1['A13'] = 'W_PPI'
ws1['B13'] = '0.30 (PPI拓扑)'
ws1['A14'] = 'W_BCP'
ws1['B14'] = '0 (删除先验)'
ws1['A15'] = 'W_CUPRO'
ws1['B15'] = '0 (删除先验)'

ws1.column_dimensions['A'].width = 18
ws1.column_dimensions['B'].width = 40

# Sheet 2: Tier1核心靶点
tier1 = pd.read_csv('results/stage8_final_targets/tier1_targets.csv')
ws2 = wb.create_sheet('Tier1核心靶点')
ws2['A1'] = 'Tier1核心靶点 (高置信度, 40基因)'
ws2['A1'].font = Font(bold=True, size=14, color='1F4E78')
ws2.merge_cells('A1:G1')

# 写入列名
for col_idx, col_name in enumerate(tier1.columns, 1):
    cell = ws2.cell(row=3, column=col_idx, value=col_name)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='4472C4')
    cell.alignment = Alignment(horizontal='center')

# 写入数据
for row_idx, row_data in enumerate(tier1.values, 4):
    for col_idx, value in enumerate(row_data, 1):
        ws2.cell(row=row_idx, column=col_idx, value=value)

# 设置列宽
ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 15
ws2.column_dimensions['D'].width = 15
ws2.column_dimensions['E'].width = 15
ws2.column_dimensions['F'].width = 15
ws2.column_dimensions['G'].width = 12

# Sheet 3: GAT排名
gat = pd.read_csv('results/stage9_ppi_gat/gat_gene_ranking.csv')
ws3 = wb.create_sheet('GAT排名')
ws3['A1'] = 'GAT节点回归排名 (Top 100)'
ws3['A1'].font = Font(bold=True, size=14, color='1F4E78')
ws3.merge_cells('A1:D1')

# 写入Top100
top100 = gat.head(100)
for col_idx, col_name in enumerate(top100.columns, 1):
    cell = ws3.cell(row=3, column=col_idx, value=col_name)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='70AD47')
    cell.alignment = Alignment(horizontal='center')

for row_idx, row_data in enumerate(top100.values, 4):
    for col_idx, value in enumerate(row_data, 1):
        ws3.cell(row=row_idx, column=col_idx, value=value)

ws3.column_dimensions['A'].width = 12
ws3.column_dimensions['B'].width = 15
ws3.column_dimensions['C'].width = 15
ws3.column_dimensions['D'].width = 12

# Sheet 4: GRN扰动评分
grn = pd.read_csv('results/stage6_graphsage_knockout/gene_perturbation_scores.csv')
ws4 = wb.create_sheet('GRN扰动评分')
ws4['A1'] = 'GRN虚拟敲除扰动评分'
ws4['A1'].font = Font(bold=True, size=14, color='1F4E78')
ws4.merge_cells('A1:E1')

for col_idx, col_name in enumerate(grn.columns, 1):
    cell = ws4.cell(row=3, column=col_idx, value=col_name)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='FFC000')
    cell.alignment = Alignment(horizontal='center')

for row_idx, row_data in enumerate(grn.values, 4):
    for col_idx, value in enumerate(row_data, 1):
        ws4.cell(row=row_idx, column=col_idx, value=value)

ws4.column_dimensions['A'].width = 12
ws4.column_dimensions['B'].width = 18
ws4.column_dimensions['C'].width = 15
ws4.column_dimensions['D'].width = 12
ws4.column_dimensions['E'].width = 12

# Sheet 5: ML SHAP重要性
ml = pd.read_csv('results/stage7_ml_shap/gene_shap_importance.csv')
ws5 = wb.create_sheet('ML_SHAP重要性')
ws5['A1'] = '机器学习SHAP重要性'
ws5['A1'].font = Font(bold=True, size=14, color='1F4E78')
ws5.merge_cells('A1:F1')

for col_idx, col_name in enumerate(ml.columns, 1):
    cell = ws5.cell(row=3, column=col_idx, value=col_name)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='5B9BD5')
    cell.alignment = Alignment(horizontal='center')

for row_idx, row_data in enumerate(ml.values, 4):
    for col_idx, value in enumerate(row_data, 1):
        ws5.cell(row=row_idx, column=col_idx, value=value)

ws5.column_dimensions['A'].width = 12
ws5.column_dimensions['B'].width = 18
ws5.column_dimensions['C'].width = 15
ws5.column_dimensions['D'].width = 15
ws5.column_dimensions['E'].width = 15
ws5.column_dimensions['F'].width = 15

# Sheet 6: 铜死亡基因排名验证
cupro_genes = ['FDX1', 'LIAS', 'LIPT1', 'DLAT', 'PDHA1', 'PDHB', 'MTF1', 'GLS', 'CDKN2A', 'SLC31A1', 'ATP7A', 'ATP7B', 'DLD', 'DBT', 'DLST', 'PDHA2', 'GCSH']
ws6 = wb.create_sheet('铜死亡基因验证')
ws6['A1'] = '铜死亡核心基因在无偏管线中的排名 (验证数据驱动性)'
ws6['A1'].font = Font(bold=True, size=14, color='1F4E78')
ws6.merge_cells('A1:E1')

# 表头
headers = ['基因', 'GAT排名', 'GAT得分', 'Stage8排名', 'Stage8得分']
for col_idx, header in enumerate(headers, 1):
    cell = ws6.cell(row=3, column=col_idx, value=header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='C55A11')
    cell.alignment = Alignment(horizontal='center')

# 获取铜死亡基因排名
cupro_data = []
for gene in cupro_genes:
    gat_row = gat[gat['Gene'] == gene]
    stage8_row = tier1[tier1['Gene'] == gene] if gene in tier1['Gene'].values else None
    
    gat_rank = int(gat_row['Rank'].values[0]) if len(gat_row) > 0 else 'N/A'
    gat_score = float(gat_row['GAT_score'].values[0]) if len(gat_row) > 0 else 'N/A'
    
    if stage8_row is not None and len(stage8_row) > 0:
        stage8_rank = int(list(tier1['Gene']).index(gene) + 1)
        stage8_score = float(stage8_row['Comprehensive'].values[0])
    else:
        stage8_rank = '>40'
        stage8_score = 'N/A'
    
    cupro_data.append([gene, gat_rank, gat_score, stage8_rank, stage8_score])

for row_idx, row_data in enumerate(cupro_data, 4):
    for col_idx, value in enumerate(row_data, 1):
        ws6.cell(row=row_idx, column=col_idx, value=value)

# 添加统计
cupro_gat = gat[gat['Gene'].isin(cupro_genes)]
avg_rank = cupro_gat['Rank'].mean()
top10_count = len(cupro_gat[cupro_gat['Rank'] <= 983])

ws6['A22'] = '统计摘要'
ws6['A22'].font = Font(bold=True)
ws6['A23'] = '平均GAT排名'
ws6['B23'] = f'{avg_rank:.0f}/9831'
ws6['A24'] = '前10%数量'
ws6['B24'] = f'{top10_count}/17'
ws6['A25'] = '结论'
ws6['B25'] = '铜死亡基因在当前CIRI数据中信号较弱，支持无偏管线'

ws6.column_dimensions['A'].width = 12
ws6.column_dimensions['B'].width = 15
ws6.column_dimensions['C'].width = 15
ws6.column_dimensions['D'].width = 15
ws6.column_dimensions['E'].width = 15

# 保存
output_path = 'CIRI和BCP靶点/结果汇总_v8.xlsx'
wb.save(output_path)
print(f'Excel汇总文件已保存: {output_path}')
