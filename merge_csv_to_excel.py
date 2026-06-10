"""将所有CSV文件汇总到一个Excel工作簿"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
import os

RESULTS_DIR = Path("results")
SUMMARY_DIR = Path("石竹烯_CIRI_预测结果汇总")

def style_worksheet(ws, df):
    """为工作表添加格式"""
    # 标题行格式
    header_fill = PatternFill('solid', fgColor='4472C4')
    header_font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 数据行格式
    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(vertical='center', wrap_text=True)
    
    # 边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 设置标题行
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 设置列宽
    for col_idx in range(1, len(df.columns) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 自动筛选
    ws.auto_filter.ref = ws.dimensions

def main():
    output_file = SUMMARY_DIR / "石竹烯_CIRI_预测结果汇总.xlsx"
    
    # 获取所有CSV文件
    csv_files = sorted(RESULTS_DIR.glob("*.csv"))
    
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认工作表
    
    for csv_file in csv_files:
        # 跳过重复的大文件
        if csv_file.name == "full_prediction_summary.csv":
            continue
        
        # 读取CSV（尝试多种编码）
        df = None
        for encoding in ['utf-8', 'utf-8-sig', 'gbk', 'gb2312', 'latin1']:
            try:
                df = pd.read_csv(csv_file, encoding=encoding)
                break
            except (UnicodeDecodeError, Exception):
                continue
        
        if df is None:
            print(f"⚠ 无法读取: {csv_file.name}，跳过")
            continue
        
        # 创建工作表名（去除.csv，限制31字符）
        sheet_name = csv_file.stem[:31]
        
        # 创建工作表
        ws = wb.create_sheet(sheet_name)
        
        # 写入数据
        for row_idx, row in enumerate(df.itertuples(index=False), 1):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 添加格式
        style_worksheet(ws, df)
    
    # 保存
    wb.save(output_file)
    print(f"汇总完成: {output_file}")
    print(f"工作表数: {len(wb.sheetnames)}")
    print("工作表列表:")
    for name in wb.sheetnames:
        print(f"  - {name}")

if __name__ == "__main__":
    main()
