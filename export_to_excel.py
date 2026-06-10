#!/usr/bin/env python3
"""
石竹烯-CIRI GAT 结果导出为 Excel
整合 Top 靶点、训练指标、混淆矩阵信息到一个工作簿
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def create_excel_report(output_path: str = "./石竹烯-CIRI_GAT结果汇总.xlsx"):
    """创建 Excel 汇总报告"""
    
    wb = Workbook()
    
    # ========== Sheet 1: Top-50 候选靶点 ==========
    ws1 = wb.active
    ws1.title = "Top50候选靶点"
    
    df50 = pd.read_csv("results/top_targets_50.csv")
    df50.columns = ["基因符号", "靶点概率", "排名", "度中心性", "PageRank", "聚类系数", "铜死亡基因", "炎症相关"]
    
    for r_idx, row in enumerate(dataframe_to_rows(df50, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            else:
                # 铜死亡基因高亮
                if c_idx == 7 and value == 1:
                    cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                # 炎症相关高亮
                if c_idx == 8 and value == 1:
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    # 调整列宽
    ws1.column_dimensions["A"].width = 15
    ws1.column_dimensions["B"].width = 12
    ws1.column_dimensions["C"].width = 8
    for col in ["D", "E", "F"]:
        ws1.column_dimensions[col].width = 15
    
    # ========== Sheet 2: Top-100 候选靶点 ==========
    ws2 = wb.create_sheet("Top100候选靶点")
    df100 = pd.read_csv("results/top_targets_100.csv")
    df100.columns = ["基因符号", "靶点概率", "排名", "度中心性", "PageRank", "聚类系数", "铜死亡基因", "炎症相关"]
    
    for r_idx, row in enumerate(dataframe_to_rows(df100, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            else:
                if c_idx == 7 and value == 1:
                    cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                if c_idx == 8 and value == 1:
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 8
    for col in ["D", "E", "F"]:
        ws2.column_dimensions[col].width = 15
    
    # ========== Sheet 3: Top-200 候选靶点 ==========
    ws3 = wb.create_sheet("Top200候选靶点")
    df200 = pd.read_csv("results/top_targets_200.csv")
    df200.columns = ["基因符号", "靶点概率", "排名", "度中心性", "PageRank", "聚类系数", "铜死亡基因", "炎症相关"]
    
    for r_idx, row in enumerate(dataframe_to_rows(df200, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            else:
                if c_idx == 7 and value == 1:
                    cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                if c_idx == 8 and value == 1:
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    ws3.column_dimensions["A"].width = 15
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 8
    for col in ["D", "E", "F"]:
        ws3.column_dimensions[col].width = 15
    
    # ========== Sheet 4: 全部未知节点预测 ==========
    ws4 = wb.create_sheet("全部未知节点预测")
    df_all = pd.read_csv("results/all_unknown_predictions.csv")
    df_all.columns = ["基因符号", "靶点概率", "排名", "度中心性", "PageRank", "聚类系数", "铜死亡基因", "炎症相关"]
    
    for r_idx, row in enumerate(dataframe_to_rows(df_all, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws4.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            else:
                if c_idx == 7 and value == 1:
                    cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                if c_idx == 8 and value == 1:
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    ws4.column_dimensions["A"].width = 15
    ws4.column_dimensions["B"].width = 12
    ws4.column_dimensions["C"].width = 8
    for col in ["D", "E", "F"]:
        ws4.column_dimensions[col].width = 15
    
    # ========== Sheet 5: 训练指标汇总 ==========
    ws5 = wb.create_sheet("训练指标")
    
    metrics_data = [
        ["指标", "数值"],
        ["最佳 Epoch", 38],
        ["验证集 F1", 0.6522],
        ["测试集 Accuracy", 0.6933],
        ["测试集 Precision", 0.4231],
        ["测试集 Recall", 0.5789],
        ["测试集 F1", 0.4889],
        ["测试集 ROC-AUC", 0.7500],
        ["测试集 PR-AUC", 0.4796],
        ["", ""],
        ["数据集统计", ""],
        ["基因池总数", 3918],
        ["阳性标签", 93],
        ["阴性标签", 278],
        ["未知标签", 3547],
        ["边数", 34557],
    ]
    
    for r_idx, row in enumerate(metrics_data, 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws5.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1 or row[0] == "数据集统计":
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            if c_idx == 2 and isinstance(value, float):
                cell.number_format = "0.0000"
    
    ws5.column_dimensions["A"].width = 20
    ws5.column_dimensions["B"].width = 15
    
    # ========== Sheet 6: 铜死亡基因专项 ==========
    ws6 = wb.create_sheet("铜死亡基因专项")
    
    cupro_df = df_all[df_all["铜死亡基因"] == 1].copy()
    cupro_df = cupro_df.sort_values("靶点概率", ascending=False)
    
    for r_idx, row in enumerate(dataframe_to_rows(cupro_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws6.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
    
    ws6.column_dimensions["A"].width = 15
    ws6.column_dimensions["B"].width = 12
    ws6.column_dimensions["C"].width = 8
    for col in ["D", "E", "F"]:
        ws6.column_dimensions[col].width = 15
    
    # 保存
    wb.save(output_path)
    print(f"Excel 报告已生成: {output_path}")
    print(f"包含工作表: {wb.sheetnames}")


if __name__ == "__main__":
    create_excel_report()
