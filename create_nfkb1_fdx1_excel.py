import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# 设置路径
work_dir = "C:/Users/Jy-Mentor-7/Desktop/你信我 这不是重蹈覆辙"
result_dir = os.path.join(work_dir, "String_Network_Systematic_Analysis")
output_file = os.path.join(work_dir, "NFKB1_FDX1_Pathway_Analysis.xlsx")

# 创建workbook
wb = Workbook()
wb.remove(wb.active)

# 定义样式
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
title_font = Font(bold=True, size=14, color="1F4E78")
subtitle_font = Font(bold=True, size=11)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ==================== Sheet 1: 分析概述 ====================
ws_summary = wb.create_sheet("分析概述")

summary_data = [
    ["NFKB1到FDX1最短路径分析报告"],
    [""],
    ["分析背景"],
    ["分析网络", "g_k3 (K>=3核心子网络)"],
    ["网络节点数", 124],
    ["网络边数", 1924],
    [""],
    ["最短路径统计"],
    ["源节点", "NFKB1 (铜死亡通路关键调控因子)"],
    ["目标节点", "FDX1 (铜死亡核心基因)"],
    ["最短路径数", 8],
    ["最短路径长度", "2跳 (3个节点)"],
    [""],
    ["桥接节点分析"],
    ["桥接节点总数", 2],
    ["HSPA5通路", "4条路径 (50%)"],
    ["HMOX1通路", "4条路径 (50%)"],
    [""],
    ["关键发现"],
    ["通路1", "NFKB1 → HSPA5 → FDX1"],
    ["通路2", "NFKB1 → HMOX1 → FDX1"],
    ["生物学意义", "NFKB1可能通过HSPA5或HMOX1间接调控FDX1"],
]

for row_idx, row_data in enumerate(summary_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.font = Font(bold=True, size=16, color="1F4E78")
            cell.alignment = Alignment(horizontal="center")
        elif value in ["分析背景", "最短路径统计", "桥接节点分析", "关键发现"]:
            cell.font = subtitle_font
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

ws_summary.column_dimensions['A'].width = 20
ws_summary.column_dimensions['B'].width = 45

# ==================== Sheet 2: 8条最短路径 ====================
ws_paths = wb.create_sheet("8条最短路径")

paths_df = pd.read_csv(os.path.join(result_dir, "05_nfkb1_fdx1_shortest_paths.txt"), sep="\t")

ws_paths['A1'] = "NFKB1到FDX1的8条最短路径"
ws_paths['A1'].font = title_font

# 简化显示：去重后的路径
unique_paths = [
    ["路径1-4", "NFKB1 → HSPA5 → FDX1", "HSPA5通路"],
    ["路径5-8", "NFKB1 → HMOX1 → FDX1", "HMOX1通路"]
]

ws_paths['A3'] = "去重后的通路（共2条）"
ws_paths['A3'].font = subtitle_font

headers = ['路径ID', '通路', "通路类型"]
for col, header in enumerate(headers, 1):
    cell = ws_paths.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

for i, path_data in enumerate(unique_paths, 1):
    for j, val in enumerate(path_data, 1):
        ws_paths.cell(row=4+i, column=j, value=val).border = thin_border

# 详细路径
ws_paths['A8'] = "详细路径列表（8条）"
ws_paths['A8'].font = subtitle_font

for r_idx, row in enumerate(dataframe_to_rows(paths_df, index=False, header=True), 9):
    for c_idx, value in enumerate(row, 1):
        cell = ws_paths.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        if r_idx == 9:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

for col in ['A', 'B', 'C', 'D']:
    ws_paths.column_dimensions[col].width = 20

# ==================== Sheet 3: 桥接节点分析 ====================
ws_bridge = wb.create_sheet("桥接节点分析")

bridge_df = pd.read_csv(os.path.join(result_dir, "05_nfkb1_fdx1_bridge_frequency.txt"), sep="\t")

ws_bridge['A1'] = "桥接节点频率分析"
ws_bridge['A1'].font = title_font

for r_idx, row in enumerate(dataframe_to_rows(bridge_df, index=False, header=True), 2):
    for c_idx, value in enumerate(row, 1):
        cell = ws_bridge.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        if r_idx == 2:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

# 添加桥接节点说明
ws_bridge['A6'] = "桥接节点功能说明"
ws_bridge['A6'].font = subtitle_font

bridge_info = [
    ["HSPA5", "GRP78/BiP", "内质网应激分子伴侣，参与蛋白折叠和应激反应"],
    ["HMOX1", "血红素加氧酶-1", "抗氧化应激关键酶，参与氧化还原调控"]
]

for col, header in enumerate(["基因", "别名", "功能"], 1):
    cell = ws_bridge.cell(row=7, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

for i, info in enumerate(bridge_info, 1):
    for j, val in enumerate(info, 1):
        ws_bridge.cell(row=7+i, column=j, value=val).border = thin_border

ws_bridge.column_dimensions['A'].width = 15
ws_bridge.column_dimensions['B'].width = 15
ws_bridge.column_dimensions['C'].width = 50

# ==================== Sheet 4: 详细路径节点表 ====================
ws_detailed = wb.create_sheet("详细路径节点表")

detailed_df = pd.read_csv(os.path.join(result_dir, "05_nfkb1_fdx1_detailed_paths.txt"), sep="\t")

ws_detailed['A1'] = "详细路径节点表（按路径和步骤）"
ws_detailed['A1'].font = title_font

# 简化显示
for r_idx, row in enumerate(dataframe_to_rows(detailed_df, index=False, header=True), 2):
    for c_idx, value in enumerate(row, 1):
        cell = ws_detailed.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        if r_idx == 2:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_detailed.column_dimensions[col].width = 15

# ==================== Sheet 5: 生物学意义解读 ====================
ws_bio = wb.create_sheet("生物学意义解读")

bio_content = [
    ["NFKB1到FDX1通路的生物学意义"],
    [""],
    ["1. 通路概述"],
    ["NFKB1（核因子κB1）是炎症反应和细胞存活的关键转录因子。"],
    ["FDX1（铁氧还蛋白1）是铜死亡（Cuproptosis）的核心调控基因。"],
    ["本分析发现NFKB1到FDX1的最短距离仅为2跳，表明两者存在紧密的功能联系。"],
    [""],
    ["2. 两条调控通路"],
    ["通路1: NFKB1 → HSPA5 → FDX1"],
    ["  • HSPA5（GRP78/BiP）是内质网应激的关键分子伴侣"],
    ["  • 铜死亡诱导内质网应激，HSPA5可能作为信号中介"],
    ["  • NFKB1可能通过调控HSPA5影响FDX1介导的铜死亡"],
    [""],
    ["通路2: NFKB1 → HMOX1 → FDX1"],
    ["  • HMOX1（血红素加氧酶-1）是重要的抗氧化酶"],
    ["  • 铜死亡涉及氧化应激，HMOX1可能参与氧化还原调控"],
    ["  • NFKB1可能通过HMOX1调控FDX1相关的氧化应激反应"],
    [""],
    ["3. 研究意义"],
    ["• 这两条通路为理解BCP（β-石竹烯）调控铜死亡的分子机制提供了线索"],
    ["• BCP可能通过抑制NFKB1，进而影响HSPA5/HMOX1，最终调控FDX1介导的铜死亡"],
    ["• HSPA5和HMOX1可作为潜在的药物靶点进行验证"],
    [""],
    ["4. 后续实验建议"],
    ["• 验证NFKB1对HSPA5和HMOX1的转录调控"],
    ["• 检测HSPA5和HMOX1对FDX1表达的影响"],
    ["• 在铜死亡模型中验证这两条通路的功能重要性"],
    ["• 评估BCP对这些节点表达的影响"]
]

for row_idx, row_data in enumerate(bio_content, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_bio.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.font = title_font
        elif value in ["1. 通路概述", "2. 两条调控通路", "3. 研究意义", "4. 后续实验建议"]:
            cell.font = subtitle_font
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

ws_bio.column_dimensions['A'].width = 80

# 保存文件
wb.save(output_file)
print(f"NFKB1到FDX1路径分析汇总Excel已生成: {output_file}")
print("包含以下sheet:")
for sheet in wb.sheetnames:
    print(f"  - {sheet}")
