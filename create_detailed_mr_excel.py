#!/usr/bin/env python3
# ================================================================================
# MR 分析详细结果汇总 - 完整 Excel 报告
# 整合所有 MR 分析数据，生成详细的研究报告
# ================================================================================

import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import math

# ================================================================================
# 辅助函数
# ================================================================================

def safe_float(value, default=''):
    """安全转换为浮点数"""
    if value is None or value == '' or value == 'NA':
        return default
    try:
        return float(value)
    except:
        return default

def safe_int(value, default=0):
    """安全转换为整数"""
    if value is None or value == '' or value == 'NA':
        return default
    try:
        return int(float(value))
    except:
        return default

def format_pvalue(pval):
    """格式化 P 值显示"""
    if pval == '' or pval is None:
        return 'NA'
    try:
        p = float(pval)
        if p < 0.0001:
            return f'<0.0001'
        elif p < 0.001:
            return f'{p:.4f}'
        elif p < 0.01:
            return f'{p:.3f}'
        elif p < 0.05:
            return f'{p:.3f}*'
        elif p < 0.1:
            return f'{p:.3f}'
        else:
            return f'{p:.3f}'
    except:
        return 'NA'

def read_csv_file(filepath):
    """读取 CSV 文件"""
    if not os.path.exists(filepath):
        return []
    
    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        data = list(reader)
    
    return data

# ================================================================================
# 数据加载
# ================================================================================

def load_all_data():
    """加载所有数据源"""
    print("="*70)
    print("加载 MR 分析数据...")
    print("="*70)
    
    # 1. MR 主要结果
    mr_file = r"D:\下载\MR_batch_results\20260508_optimized_fixed_v2\MR_results_main_optimized.csv"
    mr_data = read_csv_file(mr_file)
    print(f"✓ MR 主要结果：{len(mr_data)} 个基因")
    
    # 2. 分析日志
    log_file = r"D:\下载\MR_batch_results\20260508_optimized_fixed_v2\MR_analysis_log_optimized.csv"
    log_data = read_csv_file(log_file)
    print(f"✓ 分析日志：{len(log_data)} 条记录")
    
    # 3. Reactome 富集
    reactome_file = r"D:\下载\MR_batch_results\20260508_optimized_fixed_v2\functional_enrichment\Reactome_results.csv"
    reactome_data = read_csv_file(reactome_file)
    print(f"✓ Reactome 富集：{len(reactome_data)} 个通路")
    
    # 4. 药物靶点
    drug_file = r"D:\下载\MR_batch_results\20260508_optimized_fixed_v2\drug_targets\drug_targets_summary.csv"
    drug_data = read_csv_file(drug_file)
    print(f"✓ 药物靶点：{len(drug_data)} 个基因")
    
    return mr_data, log_data, reactome_data, drug_data

# ================================================================================
# Excel 创建
# ================================================================================

def create_detailed_excel(mr_data, log_data, reactome_data, drug_data, output_file):
    """创建详细的 MR 结果 Excel"""
    
    print(f"\n创建详细 Excel 报告...")
    
    wb = Workbook()
    
    # ============================================================================
    # Sheet 1: 主要结果摘要
    # ============================================================================
    ws_summary = wb.create_sheet(title="主要结果摘要")
    
    # 表头
    headers = [
        '基因', '状态', 'FDR 显著', '发现 P 值', '发现 OR', '发现 95%CI',
        'SNP 数量', 'F 统计量', '异质性 P', '多效性 P', 'Steiger P',
        '验证', 'Meta P', '证据等级', '优先级'
    ]
    ws_summary.append(headers)
    
    # 格式化表头
    for cell in ws_summary[1]:
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill('solid', start_color='4472C4')
        cell.font = Font(bold=True, color='FFFFFF')
    
    # 数据
    for i, row in enumerate(mr_data, 2):
        gene = row.get('gene', '')
        status = row.get('status', '')
        fdr_sig = '是' if row.get('fdr_sig') == 'TRUE' else '否'
        disc_pval = row.get('discovery_pval', '')
        disc_or = row.get('discovery_or', '')
        disc_ci = row.get('discovery_ci', '')
        nsnp = row.get('nsnp', '')
        f_mean = row.get('F_mean', '')
        het_p = row.get('Q_p', '')
        pleio_p = row.get('Egger_intercept_p', '')
        steiger_p = row.get('Steiger_p', '')
        has_rep = '是' if row.get('has_replication') == 'TRUE' else '否'
        meta_p = row.get('meta_pval', '')
        
        # 证据等级
        fdr_sig_bool = row.get('fdr_sig') == 'TRUE'
        disc_pval_float = safe_float(disc_pval)
        evidence = '高' if fdr_sig_bool else '中' if disc_pval_float != '' and disc_pval_float < 0.05 else '低'
        
        # 优先级
        priority = '⭐⭐⭐' if fdr_sig_bool else '⭐⭐' if disc_pval_float != '' and disc_pval_float < 0.05 else '⭐'
        
        ws_summary.append([
            gene, status, fdr_sig, 
            format_pvalue(disc_pval),
            f'{safe_float(disc_or, "NA"):.3f}' if disc_or != '' else 'NA',
            disc_ci,
            nsnp, f'{safe_float(f_mean, 0):.1f}',
            format_pvalue(het_p), format_pvalue(pleio_p), format_pvalue(steiger_p),
            has_rep, format_pvalue(meta_p),
            evidence, priority
        ])
        
        # 显著基因高亮
        if fdr_sig == '是':
            for cell in ws_summary[i]:
                cell.fill = PatternFill('solid', start_color='FFC000')
    
    # 调整列宽
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:
        ws_summary.column_dimensions[col].width = 12
    
    # ============================================================================
    # Sheet 2: 完整 MR 统计
    # ============================================================================
    ws_full = wb.create_sheet(title="完整 MR 统计")
    
    # 使用 MR 数据的所有列
    if mr_data:
        full_headers = list(mr_data[0].keys())
        ws_full.append(full_headers)
        
        # 格式化表头
        for cell in ws_full[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', start_color='4472C4')
            cell.font = Font(bold=True, color='FFFFFF')
        
        # 数据
        for row in mr_data:
            values = [row.get(h, '') for h in full_headers]
            ws_full.append(values)
    
    # ============================================================================
    # Sheet 3: 显著基因详情
    # ============================================================================
    ws_sig = wb.create_sheet(title="显著基因详情")
    
    # 筛选显著基因
    sig_genes = [row for row in mr_data if row.get('fdr_sig') == 'TRUE']
    
    sig_headers = [
        '基因', 'FDR q 值', '发现 P 值', '发现 OR', '发现 95%CI', '发现 Beta',
        '验证 P 值', '验证 OR', 'Meta P 值', 'Meta OR',
        'SNP 数', 'F 统计量', '状态', '证据等级'
    ]
    ws_sig.append(sig_headers)
    
    for cell in ws_sig[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', start_color='00B050')
        cell.font = Font(bold=True, color='FFFFFF')
    
    for row in sig_genes:
        gene = row.get('gene', '')
        fdr_q = row.get('fdr_qval', '')
        disc_p = row.get('discovery_pval', '')
        disc_or = row.get('discovery_or', '')
        disc_ci = row.get('discovery_ci', '')
        disc_b = row.get('discovery_b', '')
        rep_p = row.get('replication_pval', '')
        rep_or = row.get('replication_or', '')
        meta_p = row.get('meta_pval', '')
        meta_or = row.get('meta_or', '')
        nsnp = row.get('nsnp', '')
        f_mean = row.get('F_mean', '')
        status = row.get('status', '')
        evidence = '高'
        
        ws_sig.append([
            gene, f'{safe_float(fdr_q, 0):.4f}', format_pvalue(disc_p),
            f'{safe_float(disc_or, 0):.3f}', disc_ci, f'{safe_float(disc_b, 0):.4f}',
            format_pvalue(rep_p), f'{safe_float(rep_or, 0):.3f}',
            format_pvalue(meta_p), f'{safe_float(meta_or, 0):.3f}',
            nsnp, f'{safe_float(f_mean, 0):.1f}', status, evidence
        ])
    
    # ============================================================================
    # Sheet 4: 功能富集分析
    # ============================================================================
    ws_enrich = wb.create_sheet(title="功能富集分析")
    
    if reactome_data:
        enrich_headers = ['ID', '通路名称', 'P 值', 'q 值', '基因数', 'Z 得分', '基因列表']
        ws_enrich.append(enrich_headers)
        
        for cell in ws_enrich[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', start_color='FFC000')
        
        for row in reactome_data:
            ws_enrich.append([
                row.get('ID', ''),
                row.get('Description', ''),
                f'{safe_float(row.get("pvalue"), 0):.6f}',
                f'{safe_float(row.get("p.adjust"), 0):.6f}',
                row.get('Count', ''),
                f'{safe_float(row.get("zscore"), 0):.2f}',
                row.get('geneID', '')
            ])
    
    # ============================================================================
    # Sheet 5: 药物靶点信息
    # ============================================================================
    ws_drug = wb.create_sheet(title="药物靶点")
    
    if drug_data:
        drug_headers = ['Gene', 'DGIdb 发现', 'DGIdb 药物数', '药物名称', 
                       'OpenTargets', '可成药性', '已知药物数', '优先级']
        ws_drug.append(drug_headers)
        
        for cell in ws_drug[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', start_color='70AD47')
            cell.font = Font(bold=True, color='FFFFFF')
        
        for row in drug_data:
            ws_drug.append([
                row.get('gene', ''),
                '是' if row.get('dgidb_found') == 'TRUE' else '否',
                row.get('dgidb_count', ''),
                row.get('dgidb_drugs', ''),
                '是' if row.get('opentargets_found') == 'TRUE' else '否',
                row.get('tractability', ''),
                row.get('known_drugs_count', ''),
                row.get('drug_priority', '')
            ])
    
    # ============================================================================
    # Sheet 6: 分析质量控制
    # ============================================================================
    ws_qc = wb.create_sheet(title="质量控制")
    
    qc_headers = ['基因', '分析状态', '弱工具变量', 'Steiger 方向', '异质性', '多效性', '验证状态', '备注']
    ws_qc.append(qc_headers)
    
    for cell in ws_qc[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', start_color='FFC000')
    
    for row in mr_data:
        gene = row.get('gene', '')
        status = row.get('status', '')
        
        # 弱工具变量
        f_mean = safe_float(row.get('F_mean'), 0)
        weak_iv = '⚠️ 是' if f_mean < 10 else '✓ 否'
        
        # Steiger 方向
        steiger_p = safe_float(row.get('Steiger_p'), 1)
        steiger_ok = '✓ 正确' if steiger_p > 0.05 or steiger_p == 0 else '⚠️ 异常'
        
        # 异质性
        q_p = safe_float(row.get('Q_p'), 1)
        het_status = '⚠️ 显著' if q_p < 0.05 and q_p > 0 else '✓ 正常'
        
        # 多效性
        egger_p = safe_float(row.get('Egger_intercept_p'), 1)
        pleio_status = '⚠️ 显著' if egger_p < 0.05 and egger_p > 0 else '✓ 正常'
        
        # 验证
        has_rep = row.get('has_replication', 'FALSE')
        rep_status = '✓ 已验证' if has_rep == 'TRUE' else '未验证'
        
        # 备注
        notes = []
        if f_mean < 10:
            notes.append('弱工具变量')
        if q_p < 0.05 and q_p > 0:
            notes.append('存在异质性')
        if egger_p < 0.05 and egger_p > 0:
            notes.append('存在多效性')
        notes_str = '; '.join(notes) if notes else '通过质控'
        
        ws_qc.append([gene, status, weak_iv, steiger_ok, het_status, pleio_status, rep_status, notes_str])
    
    # ============================================================================
    # Sheet 7: 统计分析摘要
    # ============================================================================
    ws_stats = wb.create_sheet(title="统计摘要")
    
    # 基本统计
    total_genes = len(mr_data)
    success_genes = len([r for r in mr_data if r.get('status') == 'SUCCESS'])
    fdr_sig_genes = len([r for r in mr_data if r.get('fdr_sig') == 'TRUE'])
    pval_sig_genes = len([r for r in mr_data if safe_float(r.get('discovery_pval'), 1) < 0.05])
    
    stats_data = [
        ['基本统计', '', ''],
        ['总分析基因数', total_genes, ''],
        ['成功分析基因数', success_genes, f'{success_genes/total_genes*100:.1f}%'],
        ['FDR 显著基因数', fdr_sig_genes, f'{fdr_sig_genes/total_genes*100:.1f}%'],
        ['P 值显著基因数', pval_sig_genes, f'{pval_sig_genes/total_genes*100:.1f}%'],
        ['', '', ''],
        ['效应方向', '', ''],
        ['保护效应 (OR<1)', len([r for r in mr_data if safe_float(r.get('discovery_or'), 1) < 1]), ''],
        ['风险效应 (OR>1)', len([r for r in mr_data if safe_float(r.get('discovery_or'), 1) > 1]), ''],
        ['', '', ''],
        ['工具变量统计', '', ''],
        ['平均 SNP 数量', f'{sum(safe_int(r.get("nsnp"), 0) for r in mr_data)/total_genes:.1f}', ''],
        ['平均 F 统计量', f'{sum(safe_float(r.get("F_mean"), 0) for r in mr_data)/total_genes:.1f}', ''],
        ['', '', ''],
        ['敏感性分析', '', ''],
        ['异质性显著', len([r for r in mr_data if safe_float(r.get('Q_p'), 1) < 0.05]), '基因'],
        ['多效性显著', len([r for r in mr_data if safe_float(r.get('Egger_intercept_p'), 1) < 0.05]), '基因'],
        ['', '', ''],
        ['验证情况', '', ''],
        ['有独立验证', len([r for r in mr_data if r.get('has_replication') == 'TRUE']), '基因'],
        ['有 Meta 分析', len([r for r in mr_data if r.get('has_meta') == 'TRUE']), '基因'],
    ]
    
    for row in stats_data:
        ws_stats.append(row)
    
    # 格式化
    for i in range(3):
        ws_stats[1][i].font = Font(bold=True, size=14)
    
    for i in range(3):
        ws_stats[7][i].font = Font(bold=True)
        ws_stats[7][i].fill = PatternFill('solid', start_color='FFC000')
    
    for i in range(3):
        ws_stats[11][i].font = Font(bold=True)
        ws_stats[11][i].fill = PatternFill('solid', start_color='FFC000')
    
    for i in range(3):
        ws_stats[15][i].font = Font(bold=True)
        ws_stats[15][i].fill = PatternFill('solid', start_color='FFC000')
    
    for i in range(3):
        ws_stats[19][i].font = Font(bold=True)
        ws_stats[19][i].fill = PatternFill('solid', start_color='FFC000')
    
    # ============================================================================
    # Sheet 8: 元数据
    # ============================================================================
    ws_meta = wb.create_sheet(title="元数据")
    
    metadata = [
        ['MR 分析详细结果报告', ''],
        ['创建日期', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['数据来源', 'MR 分析优化路线 B'],
        ['分析版本', 'v2.0 (优化版)'],
        ['', ''],
        ['数据说明', ''],
        ['Sheet1: 主要结果摘要', '关键结果的快速查看'],
        ['Sheet2: 完整 MR 统计', '所有 MR 分析统计量'],
        ['Sheet3: 显著基因详情', 'FDR 显著基因详细信息'],
        ['Sheet4: 功能富集分析', 'Reactome 通路富集结果'],
        ['Sheet5: 药物靶点', 'DGIdb/OpenTargets 查询结果'],
        ['Sheet6: 质量控制', '每个基因的质控状态'],
        ['Sheet7: 统计摘要', '整体统计分析'],
        ['Sheet8: 元数据', '数据说明'],
        ['', ''],
        ['统计方法', ''],
        ['MR 方法', '逆方差加权法 (IVW)'],
        ['敏感性分析', 'MR-Egger, 加权中位数，Cochran Q'],
        ['多重检验校正', 'Benjamini-Hochberg FDR'],
        ['显著性阈值', 'FDR q < 0.05'],
        ['', ''],
        ['联系方式', ''],
        ['项目', '孟德尔随机化分析'],
        ['目标期刊', 'Nature Communications'],
    ]
    
    for row in metadata:
        ws_meta.append(row)
    
    # 格式化
    for i in range(2):
        ws_meta[1][i].font = Font(bold=True, size=16)
        ws_meta[6][i].font = Font(bold=True)
        ws_meta[6][i].fill = PatternFill('solid', start_color='4472C4')
        ws_meta[6][i].font = Font(bold=True, color='FFFFFF')
        ws_meta[15][i].font = Font(bold=True)
        ws_meta[15][i].fill = PatternFill('solid', start_color='4472C4')
        ws_meta[15][i].font = Font(bold=True, color='FFFFFF')
        ws_meta[21][i].font = Font(bold=True)
        ws_meta[21][i].fill = PatternFill('solid', start_color='4472C4')
        ws_meta[21][i].font = Font(bold=True, color='FFFFFF')
    
    # 调整列宽
    for col in ['A', 'B', 'C']:
        ws_meta.column_dimensions[col].width = 30
    
    # 保存
    print(f"保存 Excel 文件...")
    wb.save(output_file)
    print(f"✓ 文件已保存：{output_file}")

# ================================================================================
# 主函数
# ================================================================================

def main():
    print("="*70)
    print("MR 分析详细结果汇总")
    print("="*70)
    
    # 加载数据
    mr_data, log_data, reactome_data, drug_data = load_all_data()
    
    # 创建 Excel
    output_dir = r"D:\下载\MR_batch_results\20260508_optimized_fixed_v2"
    output_file = os.path.join(output_dir, "MR_详细结果汇总_20260508.xlsx")
    
    create_detailed_excel(mr_data, log_data, reactome_data, drug_data, output_file)
    
    # 打印总结
    print("\n" + "="*70)
    print("完成!")
    print("="*70)
    print(f"\n输出文件：{output_file}")
    print("\n包含 Sheets:")
    print("  1. 主要结果摘要 - 关键结果快速查看")
    print("  2. 完整 MR 统计 - 所有统计量")
    print("  3. 显著基因详情 - FDR 显著基因")
    print("  4. 功能富集分析 - Reactome 通路")
    print("  5. 药物靶点 - DGIdb/OpenTargets")
    print("  6. 质量控制 - 质控状态")
    print("  7. 统计摘要 - 整体统计")
    print("  8. 元数据 - 数据说明")
    print("="*70)

if __name__ == "__main__":
    main()
