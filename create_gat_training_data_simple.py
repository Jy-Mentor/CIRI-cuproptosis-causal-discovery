#!/usr/bin/env python3
# ================================================================================
# MR 分析结果汇总 - GAT 训练数据准备 (纯 Python 版本，无需 numpy/pandas)
# 整合所有 MR 分析结果，生成适合图注意力网络 (GAT) 训练的结构化数据
# ================================================================================

import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import math

# ================================================================================
# 第一部分：数据加载函数
# ================================================================================

def read_csv_file(filepath):
    """读取 CSV 文件"""
    if not os.path.exists(filepath):
        print(f"✗ 文件不存在：{filepath}")
        return []
    
    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        data = list(reader)
    
    print(f"✓ 加载 {len(data)} 条记录")
    return data

def load_mr_results():
    """加载 MR 主要结果"""
    mr_file = r"D:\下载\MR_batch_results\20260508_optimized_fixed_v2\MR_results_main_optimized.csv"
    print(f"\n加载 MR 结果...")
    return read_csv_file(mr_file)

def load_reactome_results():
    """加载 Reactome 富集结果"""
    reactome_file = r"D:\下载\MR_batch_results\20260508_optimized_fixed_v2\functional_enrichment\Reactome_results.csv"
    print(f"加载 Reactome 富集...")
    return read_csv_file(reactome_file)

def load_drug_targets():
    """加载药物靶点结果"""
    drug_file = r"D:\下载\MR_batch_results\20260508_optimized_fixed_v2\drug_targets\drug_targets_summary.csv"
    print(f"加载药物靶点...")
    return read_csv_file(drug_file)

# ================================================================================
# 第二部分：特征提取函数
# ================================================================================

def safe_float(value, default=0.0):
    """安全转换为浮点数"""
    if value is None or value == '' or value == 'NA':
        return default
    try:
        return float(value)
    except:
        return default

def safe_int(value, default=0):
    """安全转换为整数"""
    if value is None or value == '' or value == 'NA':
        return default
    try:
        return int(float(value))
    except:
        return default

def log10(value):
    """计算 log10，处理边界情况"""
    if value is None or value <= 0:
        return 0.0
    return -math.log10(value)

def extract_features(mr_row):
    """从 MR 行提取特征"""
    features = {}
    
    # MR 统计特征
    features['mr_beta'] = safe_float(mr_row.get('discovery_b'))
    features['mr_se'] = safe_float(mr_row.get('discovery_se'))
    features['mr_pval'] = safe_float(mr_row.get('discovery_pval'))
    features['mr_pval_log10'] = log10(safe_float(mr_row.get('discovery_pval', 1)))
    features['fdr_qval'] = safe_float(mr_row.get('fdr_qval'))
    
    # 工具变量特征
    features['nsnp'] = safe_int(mr_row.get('nsnp'))
    features['f_mean'] = safe_float(mr_row.get('F_mean'))
    
    # 敏感性分析特征
    features['heterogeneity_p'] = safe_float(mr_row.get('Q_p'))
    features['pleiotropy_p'] = safe_float(mr_row.get('Egger_intercept_p'))
    features['steiger_p'] = safe_float(mr_row.get('Steiger_p'))
    
    # 验证特征
    has_rep = mr_row.get('has_replication', 'FALSE')
    features['has_replication'] = 1 if str(has_rep).upper() == 'TRUE' else 0
    
    has_meta = mr_row.get('has_meta', 'FALSE')
    features['has_meta'] = 1 if str(has_meta).upper() == 'TRUE' else 0
    
    # 状态编码
    status = mr_row.get('status', 'UNKNOWN')
    features['is_success'] = 1 if status == 'SUCCESS' else 0
    
    fdr_sig = mr_row.get('fdr_sig', 'FALSE')
    features['is_fdr_sig'] = 1 if str(fdr_sig).upper() == 'TRUE' else 0
    
    return features

def create_target(mr_row):
    """创建目标变量"""
    target = {}
    
    # 是否显著
    fdr_qval = safe_float(mr_row.get('fdr_qval'))
    is_sig = 1 if fdr_qval < 0.05 and fdr_qval > 0 else 0
    target['is_significant'] = is_sig
    
    # 效应方向
    if is_sig:
        beta = safe_float(mr_row.get('discovery_b'))
        if beta < 0:
            target['effect_direction'] = 1  # 保护效应
        elif beta > 0:
            target['effect_direction'] = 2  # 风险效应
        else:
            target['effect_direction'] = 0
    else:
        target['effect_direction'] = 0
    
    # 置信度
    if is_sig:
        pval = safe_float(mr_row.get('discovery_pval', 1))
        if pval > 0:
            target['confidence_score'] = min(1.0, log10(pval) / 10)
        else:
            target['confidence_score'] = 0.0
    else:
        target['confidence_score'] = 0.0
    
    return target

def create_network_edges(mr_data):
    """创建网络边"""
    edges = []
    
    # 找出所有显著基因
    significant_genes = []
    for row in mr_data:
        fdr_qval = safe_float(row.get('fdr_qval'))
        if fdr_qval < 0.05 and fdr_qval > 0:
            gene = row.get('gene', '')
            if gene:
                significant_genes.append(gene)
    
    print(f"  找到 {len(significant_genes)} 个显著基因")
    
    # 创建完全连接
    for i, gene1 in enumerate(significant_genes):
        for gene2 in significant_genes[i+1:]:
            edges.append({
                'source': gene1,
                'target': gene2,
                'edge_type': 'co_significant',
                'weight': 1.0
            })
    
    return edges

# ================================================================================
# 第三部分：Excel 创建函数
# ================================================================================

def create_excel_file(features_list, targets_list, edges_list, output_file):
    """创建 Excel 文件"""
    
    print(f"\n创建 Excel 文件：{output_file}")
    
    wb = Workbook()
    
    # Sheet 1: 节点特征
    ws_features = wb.create_sheet(title="Node_Features")
    
    # 表头
    if features_list:
        feature_names = list(features_list[0].keys())
        headers = ['gene_id'] + feature_names
        ws_features.append(headers)
        
        # 数据
        for feat in features_list:
            row = [feat['gene']] + [feat[name] for name in feature_names]
            ws_features.append(row)
        
        # 格式化表头
        for cell in ws_features[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', start_color='FFFF00')
    
    # Sheet 2: 目标变量
    ws_targets = wb.create_sheet(title="Target_Variables")
    ws_targets.append(['gene_id', 'is_significant', 'effect_direction', 'confidence_score'])
    
    for tgt in targets_list:
        ws_targets.append([
            tgt['gene'],
            tgt['is_significant'],
            tgt['effect_direction'],
            round(tgt['confidence_score'], 4)
        ])
    
    for cell in ws_targets[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', start_color='FFFF00')
    
    # Sheet 3: 边列表
    ws_edges = wb.create_sheet(title="Edge_List")
    ws_edges.append(['source_gene', 'target_gene', 'edge_type', 'weight'])
    
    for edge in edges_list:
        ws_edges.append([
            edge['source'],
            edge['target'],
            edge['edge_type'],
            round(edge['weight'], 4)
        ])
    
    for cell in ws_edges[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', start_color='FFFF00')
    
    # Sheet 4: 统计信息
    ws_stats = wb.create_sheet(title="Statistics")
    
    num_genes = len(features_list)
    num_features = len(feature_names) if features_list else 0
    num_sig = sum(1 for t in targets_list if t['is_significant'] == 1)
    num_protective = sum(1 for t in targets_list if t['effect_direction'] == 1)
    num_risk = sum(1 for t in targets_list if t['effect_direction'] == 2)
    num_edges = len(edges_list)
    
    stats = [
        ['数据集统计', '', ''],
        ['总基因数', num_genes, ''],
        ['特征维度', num_features, '(不包括 gene_id)'],
        ['显著基因数', num_sig, '(FDR < 0.05)'],
        ['保护效应基因', num_protective, '(beta < 0)'],
        ['风险效应基因', num_risk, '(beta > 0)'],
        ['网络边数', num_edges, ''],
        ['', '', ''],
        ['特征列表', '', ''],
    ]
    
    for row in stats:
        ws_stats.append(row)
    
    # 添加特征名
    if features_list:
        for i, fname in enumerate(feature_names, 1):
            ws_stats.append([f'特征 {i}', fname, ''])
    
    # 格式化
    for cell in ws_stats[1][0:3]:
        cell.font = Font(bold=True, size=12)
    
    for cell in ws_stats[9][0:3]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', start_color='FFFF00')
    
    # Sheet 5: 元数据
    ws_metadata = wb.create_sheet(title="Metadata")
    
    metadata = [
        ['数据集信息', ''],
        ['创建日期', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['数据来源', 'MR 分析优化路线 B'],
        ['目标期刊', 'Nature Communications'],
        ['', ''],
        ['特征说明', ''],
        ['mr_beta', 'MR 分析的效应量 (beta)'],
        ['mr_se', 'MR 效应量的标准误'],
        ['mr_pval', 'MR 分析的 P 值'],
        ['mr_pval_log10', '-log10(P 值)'],
        ['fdr_qval', 'FDR 校正后的 q 值'],
        ['nsnp', '工具变量 (SNP) 数量'],
        ['f_mean', '平均 F 统计量'],
        ['heterogeneity_p', '异质性检验 P 值'],
        ['pleiotropy_p', '多效性检验 P 值'],
        ['steiger_p', 'Steiger 方向检验 P 值'],
        ['has_replication', '是否有独立验证 (0/1)'],
        ['has_meta', '是否有 Meta 分析 (0/1)'],
        ['is_success', '分析是否成功 (0/1)'],
        ['is_fdr_sig', '是否 FDR 显著 (0/1)'],
        ['', ''],
        ['目标变量说明', ''],
        ['is_significant', '是否显著 (0: 不显著，1: FDR<0.05)'],
        ['effect_direction', '效应方向 (0: 无，1: 保护，2: 风险)'],
        ['confidence_score', '置信度得分 (0-1)'],
        ['', ''],
        ['GAT 训练建议', ''],
        ['输入维度', num_features],
        ['输出类别', '3 (无显著/保护/风险)'],
        ['图结构', '基于共显著性的边'],
        ['推荐隐藏层', '2-3 层'],
        ['推荐注意力头', '4-8 头'],
    ]
    
    for row in metadata:
        ws_metadata.append(row)
    
    # 格式化
    for i in range(2):
        ws_metadata[1][i].font = Font(bold=True, size=12)
        ws_metadata[6][i].font = Font(bold=True)
        ws_metadata[6][i].fill = PatternFill('solid', start_color='FFFF00')
        ws_metadata[21][i].font = Font(bold=True)
        ws_metadata[21][i].fill = PatternFill('solid', start_color='FFFF00')
        ws_metadata[25][i].font = Font(bold=True)
        ws_metadata[25][i].fill = PatternFill('solid', start_color='FFFF00')
    
    # 调整列宽
    for col in ['A', 'B', 'C']:
        ws_metadata.column_dimensions[col].width = 25
    
    # 保存
    wb.save(output_file)
    print(f"✓ Excel 文件已保存")

def create_gat_training_code(output_dir):
    """创建 GAT 训练代码"""
    
    code = """#!/usr/bin/env python3
\"\"\"
GAT 训练代码 - MR 分析结果
使用 PyTorch Geometric
\"\"\"

import torch
import torch.nn.functional as F
from torch_geometric.nn import GATConv
from torch_geometric.data import Data
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score, classification_report

class GAT(torch.nn.Module):
    def __init__(self, num_features, num_classes=3, hidden_channels=64, heads=4, dropout=0.6):
        super(GAT, self).__init__()
        
        self.conv1 = GATConv(num_features, hidden_channels, heads=heads, dropout=dropout)
        self.conv2 = GATConv(hidden_channels * heads, num_classes, heads=1, dropout=dropout)
        self.dropout = torch.nn.Dropout(dropout)
    
    def forward(self, x, edge_index):
        x = F.elu(self.conv1(x, edge_index))
        x = self.dropout(x)
        x = self.conv2(x, edge_index)
        return F.log_softmax(x, dim=1)

def load_data(excel_file):
    # 加载数据
    features_df = pd.read_excel(excel_file, sheet_name='Node_Features')
    targets_df = pd.read_excel(excel_file, sheet_name='Target_Variables')
    edges_df = pd.read_excel(excel_file, sheet_name='Edge_List')
    
    # 基因映射
    genes = features_df['gene_id'].tolist()
    gene2idx = {g: i for i, g in enumerate(genes)}
    
    # 特征矩阵
    feature_cols = [c for c in features_df.columns if c != 'gene_id']
    X = torch.FloatTensor(features_df[feature_cols].fillna(0).values)
    
    # 目标
    y = torch.LongTensor(targets_df['effect_direction'].values)
    
    # 边
    edge_list = []
    for _, row in edges_df.iterrows():
        src = gene2idx.get(row['source_gene'])
        tgt = gene2idx.get(row['target_gene'])
        if src is not None and tgt is not None:
            edge_list.append([src, tgt])
            edge_list.append([tgt, src])
    
    edge_index = torch.LongTensor(edge_list).t().contiguous()
    
    return Data(x=X, edge_index=edge_index, y=y), genes, gene2idx

def train():
    data, genes, gene2idx = load_data('MR_GAT_Training_Data_20260508.xlsx')
    
    # 训练集划分
    n = len(data.y)
    indices = torch.randperm(n)
    train_idx = indices[:int(0.7*n)]
    test_idx = indices[int(0.7*n):]
    
    train_mask = torch.zeros(n, dtype=torch.bool)
    test_mask = torch.zeros(n, dtype=torch.bool)
    train_mask[train_idx] = True
    test_mask[test_idx] = True
    
    # 模型
    model = GAT(num_features=data.num_node_features)
    optimizer = torch.optim.Adam(model.parameters(), lr=0.01, weight_decay=5e-4)
    
    # 训练
    for epoch in range(200):
        model.train()
        optimizer.zero_grad()
        out = model(data.x, data.edge_index)
        loss = F.nll_loss(out[train_mask], data.y[train_mask])
        loss.backward()
        optimizer.step()
        
        if (epoch+1) % 20 == 0:
            model.eval()
            pred = out.argmax(dim=1)
            train_acc = accuracy_score(data.y[train_mask], pred[train_mask])
            test_acc = accuracy_score(data.y[test_mask], pred[test_mask])
            print(f'Epoch {epoch+1}: Loss={loss.item():.4f}, Train Acc={train_acc:.4f}, Test Acc={test_acc:.4f}')
    
    # 评估
    model.eval()
    pred = model(data.x, data.edge_index).argmax(dim=1)
    print('\\n分类报告:')
    print(classification_report(data.y[test_mask], pred[test_mask], 
                                target_names=['无显著', '保护', '风险']))

if __name__ == '__main__':
    train()
"""
    
    code_file = os.path.join(output_dir, "gat_training_example.py")
    with open(code_file, 'w', encoding='utf-8') as f:
        f.write(code)
    
    print(f"✓ GAT 训练代码已保存：{code_file}")

# ================================================================================
# 第四部分：主函数
# ================================================================================

def main():
    print("="*70)
    print("MR 分析结果汇总 - GAT 训练数据 (纯 Python 版本)")
    print("="*70)
    
    # 加载数据
    mr_data = load_mr_results()
    reactome_data = load_reactome_results()
    drug_data = load_drug_targets()
    
    if not mr_data:
        print("错误：无法加载 MR 数据")
        return
    
    # 提取特征
    print("\\n提取特征...")
    features_list = []
    targets_list = []
    
    for i, row in enumerate(mr_data):
        gene = row.get('gene', f'Gene_{i}')
        print(f"  处理 {i+1}/{len(mr_data)}: {gene}")
        
        features = extract_features(row)
        features['gene'] = gene
        features_list.append(features)
        
        target = create_target(row)
        target['gene'] = gene
        targets_list.append(target)
    
    # 创建网络
    print("\\n创建基因网络...")
    edges_list = create_network_edges(mr_data)
    
    # 创建 Excel
    output_dir = r"D:\下载\MR_batch_results\20260508_optimized_fixed_v2"
    output_file = os.path.join(output_dir, "MR_GAT_Training_Data_20260508.xlsx")
    
    create_excel_file(features_list, targets_list, edges_list, output_file)
    create_gat_training_code(output_dir)
    
    # 打印总结
    print("\\n" + "="*70)
    print("完成!")
    print("="*70)
    print(f"\\n输出文件:")
    print(f"  Excel: {output_file}")
    print(f"  训练代码：{os.path.join(output_dir, 'gat_training_example.py')}")
    print(f"\\n数据统计:")
    print(f"  基因数：{len(features_list)}")
    print(f"  特征数：{len(features_list[0]) - 1}")
    print(f"  显著基因：{sum(1 for t in targets_list if t['is_significant']==1)}")
    print(f"  网络边：{len(edges_list)}")
    print("="*70)

if __name__ == "__main__":
    main()
