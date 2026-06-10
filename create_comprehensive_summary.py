#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GSE174574 单细胞分析结果综合汇总
整合所有关键结果到一个Excel文件
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

WORK_DIR = r"C:\Users\Jy-Mentor-7\Desktop\你信我 这不是重蹈覆辙"
RESULTS_DIR = os.path.join(WORK_DIR, "GSE174574_SCISSOR_Results")
OUTPUT_FILE = os.path.join(WORK_DIR, "GSE174574_综合分析结果汇总.xlsx")

def create_summary_excel():
    """创建综合汇总Excel文件"""
    
    # 创建工作簿
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认sheet
    
    # 定义样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== Sheet 1: 分析概览 =====
    ws_overview = wb.create_sheet("分析概览")
    overview_data = [
        ["GSE174574 单细胞RNA-seq综合分析结果汇总"],
        [],
        ["分析项目", "结果"],
        ["数据集", "GSE174574 (小鼠 MCAO 单细胞)"],
        ["总细胞数", "57,224"],
        ["使用基因数", "3,000 (HVG)"],
        ["样本分组", "Sham (n=3) vs MCAO (n=3)"],
        ["细胞类型数", "7种"],
        ["Hub基因数", "10个"],
        [],
        ["关键发现"],
        ["1. NFKB1-FDX1相关性 (MCAO组)", "ρ = 0.443, p < 0.001"],
        ["2. 最敏感细胞类型", "Microglia (Effect Size = 最强)"],
        ["3. 虚拟敲除靶点", "Nfkb1, Fdx1, Stat3, Tlr4等8个基因"],
    ]
    
    for row_idx, row_data in enumerate(overview_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_overview.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
                cell.alignment = Alignment(horizontal='center')
            elif row_idx == 3:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
            elif row_idx == 11:
                cell.font = Font(bold=True, size=12)
    
    ws_overview.column_dimensions['A'].width = 35
    ws_overview.column_dimensions['B'].width = 40
    
    # ===== Sheet 2: 细胞类型差异检验 =====
    ws_diff = wb.create_sheet("细胞类型差异检验")
    
    # 读取差异检验结果
    diff_file = os.path.join(RESULTS_DIR, "02_net_score_by_celltype_v3.csv")
    if os.path.exists(diff_file):
        df_diff = pd.read_csv(diff_file)
        
        # 添加标题
        ws_diff.append(["SCISSOR Net Score 细胞类型差异检验结果"])
        ws_diff.append([])
        
        # 写入表头
        headers = ["细胞类型", "MCAO中位数", "Sham中位数", "P值", "校正P值", "效应量", "MCAO样本数", "Sham样本数"]
        ws_diff.append(headers)
        for cell in ws_diff[3]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        
        # 写入数据
        for _, row in df_diff.iterrows():
            ws_diff.append([
                row['CellType'],
                round(row['MCAO_Median'], 4),
                round(row['Sham_Median'], 4),
                f"{row['P_value']:.2e}",
                f"{row['P_adj']:.2e}",
                round(row['Effect_Size'], 4),
                int(row['MCAO_N']),
                int(row['Sham_N'])
            ])
        
        # 设置列宽
        ws_diff.column_dimensions['A'].width = 20
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws_diff.column_dimensions[col].width = 15
    
    # ===== Sheet 3: NFKB1-FDX1相关性 =====
    ws_cor = wb.create_sheet("NFKB1-FDX1相关性")
    
    cor_file = os.path.join(RESULTS_DIR, "03_nfkb1_fdx1_correlation_v3.csv")
    if os.path.exists(cor_file):
        df_cor = pd.read_csv(cor_file)
        
        ws_cor.append(["NFKB1与FDX1表达相关性分析"])
        ws_cor.append([])
        
        headers = ["细胞类型", "分组", "Spearman相关系数", "P值", "样本数"]
        ws_cor.append(headers)
        for cell in ws_cor[3]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        
        for _, row in df_cor.iterrows():
            ws_cor.append([
                row['CellType'],
                row['Group'],
                round(row['Correlation'], 4),
                f"{row['P_value']:.2e}",
                int(row['N'])
            ])
        
        ws_cor.append([])
        ws_cor.append(["关键发现:"])
        ws_cor.append(["MCAO组中NFKB1与FDX1呈显著正相关 (ρ=0.4429, p<0.001),"])
        ws_cor.append(["提示二者在脑卒中病理过程中可能存在协同调控机制。"])
        
        ws_cor.column_dimensions['A'].width = 18
        ws_cor.column_dimensions['B'].width = 12
        ws_cor.column_dimensions['C'].width = 20
        ws_cor.column_dimensions['D'].width = 15
        ws_cor.column_dimensions['E'].width = 12
    
    # ===== Sheet 4: 虚拟敲除统计 =====
    ws_kd = wb.create_sheet("虚拟敲除统计")
    
    kd_stats_mcaofile = os.path.join(RESULTS_DIR, "virtual_knockdown", "knockdown_stats_MCAO.csv")
    kd_stats_shamfile = os.path.join(RESULTS_DIR, "virtual_knockdown", "knockdown_stats_Sham.csv")
    
    row_idx = 1
    ws_kd.append(["虚拟敲除分析统计汇总"])
    row_idx += 2
    
    # MCAO组
    if os.path.exists(kd_stats_mcaofile):
        df_kd_m = pd.read_csv(kd_stats_mcaofile)
        ws_kd.append(["MCAO组虚拟敲除统计"])
        row_idx += 1
        
        headers = list(df_kd_m.columns)
        ws_kd.append(headers)
        for cell in ws_kd[row_idx]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        row_idx += 1
        
        for _, row in df_kd_m.iterrows():
            ws_kd.append(list(row))
            row_idx += 1
        
        ws_kd.append([])
        row_idx += 1
    
    # Sham组
    if os.path.exists(kd_stats_shamfile):
        df_kd_s = pd.read_csv(kd_stats_shamfile)
        ws_kd.append(["Sham组虚拟敲除统计"])
        row_idx += 1
        
        headers = list(df_kd_s.columns)
        ws_kd.append(headers)
        for cell in ws_kd[row_idx]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        row_idx += 1
        
        for _, row in df_kd_s.iterrows():
            ws_kd.append(list(row))
            row_idx += 1
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_kd.column_dimensions[col].width = 18
    
    # ===== Sheet 5: Hub基因列表 =====
    ws_hub = wb.create_sheet("Hub基因列表")
    
    hub_genes_data = [
        ["Hub基因列表 (跨物种映射后)"],
        [],
        ["人类基因", "小鼠同源基因", "功能分类", "在数据中"],
        ["NFKB1", "Nfkb1", "炎症调控", "✓"],
        ["FDX1", "Fdx1", "铜死亡核心", "✓"],
        ["STAT3", "Stat3", "转录因子", "✓"],
        ["HSPA5", "Hspa5", "内质网应激", "✓"],
        ["HMOX1", "Hmox1", "抗氧化应激", "✓"],
        ["HIF1A", "Hif1a", "缺氧应答", "✓"],
        ["TNF", "Tnf", "炎症因子", "✓"],
        ["IL6", "Il6", "炎症因子", "✓"],
        ["GPX4", "Gpx4", "铁死亡/抗氧化", "✓"],
        ["DLAT", "Dlat", "铜死亡相关", "✓"],
        [],
        ["注: 所有Hub基因均成功映射到GSE174574数据集中"]
    ]
    
    for row_data in hub_genes_data:
        ws_hub.append(row_data)
    
    # 设置表头样式
    for cell in ws_hub[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    ws_hub.column_dimensions['A'].width = 15
    ws_hub.column_dimensions['B'].width = 18
    ws_hub.column_dimensions['C'].width = 20
    ws_hub.column_dimensions['D'].width = 12
    
    # ===== Sheet 6: 方法学说明 =====
    ws_method = wb.create_sheet("方法学说明")
    
    method_data = [
        ["分析方法学说明"],
        [],
        ["1. 数据预处理"],
        ["   - 使用Scanpy进行单细胞数据分析"],
        ["   - 质控标准: n_genes > 200, n_genes < 7500, pct_mt < 10%"],
        ["   - 标准化: Normalize_total + log1p"],
        ["   - 高变基因选择: 3,000个HVG (强制包含Hub基因)"],
        [],
        ["2. SCISSOR-like评分"],
        ["   - MCAO特征基因: 14个 (Il6, Tnf, Nfkb1等)"],
        ["   - Sham特征基因: 8个 (Bdnf, Ngf, Snap25等)"],
        ["   - Net Score = MCAO_Score - Sham_Score"],
        [],
        ["3. 细胞类型注释"],
        ["   - 方法: 基于已知marker基因的评分注释"],
        ["   - 细胞类型: Microglia, Astrocytes, Neuron, Oligodendrocytes, Endothelial, OPC, Pericytes"],
        [],
        ["4. 统计分析"],
        ["   - 差异检验: Mann-Whitney U检验"],
        ["   - 多重检验校正: Bonferroni方法"],
        ["   - 相关性分析: Spearman秩相关"],
        [],
        ["5. 虚拟敲除分析"],
        ["   - 细胞类型: Microglia"],
        ["   - 方法: 基于相关性的基因表达预测"],
        ["   - 显著性阈值: |Predicted_logFC| > 0.1且P < 0.05"],
    ]
    
    for row_data in method_data:
        ws_method.append(row_data)
    
    ws_method.column_dimensions['A'].width = 80
    
    # 保存工作簿
    wb.save(OUTPUT_FILE)
    print(f"✅ 综合汇总Excel已保存: {OUTPUT_FILE}")
    
    # 打印汇总信息
    print("\n" + "="*60)
    print("汇总文件内容:")
    print("="*60)
    for sheet_name in wb.sheetnames:
        print(f"  - {sheet_name}")
    print("="*60)

if __name__ == "__main__":
    create_summary_excel()
