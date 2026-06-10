# -*- coding: utf-8 -*-
"""
汇总Stage1-9所有结果到Excel报告
"""
import os
import sys
import json
import time
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = r"c:\Users\Jy-Mentor-7\Desktop\你信我 这不是重蹈覆辙"
RESULTS_DIR = os.path.join(BASE_DIR, "results")

def main():
    wb = Workbook()
    
    # Sheet 1: 总览
    ws1 = wb.active
    ws1.title = "总览"
    ws1['A1'] = "多组学整合分析揭示 BCP 治疗 CIRI 的潜在靶点 - 结果总览 (v8 无偏管线)"
    ws1['A1'].font = Font(bold=True, size=14)
    ws1.merge_cells('A1:F1')
    
    ws1['A3'] = "阶段"
    ws1['B3'] = "功能"
    ws1['C3'] = "输入"
    ws1['D3'] = "关键结果"
    ws1['E3'] = "状态"
    ws1['F3'] = "问题"
    for c in range(1, 7):
        ws1.cell(row=3, column=c).font = Font(bold=True)
        ws1.cell(row=3, column=c).fill = PatternFill("solid", fgColor="4472C4")
        ws1.cell(row=3, column=c).font = Font(bold=True, color="FFFFFF")
    
    stages = [
        ["Stage1", "RMA标准化+limma差异分析", "GSE61616原始数据", "大鼠脑组织5 Sham vs 5 MCAO, 1836 DEGs", "✅", ""],
        ["Stage2", "单细胞聚类分析", "GSE210986小鼠scRNA-seq", "10567细胞 (5411 MCAO, 5156 Sham)", "✅", ""],
        ["Stage3", "GO/KEGG/GSEA富集分析", "Stage1 DEGs+大鼠→人类映射", "人类同源DEGs, 通路富集", "✅", ""],
        ["Stage4", "三层种子池+WGCNA(v8 无偏)", "Bulk+SC+Homology, 全种子池", "~2059种子基因(无先验过滤)", "✅", "无偏管线: 不做铜死亡过滤"],
        ["Stage5", "STRING PPI网络+拓扑", "种子池基因(人类9606)", "9K+节点, 全PPI网络", "✅", "v8无偏: 全种子网络"],
        ["Stage6", "GRN虚拟敲除(Spearman+Top-N)", "单细胞MCAO数据", "基因扰动评分", "✅", "Top-N相关性策略"],
        ["Stage7", "ML+LOO-CV+Bootstrap(v5)", "GSE174574 pseudo-bulk 6样本", "LR LOO-CV AUC=1.0, Bootstrap×100", "✅", "Bootstrap稳定性加权"],
        ["Stage8", "多组学融合靶点排名(v8 无偏)", "Stage5-7数据驱动整合", "纯数据驱动: GRN(0.35)+ML(0.131)+PPI(0.30)", "✅", "删除BCP/铜死亡先验权重"],
        ["Stage9", "PPI-GAT节点回归(v8 无偏)", "PPI网络, Stage8综合得分标签", "GAT R²=?, 8D特征(无先验)", "✅", "数据驱动标签: Stage8得分"],
    ]
    
    for i, row in enumerate(stages, 4):
        for j, val in enumerate(row):
            cell = ws1.cell(row=i, column=j+1, value=val)
            if val == "✅":
                cell.fill = PatternFill("solid", fgColor="C6EFCE")
            elif val == "⚠️":
                cell.fill = PatternFill("solid", fgColor="FFEB9C")
    
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 30
    ws1.column_dimensions['C'].width = 35
    ws1.column_dimensions['D'].width = 35
    ws1.column_dimensions['E'].width = 10
    ws1.column_dimensions['F'].width = 20
    
    # Sheet 2: Tier1靶点 (only Tier1 genes from core_targets.csv)
    ws2 = wb.create_sheet("Tier1靶点")
    core_file = os.path.join(RESULTS_DIR, "stage8_final_targets", "core_targets.csv")
    tier1_file = os.path.join(RESULTS_DIR, "stage8_final_targets", "tier1_targets.csv")
    if os.path.exists(tier1_file):
        df_t1 = pd.read_csv(tier1_file)
        ws2['A1'] = f"Tier1核心靶点 (高置信, {len(df_t1)} genes)"
        ws2['A1'].font = Font(bold=True, size=14)
        for i, col in enumerate(df_t1.columns, 1):
            cell = ws2.cell(row=2, column=i, value=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.font = Font(bold=True, color="FFFFFF")
        for r, row_data in enumerate(df_t1.values, 3):
            for c, val in enumerate(row_data, 1):
                ws2.cell(row=r, column=c, value=val)
        for col_idx in range(1, len(df_t1.columns)+1):
            ws2.column_dimensions[get_column_letter(col_idx)].width = 18
    
    # Sheet 3: GRN敲除结果
    ws3 = wb.create_sheet("GRN敲除")
    grn_file = os.path.join(RESULTS_DIR, "stage6_sctenifold_knockout", "gene_perturbation_scores.csv")
    if os.path.exists(grn_file):
        df_grn = pd.read_csv(grn_file)
        ws3['A1'] = "GRN虚拟敲除 - 扰动评分"
        ws3['A1'].font = Font(bold=True, size=14)
        for i, col in enumerate(df_grn.columns, 1):
            cell = ws3.cell(row=2, column=i, value=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.font = Font(bold=True, color="FFFFFF")
        for r, row_data in enumerate(df_grn.values, 3):
            for c, val in enumerate(row_data, 1):
                ws3.cell(row=r, column=c, value=val)
    
    # Sheet 4: GAT排名
    ws4 = wb.create_sheet("GAT排名")
    gat_file = os.path.join(RESULTS_DIR, "stage9_ppi_gat", "gat_gene_ranking.csv")
    if os.path.exists(gat_file):
        df_gat = pd.read_csv(gat_file)
        ws4['A1'] = "GAT节点回归 - 基因排名"
        ws4['A1'].font = Font(bold=True, size=14)
        for i, col in enumerate(df_gat.columns, 1):
            cell = ws4.cell(row=2, column=i, value=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.font = Font(bold=True, color="FFFFFF")
        for r, row_data in enumerate(df_gat.values, 3):
            for c, val in enumerate(row_data, 1):
                ws4.cell(row=r, column=c, value=val)
    
    # Sheet 5: ML结果
    ws5 = wb.create_sheet("ML_SHAP")
    ml_file = os.path.join(RESULTS_DIR, "stage7_ml_shap", "ml_model_performance.csv")
    if os.path.exists(ml_file):
        df_ml = pd.read_csv(ml_file)
        ws5['A1'] = "ML模型性能"
        ws5['A1'].font = Font(bold=True, size=14)
        for i, col in enumerate(df_ml.columns, 1):
            cell = ws5.cell(row=2, column=i, value=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.font = Font(bold=True, color="FFFFFF")
        for r, row_data in enumerate(df_ml.values, 3):
            for c, val in enumerate(row_data, 1):
                ws5.cell(row=r, column=c, value=val)
    
    # Sheet 6: 铜死亡基因排名 (use core_targets.csv for all 17 copper genes)
    ws6 = wb.create_sheet("铜死亡排名")
    cupro_genes = ["FDX1","LIAS","LIPT1","DLAT","PDHA1","PDHB","MTF1","GLS","CDKN2A","SLC31A1","ATP7A","ATP7B","DLD","DBT","DLST","PDHA2","GCSH"]
    core_file = os.path.join(RESULTS_DIR, "stage8_final_targets", "core_targets.csv")
    if os.path.exists(core_file):
        df_all_core = pd.read_csv(core_file)
        df_cupro = df_all_core[df_all_core["Gene"].isin(cupro_genes)].sort_values("Comprehensive", ascending=False)
    else:
        df_cupro = pd.DataFrame()
    ws6['A1'] = "铜死亡核心基因在综合排名中的位置"
    ws6['A1'].font = Font(bold=True, size=14)
    if not df_cupro.empty:
        for i, col in enumerate(df_cupro.columns, 1):
            cell = ws6.cell(row=2, column=i, value=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.font = Font(bold=True, color="FFFFFF")
        for r, row_data in enumerate(df_cupro.values, 3):
            for c, val in enumerate(row_data, 1):
                ws6.cell(row=r, column=c, value=val)
    
    output_file = os.path.join(RESULTS_DIR, f"complete_pipeline_results_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(output_file)
    print(f"Excel报告已保存: {output_file}")

if __name__ == "__main__":
    main()
