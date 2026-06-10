import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 文件路径
work_dir = "C:/Users/Jy-Mentor-7/Desktop/你信我 这不是重蹈覆辙"
mapped_excel = f"{work_dir}/PPI_Input_Genes_Human.xlsx"
mapping_file = f"{work_dir}/大创/大鼠 小鼠 人类映射库.txt"
output_file = f"{work_dir}/PPI_Input_Genes_Human_Complete_Final.xlsx"

# 读取已映射的人类基因
mapped_df = pd.read_excel(mapped_excel, sheet_name="PPI输入基因 (Human)", skiprows=7)
mapped_human_genes = set()
for col in mapped_df.columns:
    for val in mapped_df[col]:
        if pd.notna(val) and str(val).strip() != '' and str(val) != '人类基因':
            mapped_human_genes.add(str(val).strip().upper())

print(f"已映射的人类基因数: {len(mapped_human_genes)}")

# 读取未映射的大鼠基因
unmapped_df = pd.read_excel(mapped_excel, sheet_name="未映射基因", skiprows=3, header=None)
unmapped_rat_genes = []
for val in unmapped_df[0]:
    if pd.notna(val) and str(val).strip() != '':
        unmapped_rat_genes.append(str(val).strip().upper())

print(f"需要映射的未映射大鼠基因数: {len(unmapped_rat_genes)}")

# 读取本地映射库建立完整映射
rat_to_human = {}
mapping_lines = open(mapping_file, 'r', encoding='utf-8').readlines()
header_line = None
for i, line in enumerate(mapping_lines):
    if line.startswith('RAT_GENE_SYMBOL'):
        header_line = i
        break

for line in mapping_lines[header_line+1:]:
    parts = line.strip().split('\t')
    if len(parts) >= 2:
        rat_gene = parts[0].strip().upper()
        human_ortholog = parts[1].strip().upper()
        
        if rat_gene and human_ortholog and human_ortholog != 'N/A':
            human_genes = [g.strip() for g in human_ortholog.split('|') if g.strip()]
            if rat_gene not in rat_to_human:
                rat_to_human[rat_gene] = []
            rat_to_human[rat_gene].extend(human_genes)

# 对映射去重
for rat_gene in rat_to_human:
    rat_to_human[rat_gene] = list(set(rat_to_human[rat_gene]))

print(f"本地映射库中的映射关系: {len(rat_to_human)}")

# 处理未映射基因
additional_mappings = []
mapping_details = []

for gene in unmapped_rat_genes:
    mapped = False
    
    # 1. 直接匹配本地映射库
    if gene in rat_to_human:
        human_genes = rat_to_human[gene]
        additional_mappings.extend(human_genes)
        for hg in human_genes:
            mapping_details.append({
                'Rat_Gene': gene,
                'Human_Gene': hg,
                'Method': 'Direct_Mapping'
            })
        mapped = True
        continue
    
    # 2. 尝试去除数字后缀（如Rpl3a -> RPL3）
    import re
    base_gene = re.sub(r'\d+[A-Z]?$', '', gene)
    if base_gene in rat_to_human:
        human_genes = rat_to_human[base_gene]
        additional_mappings.extend(human_genes)
        for hg in human_genes:
            mapping_details.append({
                'Rat_Gene': gene,
                'Human_Gene': hg,
                'Method': 'Base_Gene_Mapping'
            })
        mapped = True
        continue
    
    # 3. 尝试基因名保守性（假设大鼠和人类基因名相同）
    # 检查是否为常见的保守基因（非LOC/RGD开头）
    if not gene.startswith('LOC') and not gene.startswith('RGD') and not gene.startswith('MGC'):
        additional_mappings.append(gene)
        mapping_details.append({
            'Rat_Gene': gene,
            'Human_Gene': gene,
            'Method': 'Conserved_Gene_Name'
        })
        mapped = True
        continue
    
    # 4. 记录无法映射的基因
    if not mapped:
        mapping_details.append({
            'Rat_Gene': gene,
            'Human_Gene': 'N/A',
            'Method': 'Unmapped'
        })

# 去重
additional_mappings = list(set([g.upper() for g in additional_mappings]))
print(f"额外映射的人类基因数: {len(additional_mappings)}")

# 合并所有人类基因
all_human_genes = sorted(list(mapped_human_genes.union(set(additional_mappings))))
print(f"最终人类基因总数: {len(all_human_genes)}")

# 统计各方法映射数
method_counts = {}
for detail in mapping_details:
    method = detail['Method']
    method_counts[method] = method_counts.get(method, 0) + 1

print("\n映射方法统计:")
for method, count in method_counts.items():
    print(f"  {method}: {count}")

# 保存文本格式
with open(f"{work_dir}/PPI_Input_Genes_Human_Complete_Final.txt", 'w') as f:
    for gene in all_human_genes:
        f.write(gene + '\n')

# 创建Excel
wb = Workbook()
ws = wb.active
ws.title = "PPI输入基因 (Human完整)"

# 样式
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")
title_font = Font(bold=True, size=14, color="1F4E78")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 标题
ws['A1'] = "PPI网络输入基因 (完整人类基因符号)"
ws['A1'].font = title_font
ws.merge_cells('A1:D1')

# 统计信息
ws['A3'] = f"原始大鼠基因数: 4014"
ws['A4'] = f"本地库直接映射: {len(mapped_human_genes)}"
ws['A5'] = f"补充映射基因: {len(additional_mappings)}"
ws['A6'] = f"去重后人类基因总数: {len(all_human_genes)}"

# 映射方法统计
ws['A8'] = "映射方法统计:"
ws['A8'].font = Font(bold=True)
row = 9
for method, count in method_counts.items():
    ws.cell(row=row, column=1, value=method)
    ws.cell(row=row, column=2, value=count)
    row += 1

# 表头
gene_start_row = row + 2
headers = ['序号', '人类基因符号']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=gene_start_row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# 数据
for idx, gene in enumerate(all_human_genes, 1):
    ws.cell(row=gene_start_row + idx, column=1, value=idx).border = thin_border
    ws.cell(row=gene_start_row + idx, column=2, value=gene).border = thin_border

# 设置列宽
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 20

# 第二个sheet：映射详情
ws2 = wb.create_sheet("映射详情")
ws2['A1'] = "基因映射详情"
ws2['A1'].font = title_font

# 表头
detail_headers = ['大鼠基因', '人类基因', '映射方法']
for col, header in enumerate(detail_headers, 1):
    cell = ws2.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# 数据
for idx, detail in enumerate(mapping_details, 1):
    ws2.cell(row=3 + idx, column=1, value=detail['Rat_Gene']).border = thin_border
    ws2.cell(row=3 + idx, column=2, value=detail['Human_Gene']).border = thin_border
    ws2.cell(row=3 + idx, column=3, value=detail['Method']).border = thin_border

ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 20

# 保存
wb.save(output_file)
print(f"\n结果已保存:")
print(f"  - Excel格式: {output_file}")
print(f"  - 文本格式: {work_dir}/PPI_Input_Genes_Human_Complete_Final.txt")
