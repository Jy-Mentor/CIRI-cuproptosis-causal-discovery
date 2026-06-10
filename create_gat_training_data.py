#!/usr/bin/env python3
# ================================================================================
# MR 分析结果汇总 - GAT 训练数据准备
# 整合所有 MR 分析结果，生成适合图注意力网络 (GAT) 训练的结构化数据
# ================================================================================

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime

# ================================================================================
# 第一部分：数据加载函数
# ================================================================================

def load_mr_results():
    """加载 MR 主要结果"""
    mr_file = r"D:\下载\MR_batch_results\20260508_optimized_fixed_v2\MR_results_main_optimized.csv"
    if os.path.exists(mr_file):
        df = pd.read_csv(mr_file)
        print(f"✓ 加载 MR 结果：{len(df)} 个基因")
        return df
    else:
        print(f"✗ MR 结果文件不存在：{mr_file}")
        return None

def load_functional_enrichment():
    """加载功能富集分析结果"""
    reactome_file = r"D:\下载\MR_batch_results\20260508_optimized_fixed_v2\functional_enrichment\Reactome_results.csv"
    if os.path.exists(reactome_file):
        df = pd.read_csv(reactome_file)
        print(f"✓ 加载 Reactome 富集：{len(df)} 个通路")
        return df
    else:
        print("✗ Reactome 结果文件不存在")
        return None

def load_drug_targets():
    """加载药物靶点预测结果"""
    drug_file = r"D:\下载\MR_batch_results\20260508_optimized_fixed_v2\drug_targets\drug_targets_summary.csv"
    if os.path.exists(drug_file):
        df = pd.read_csv(drug_file)
        print(f"✓ 加载药物靶点：{len(df)} 个基因")
        return df
    else:
        print("✗ 药物靶点文件不存在")
        return None

def load_gene_details(gene_name):
    """加载单个基因的详细结果"""
    detail_file = r"D:\下载\MR_batch_results\20260508_optimized_fixed_v2\details\{}_detail.csv".format(gene_name)
    if os.path.exists(detail_file):
        return pd.read_csv(detail_file)
    return None

# ================================================================================
# 第二部分：特征工程函数
# ================================================================================

def extract_numerical_features(mr_row):
    """从 MR 结果中提取数值特征"""
    features = {}
    
    # MR 统计特征
    features['mr_beta'] = mr_row.get('discovery_b', np.nan)
    features['mr_se'] = mr_row.get('discovery_se', np.nan)
    features['mr_pval'] = mr_row.get('discovery_pval', np.nan)
    features['mr_pval_log10'] = -np.log10(mr_row.get('discovery_pval', 1)) if pd.notna(mr_row.get('discovery_pval')) else np.nan
    features['fdr_qval'] = mr_row.get('fdr_qval', np.nan)
    
    # 工具变量特征
    features['nsnp'] = mr_row.get('nsnp', 0)
    features['f_mean'] = mr_row.get('F_mean', np.nan)
    
    # 敏感性分析特征
    features['heterogeneity_p'] = mr_row.get('Q_p', np.nan)
    features['pleiotropy_p'] = mr_row.get('Egger_intercept_p', np.nan)
    features['steiger_p'] = mr_row.get('Steiger_p', np.nan)
    
    # 验证特征
    features['has_replication'] = 1 if mr_row.get('has_replication', False) else 0
    features['has_meta'] = 1 if mr_row.get('has_meta', False) else 0
    
    # 状态编码
    status = mr_row.get('status', 'UNKNOWN')
    features['is_success'] = 1 if status == 'SUCCESS' else 0
    features['is_fdr_sig'] = 1 if pd.notna(mr_row.get('fdr_sig')) and mr_row.get('fdr_sig') else 0
    
    return features

def create_gene_network_features(mr_df):
    """创建基因网络特征（用于 GAT 的边信息）"""
    # 这里可以基于通路共现、蛋白互作等构建网络
    # 简化版本：基于功能相似性
    
    network_edges = []
    
    # 显著基因之间建立连接
    significant_genes = mr_df[mr_df['fdr_sig'] == True]['gene'].tolist()
    
    # 创建完全连接（后续可以基于生物学知识修剪）
    for i, gene1 in enumerate(significant_genes):
        for gene2 in significant_genes[i+1:]:
            network_edges.append({
                'source': gene1,
                'target': gene2,
                'edge_type': 'co_significant',
                'weight': 1.0
            })
    
    return network_edges

def create_target_variable(mr_row):
    """创建 GAT 预测的目标变量"""
    # 多标签分类：基于 FDR 显著性和效应方向
    
    target = {
        'is_significant': 0,
        'effect_direction': 0,  # 0: 无显著，1: 保护，2: 风险
        'confidence_score': 0.0
    }
    
    if pd.notna(mr_row.get('fdr_qval')) and mr_row.get('fdr_qval') < 0.05:
        target['is_significant'] = 1
        
        if pd.notna(mr_row.get('discovery_b')):
            beta = mr_row.get('discovery_b')
            target['effect_direction'] = 1 if beta < 0 else 2
            
            # 置信度基于 P 值
            pval = mr_row.get('discovery_pval', 1)
            if pd.notna(pval) and pval > 0:
                target['confidence_score'] = min(1.0, -np.log10(pval) / 10)
    
    return target

# ================================================================================
# 第三部分：数据整合函数
# ================================================================================

def integrate_all_data():
    """整合所有数据源"""
    print("\n" + "="*70)
    print("整合 MR 分析数据用于 GAT 训练")
    print("="*70)
    
    # 1. 加载 MR 主要结果
    mr_df = load_mr_results()
    if mr_df is None or len(mr_df) == 0:
        print("错误：无法加载 MR 结果")
        return None, None, None
    
    # 2. 加载功能富集
    reactome_df = load_functional_enrichment()
    
    # 3. 加载药物靶点
    drug_df = load_drug_targets()
    
    # 4. 提取特征
    print("\n提取基因特征...")
    all_features = []
    all_targets = []
    
    for idx, row in mr_df.iterrows():
        gene = row.get('gene', 'Unknown')
        print(f"  处理基因 {idx+1}/{len(mr_df)}: {gene}")
        
        # 数值特征
        features = extract_numerical_features(row)
        features['gene'] = gene
        
        # 目标变量
        target = create_target_variable(row)
        target['gene'] = gene
        
        all_features.append(features)
        all_targets.append(target)
    
    # 转换为 DataFrame
    features_df = pd.DataFrame(all_features)
    targets_df = pd.DataFrame(all_targets)
    
    # 5. 创建网络边
    print("\n创建基因网络边...")
    edges = create_gene_network_features(mr_df)
    edges_df = pd.DataFrame(edges)
    
    print(f"\n✓ 特征矩阵：{features_df.shape[0]} 个基因 × {features_df.shape[1]} 个特征")
    print(f"✓ 目标变量：{targets_df.shape[0]} 个基因")
    print(f"✓ 网络边：{len(edges)} 条")
    
    return features_df, targets_df, edges_df

# ================================================================================
# 第四部分：Excel 文件创建
# ================================================================================

def create_gat_training_excel(features_df, targets_df, edges_df, output_file):
    """创建适合 GAT 训练的 Excel 文件"""
    
    print(f"\n创建 Excel 文件：{output_file}")
    
    # 创建工作簿
    wb = Workbook()
    
    # 删除默认 sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # ============================================================================
    # Sheet 1: 节点特征 (Node Features)
    # ============================================================================
    ws_features = wb.create_sheet(title="Node_Features")
    
    # 添加表头
    headers = ['gene_id'] + [col for col in features_df.columns if col != 'gene']
    ws_features.append(headers)
    
    # 添加数据
    for idx, row in features_df.iterrows():
        gene = row.get('gene', f'Gene_{idx}')
        feature_values = [gene] + [row[col] if pd.notna(row[col]) else 0 for col in features_df.columns if col != 'gene']
        ws_features.append(feature_values)
    
    # 格式化
    for cell in ws_features[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', start_color='FFFF00')
    
    # ============================================================================
    # Sheet 2: 目标变量 (Target Variables)
    # ============================================================================
    ws_targets = wb.create_sheet(title="Target_Variables")
    
    # 添加表头
    target_headers = ['gene_id', 'is_significant', 'effect_direction', 'confidence_score']
    ws_targets.append(target_headers)
    
    # 添加数据
    for idx, row in targets_df.iterrows():
        gene = row.get('gene', f'Gene_{idx}')
        ws_targets.append([
            gene,
            row.get('is_significant', 0),
            row.get('effect_direction', 0),
            round(row.get('confidence_score', 0), 4)
        ])
    
    # 格式化
    for cell in ws_targets[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', start_color='FFFF00')
    
    # ============================================================================
    # Sheet 3: 边列表 (Edge List)
    # ============================================================================
    ws_edges = wb.create_sheet(title="Edge_List")
    
    # 添加表头
    edge_headers = ['source_gene', 'target_gene', 'edge_type', 'weight']
    ws_edges.append(edge_headers)
    
    # 添加数据
    for edge in edges:
        ws_edges.append([
            edge['source'],
            edge['target'],
            edge['edge_type'],
            round(edge['weight'], 4)
        ])
    
    # 格式化
    for cell in ws_edges[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', start_color='FFFF00')
    
    # ============================================================================
    # Sheet 4: 数据统计 (Statistics)
    # ============================================================================
    ws_stats = wb.create_sheet(title="Statistics")
    
    stats_data = [
        ['数据集统计', '', ''],
        ['总基因数', len(features_df), ''],
        ['特征维度', features_df.shape[1] - 1, '(不包括 gene_id)'],
        ['显著基因数', targets_df['is_significant'].sum(), '(FDR < 0.05)'],
        ['保护效应基因', (targets_df['effect_direction'] == 1).sum(), '(beta < 0)'],
        ['风险效应基因', (targets_df['effect_direction'] == 2).sum(), '(beta > 0)'],
        ['网络边数', len(edges), ''],
        ['', '', ''],
        ['特征列表', '', ''],
    ]
    
    # 添加特征名称
    feature_names = [col for col in features_df.columns if col != 'gene']
    for i, fname in enumerate(feature_names, 1):
        stats_data.append([f'特征 {i}', fname, ''])
    
    for row_data in stats_data:
        ws_stats.append(row_data)
    
    # 格式化
    for cell in ws_stats['A1:C1']:
        cell.font = Font(bold=True, size=12)
    
    for cell in ws_stats['A9:C9']:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', start_color='FFFF00')
    
    # ============================================================================
    # Sheet 5: 元数据 (Metadata)
    # ============================================================================
    ws_metadata = wb.create_sheet(title="Metadata")
    
    metadata = [
        ['数据集信息', ''],
        ['创建日期', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['数据来源', 'MR 分析优化路线 B'],
        ['目标期刊', 'Nature Communications'],
        ['', ''],
        ['特征说明', ''],
        ['mr_beta', 'MR 分析的效应量 (beta)'],
        ['mr_se', 'MR 效应量的标准误'],
        ['mr_pval', 'MR 分析的 P 值'],
        ['mr_pval_log10', '-log10(P 值)'],
        ['fdr_qval', 'FDR 校正后的 q 值'],
        ['nsnp', '工具变量 (SNP) 数量'],
        ['f_mean', '平均 F 统计量'],
        ['heterogeneity_p', '异质性检验 P 值'],
        ['pleiotropy_p', '多效性检验 P 值'],
        ['steiger_p', 'Steiger 方向检验 P 值'],
        ['has_replication', '是否有独立验证 (0/1)'],
        ['has_meta', '是否有 Meta 分析 (0/1)'],
        ['is_success', '分析是否成功 (0/1)'],
        ['is_fdr_sig', '是否 FDR 显著 (0/1)'],
        ['', ''],
        ['目标变量说明', ''],
        ['is_significant', '是否显著 (0: 不显著，1: FDR<0.05)'],
        ['effect_direction', '效应方向 (0: 无，1: 保护，2: 风险)'],
        ['confidence_score', '置信度得分 (0-1)'],
        ['', ''],
        ['GAT 训练建议', ''],
        ['输入维度', features_df.shape[1] - 1],
        ['输出类别', '3 (无显著/保护/风险)'],
        ['图结构', '基于共显著性的边'],
        ['推荐隐藏层', '2-3 层'],
        ['推荐注意力头', '4-8 头'],
    ]
    
    for row_data in metadata:
        ws_metadata.append(row_data)
    
    # 格式化
    for cell in ws_metadata['A1:B1']:
        cell.font = Font(bold=True, size=12)
    
    for cell in ws_metadata['A6:B6']:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', start_color='FFFF00')
    
    for cell in ws_metadata['A21:B21']:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', start_color='FFFF00')
    
    for cell in ws_metadata['A25:B25']:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', start_color='FFFF00')
    
    # 调整列宽
    for col in ['A', 'B', 'C']:
        ws_metadata.column_dimensions[col].width = 25
    
    # 保存文件
    wb.save(output_file)
    print(f"✓ Excel 文件已保存：{output_file}")

# ================================================================================
# 第五部分：生成 PyTorch Geometric 数据加载代码
# ================================================================================

def generate_gat_training_code(output_dir):
    """生成 GAT 训练代码示例"""
    
    code = '''#!/usr/bin/env python3
"""
GAT 训练代码示例 - 基于 MR 分析结果
使用 PyTorch Geometric
"""

import torch
import torch.nn.functional as F
from torch_geometric.nn import GATConv
from torch_geometric.data import Data
import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.metrics import classification_report, accuracy_score

# ============================================================================
# 1. 数据加载
# ============================================================================

def load_gat_data(excel_file):
    """从 Excel 文件加载 GAT 训练数据"""
    
    # 加载特征
    features_df = pd.read_excel(excel_file, sheet_name='Node_Features')
    
    # 加载目标
    targets_df = pd.read_excel(excel_file, sheet_name='Target_Variables')
    
    # 加载边
    edges_df = pd.read_excel(excel_file, sheet_name='Edge_List')
    
    # 创建基因到索引的映射
    gene_list = features_df['gene_id'].tolist()
    gene_to_idx = {gene: idx for idx, gene in enumerate(gene_list)}
    
    # 特征矩阵
    feature_cols = [col for col in features_df.columns if col != 'gene_id']
    X = features_df[feature_cols].values
    X = np.nan_to_num(X, nan=0.0)  # 处理 NaN
    X = torch.FloatTensor(X)
    
    # 目标变量
    y = torch.LongTensor(targets_df['effect_direction'].values)
    
    # 边索引
    edge_index = []
    for _, row in edges_df.iterrows():
        src_idx = gene_to_idx.get(row['source_gene'])
        tgt_idx = gene_to_idx.get(row['target_gene'])
        if src_idx is not None and tgt_idx is not None:
            edge_index.append([src_idx, tgt_idx])
            edge_index.append([tgt_idx, src_idx])  # 无向图
    
    edge_index = torch.LongTensor(edge_index).t().contiguous()
    
    # 创建 PyTorch Geometric Data 对象
    data = Data(x=X, edge_index=edge_index, y=y)
    
    print(f"数据加载完成:")
    print(f"  节点数：{data.num_nodes}")
    print(f"  边数：{data.num_edges}")
    print(f"  特征维度：{data.num_node_features}")
    print(f"  类别数：{len(torch.unique(y))}")
    
    return data, gene_list, gene_to_idx

# ============================================================================
# 2. GAT 模型定义
# ============================================================================

class GAT(torch.nn.Module):
    def __init__(self, num_features, num_classes, hidden_channels=64, heads=4, dropout=0.6):
        super(GAT, self).__init__()
        
        # 第一层 GAT
        self.conv1 = GATConv(
            num_features, 
            hidden_channels, 
            heads=heads, 
            dropout=dropout,
            concat=True
        )
        
        # 第二层 GAT
        self.conv2 = GATConv(
            hidden_channels * heads, 
            num_classes, 
            heads=1, 
            dropout=dropout,
            concat=False
        )
        
        self.dropout = torch.nn.Dropout(dropout)
    
    def forward(self, data):
        x, edge_index = data.x, data.edge_index
        
        # 第一层
        x = self.conv1(x, edge_index)
        x = F.elu(x)
        x = self.dropout(x)
        
        # 第二层
        x = self.conv2(x, edge_index)
        
        return F.log_softmax(x, dim=1)

# ============================================================================
# 3. 训练函数
# ============================================================================

def train_gat(data, epochs=200, lr=0.01, weight_decay=5e-4):
    """训练 GAT 模型"""
    
    # 划分训练集和测试集
    train_mask, test_mask = train_test_split(
        np.arange(len(data.y)), 
        test_size=0.3, 
        random_state=42
    )
    
    train_mask = torch.zeros(len(data.y), dtype=torch.bool)
    test_mask = torch.zeros(len(data.y), dtype=torch.bool)
    train_mask[train_mask] = True
    test_mask[test_mask] = True
    
    # 创建模型
    model = GAT(
        num_features=data.num_node_features,
        num_classes=3,  # 无显著/保护/风险
        hidden_channels=64,
        heads=4,
        dropout=0.6
    )
    
    optimizer = torch.optim.Adam(model.parameters(), lr=lr, weight_decay=weight_decay)
    criterion = torch.nn.NLLLoss()
    
    # 训练循环
    print("\\n开始训练...")
    for epoch in range(epochs):
        model.train()
        optimizer.zero_grad()
        
        out = model(data)
        loss = criterion(out[train_mask], data.y[train_mask])
        loss.backward()
        optimizer.step()
        
        # 每 20 轮打印一次
        if (epoch + 1) % 20 == 0:
            model.eval()
            with torch.no_grad():
                pred = out.argmax(dim=1)
                train_acc = accuracy_score(data.y[train_mask], pred[train_mask])
                test_acc = accuracy_score(data.y[test_mask], pred[test_mask])
                
            print(f"Epoch {epoch+1:03d} | Loss: {loss.item():.4f} | "
                  f"Train Acc: {train_acc:.4f} | Test Acc: {test_acc:.4f}")
    
    # 最终评估
    model.eval()
    with torch.no_grad():
        out = model(data)
        pred = out.argmax(dim=1)
        
        print("\\n测试集分类报告:")
        print(classification_report(
            data.y[test_mask], 
            pred[test_mask],
            target_names=['无显著', '保护效应', '风险效应']
        ))
    
    return model

# ============================================================================
# 4. 主函数
# ============================================================================

if __name__ == "__main__":
    # 加载数据
    excel_file = "MR_GAT_Training_Data_20260508.xlsx"
    data, gene_list, gene_to_idx = load_gat_data(excel_file)
    
    # 训练模型
    model = train_gat(data, epochs=200, lr=0.01)
    
    # 预测显著基因
    print("\\n预测最可能的显著基因:")
    model.eval()
    with torch.no_grad():
        out = model(data)
        probs = torch.exp(out)
        
        # 按保护效应排序
        protective_probs = probs[:, 1]
        top_idx = torch.argsort(protective_probs, descending=True)[:10]
        
        print("\\nTop 10 保护效应基因:")
        for idx in top_idx:
            gene = gene_list[idx]
            prob = protective_probs[idx].item()
            print(f"  {gene}: {prob:.4f}")

'''
    
    code_file = os.path.join(output_dir, "gat_training_example.py")
    with open(code_file, 'w', encoding='utf-8') as f:
        f.write(code)
    
    print(f"✓ GAT 训练代码示例已保存：{code_file}")

# ================================================================================
# 第六部分：主函数
# ================================================================================

def main():
    """主函数"""
    print("="*70)
    print("MR 分析结果汇总 - GAT 训练数据准备")
    print("="*70)
    
    # 1. 整合所有数据
    features_df, targets_df, edges_df = integrate_all_data()
    
    if features_df is None:
        print("错误：数据整合失败")
        return
    
    # 2. 创建 Excel 文件
    output_dir = r"D:\下载\MR_batch_results\20260508_optimized_fixed_v2"
    output_file = os.path.join(output_dir, "MR_GAT_Training_Data_20260508.xlsx")
    
    create_gat_training_excel(features_df, targets_df, edges_df, output_file)
    
    # 3. 生成训练代码
    generate_gat_training_code(output_dir)
    
    # 4. 打印总结
    print("\n" + "="*70)
    print("数据汇总完成!")
    print("="*70)
    print(f"\n输出文件:")
    print(f"  1. Excel 数据：{output_file}")
    print(f"  2. 训练代码：{os.path.join(output_dir, 'gat_training_example.py')}")
    print(f"\n数据结构:")
    print(f"  - 节点特征：{features_df.shape[0]} 个基因 × {features_df.shape[1]-1} 个特征")
    print(f"  - 目标变量：{targets_df.shape[0]} 个基因")
    print(f"  - 网络边：{len(edges_df)} 条")
    print(f"\nGAT 训练建议:")
    print(f"  - 输入维度：{features_df.shape[1]-1}")
    print(f"  - 输出类别：3 (无显著/保护效应/风险效应)")
    print(f"  - 推荐架构：2 层 GAT + 4-8 注意力头")
    print(f"  - 训练轮数：200-500 epochs")
    print("="*70)

if __name__ == "__main__":
    main()
