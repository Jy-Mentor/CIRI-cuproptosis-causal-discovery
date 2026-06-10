# -*- coding: utf-8 -*-
"""
Create comprehensive Excel report from Stage7-9 results
"""
import os
import sys
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import RESULTS_DIR, BCP_TARGETS, CUPROPTOSIS_GENES, CUPROPTOSIS_RELATED

OUTPUT_DIR = os.path.join(RESULTS_DIR, "final_report")
os.makedirs(OUTPUT_DIR, exist_ok=True)

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
SUBTITLE_FONT = Font(bold=True, size=11)
SUBTITLE_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CUPRO_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")


def style_header(ws, row=1, max_col=None):
    if max_col is None:
        max_col = ws.max_column
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_table(ws, start_row=2):
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = THIN_BORDER


def auto_width(ws, max_width=35):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, max_width)


wb = Workbook()

# ============================================================
# Sheet 1: Executive Summary
# ============================================================
ws1 = wb.active
ws1.title = "项目概览"

ws1.cell(row=1, column=1, value="beta-Caryophyllene x Cuproptosis x CIRI Target Screening Report")
ws1.cell(row=1, column=1).font = Font(bold=True, size=16, color="1F4E78")
ws1.merge_cells("A1:F1")

overview = [
    ["Analysis Date", datetime.now().strftime("%Y-%m-%d")],
    ["Pipeline Version", "v5.0 (GSE174574 pseudo-bulk)"],
    ["Cell Type", "Microglia (all 58,340 cells)"],
    ["Sample Size", "6 independent samples (3 Sham + 3 MCAO)"],
    ["CV Strategy", "Leave-One-Out (LOO)"],
    ["Significance Test", "Welch's t-test (alternative to permutation for n<=10)"],
    ["BCP Targets (cleaned)", len(BCP_TARGETS)],
    ["Cuproptosis Core Genes", len(CUPROPTOSIS_GENES)],
    ["Cuproptosis Related Genes", len(CUPROPTOSIS_RELATED)],
]

row = 3
ws1.cell(row=row, column=1, value="Project Overview")
ws1.cell(row=row, column=1).font = SUBTITLE_FONT
ws1.cell(row=row, column=1).fill = SUBTITLE_FILL
ws1.merge_cells(f"A{row}:B{row}")
row += 1

for item, val in overview:
    ws1.cell(row=row, column=1, value=item).font = Font(bold=True)
    ws1.cell(row=row, column=1).border = THIN_BORDER
    ws1.cell(row=row, column=2, value=val).border = THIN_BORDER
    row += 1

row += 1
ws1.cell(row=row, column=1, value="Stage7 ML Performance")
ws1.cell(row=row, column=1).font = SUBTITLE_FONT
ws1.cell(row=row, column=1).fill = SUBTITLE_FILL
ws1.merge_cells(f"A{row}:B{row}")
row += 1

ml_perf = pd.read_csv(os.path.join(RESULTS_DIR, "stage7_ml_shap", "ml_model_performance.csv"))
for _, r in ml_perf.iterrows():
    ws1.cell(row=row, column=1, value=f"  {r['Model']} CV AUC").border = THIN_BORDER
    ws1.cell(row=row, column=2, value=r['CV_AUC']).border = THIN_BORDER
    row += 1
    ws1.cell(row=row, column=1, value=f"  {r['Model']} Welch's t-test P").border = THIN_BORDER
    ws1.cell(row=row, column=2, value=r['Test_P']).border = THIN_BORDER
    row += 1

row += 1
ws1.cell(row=row, column=1, value="Stage8 Target Classification")
ws1.cell(row=row, column=1).font = SUBTITLE_FONT
ws1.cell(row=row, column=1).fill = SUBTITLE_FILL
ws1.merge_cells(f"A{row}:B{row}")
row += 1

tier1 = pd.read_csv(os.path.join(RESULTS_DIR, "stage8_final_targets", "tier1_targets.csv"))
core = pd.read_csv(os.path.join(RESULTS_DIR, "stage8_final_targets", "core_targets.csv"))
tier_counts = core['Tier'].value_counts()
ws1.cell(row=row, column=1, value="  Tier1 targets").border = THIN_BORDER
ws1.cell(row=row, column=2, value=int(tier_counts.get('Tier1', 0))).border = THIN_BORDER
row += 1
ws1.cell(row=row, column=1, value="  Tier2 targets").border = THIN_BORDER
ws1.cell(row=row, column=2, value=int(tier_counts.get('Tier2', 0))).border = THIN_BORDER
row += 1
ws1.cell(row=row, column=1, value="  Tier3 targets").border = THIN_BORDER
ws1.cell(row=row, column=2, value=int(tier_counts.get('Tier3', 0))).border = THIN_BORDER
row += 1

row += 1
ws1.cell(row=row, column=1, value="Stage9 GAT Performance")
ws1.cell(row=row, column=1).font = SUBTITLE_FONT
ws1.cell(row=row, column=1).fill = SUBTITLE_FILL
ws1.merge_cells(f"A{row}:B{row}")
row += 1

import json
with open(os.path.join(RESULTS_DIR, "stage9_ppi_gat", "gat_performance.json"), 'r') as f:
    gat_perf = json.load(f)
ws1.cell(row=row, column=1, value="  R-squared").border = THIN_BORDER
ws1.cell(row=row, column=2, value=gat_perf['R2']).border = THIN_BORDER
row += 1
ws1.cell(row=row, column=1, value="  MSE").border = THIN_BORDER
ws1.cell(row=row, column=2, value=gat_perf['MSE']).border = THIN_BORDER
row += 1
ws1.cell(row=row, column=1, value="  Genes in PPI network").border = THIN_BORDER
ws1.cell(row=row, column=2, value=gat_perf['N_genes']).border = THIN_BORDER
row += 1
ws1.cell(row=row, column=1, value="  Edges").border = THIN_BORDER
ws1.cell(row=row, column=2, value=gat_perf['N_edges']).border = THIN_BORDER

ws1.column_dimensions['A'].width = 40
ws1.column_dimensions['B'].width = 25

# ============================================================
# Sheet 2: Stage7 ML Feature Importance
# ============================================================
ws2 = wb.create_sheet("Stage7_特征重要性")
shap = pd.read_csv(os.path.join(RESULTS_DIR, "stage7_ml_shap", "gene_shap_importance.csv"))
shap.to_excel(os.path.join(OUTPUT_DIR, "_temp_shap.xlsx"), index=False)

row = 1
ws2.cell(row=row, column=1, value="Stage7: ML Feature Importance (Nested CV LOO + Welch's t-test)")
ws2.cell(row=row, column=1).font = TITLE_FONT
ws2.merge_cells("A1:F1")
row += 2

headers = shap.columns.tolist()
for j, h in enumerate(headers, 1):
    ws2.cell(row=row, column=j, value=h)
style_header(ws2, row, len(headers))
row += 1

cupro_set = set(CUPROPTOSIS_GENES) | set(CUPROPTOSIS_RELATED)
for _, r in shap.iterrows():
    ws2.cell(row=row, column=1, value=r['Gene'])
    ws2.cell(row=row, column=2, value=r['SHAP_importance'])
    ws2.cell(row=row, column=3, value=r['Rank'])
    ws2.cell(row=row, column=4, value=r['L1_LogReg_score'])
    ws2.cell(row=row, column=5, value=r['Is_cuproptosis_core'])
    ws2.cell(row=row, column=6, value=r['Is_cuproptosis_related'])
    if r['Gene'] in cupro_set:
        for c in range(1, 7):
            ws2.cell(row=row, column=c).fill = CUPRO_FILL
    row += 1

style_data_table(ws2, 3)
auto_width(ws2)

# ============================================================
# Sheet 3: Stage8 Tier1 Targets
# ============================================================
ws3 = wb.create_sheet("Stage8_Tier1靶点")
row = 1
ws3.cell(row=row, column=1, value="Stage8: Tier1 Final Target Ranking")
ws3.cell(row=row, column=1).font = TITLE_FONT
ws3.merge_cells("A1:I1")
row += 2

tier1.to_excel(os.path.join(OUTPUT_DIR, "_temp_tier1.xlsx"), index=False)
for j, h in enumerate(tier1.columns, 1):
    ws3.cell(row=row, column=j, value=h)
style_header(ws3, row, len(tier1.columns))
row += 1

for _, r in tier1.iterrows():
    for j, h in enumerate(tier1.columns, 1):
        ws3.cell(row=row, column=j, value=r[h])
    row += 1

style_data_table(ws3, 3)
auto_width(ws3)

# ============================================================
# Sheet 4: Stage8 All Targets
# ============================================================
ws4 = wb.create_sheet("Stage8_全部排名")
row = 1
ws4.cell(row=row, column=1, value="Stage8: All Target Rankings (Top 100)")
ws4.cell(row=row, column=1).font = TITLE_FONT
ws4.merge_cells("A1:I1")
row += 2

core_top100 = core.head(100)
for j, h in enumerate(core_top100.columns, 1):
    ws4.cell(row=row, column=j, value=h)
style_header(ws4, row, len(core_top100.columns))
row += 1

for _, r in core_top100.iterrows():
    for j, h in enumerate(core_top100.columns, 1):
        ws4.cell(row=row, column=j, value=r[h])
    if r['Gene'] in cupro_set:
        for c in range(1, len(core_top100.columns) + 1):
            ws4.cell(row=row, column=c).fill = CUPRO_FILL
    row += 1

style_data_table(ws4, 3)
auto_width(ws4)

# ============================================================
# Sheet 5: Stage9 GAT Gene Ranking
# ============================================================
ws5 = wb.create_sheet("Stage9_GAT排名")
gat_ranking = pd.read_csv(os.path.join(RESULTS_DIR, "stage9_ppi_gat", "gat_gene_ranking.csv"))
row = 1
ws5.cell(row=row, column=1, value="Stage9: GAT Gene Ranking (Top 50)")
ws5.cell(row=row, column=1).font = TITLE_FONT
ws5.merge_cells("A1:D1")
row += 2

gat_top50 = gat_ranking.head(50)
for j, h in enumerate(gat_top50.columns, 1):
    ws5.cell(row=row, column=j, value=h)
style_header(ws5, row, len(gat_top50.columns))
row += 1

for _, r in gat_top50.iterrows():
    for j, h in enumerate(gat_top50.columns, 1):
        ws5.cell(row=row, column=j, value=r[h])
    if r['Gene'] in cupro_set:
        for c in range(1, len(gat_top50.columns) + 1):
            ws5.cell(row=row, column=c).fill = CUPRO_FILL
    row += 1

style_data_table(ws5, 3)
auto_width(ws5)

# ============================================================
# Sheet 6: Cuproptosis Core Gene Tracking
# ============================================================
ws6 = wb.create_sheet("铜死亡核心基因追踪")
row = 1
ws6.cell(row=row, column=1, value="Cuproptosis Core Gene Tracking Across All Stages")
ws6.cell(row=row, column=1).font = TITLE_FONT
ws6.merge_cells("A1:F1")
row += 2

cupro_headers = ['Gene', 'Stage7_Rank', 'Stage7_SHAP', 'Stage8_Tier', 'Stage8_Score', 'Stage9_Rank']
for j, h in enumerate(cupro_headers, 1):
    ws6.cell(row=row, column=j, value=h)
style_header(ws6, row, len(cupro_headers))
row += 1

shap_dict = dict(zip(shap['Gene'].str.upper(), zip(shap['Rank'], shap['SHAP_importance'])))
core_dict = dict(zip(core['Gene'].str.upper(), zip(core['Tier'], core['Comprehensive'])))
gat_dict = dict(zip(gat_ranking['Gene'].str.upper(), zip(gat_ranking['Rank'], gat_ranking['GAT_score'])))

for gene in CUPROPTOSIS_GENES:
    g = gene.upper()
    s7_rank, s7_shap = shap_dict.get(g, ('N/A', 0))
    s8_tier, s8_score = core_dict.get(g, ('N/A', 0))
    s9_rank, s9_score = gat_dict.get(g, ('N/A', 0))
    
    ws6.cell(row=row, column=1, value=g)
    ws6.cell(row=row, column=2, value=s7_rank)
    ws6.cell(row=row, column=3, value=s7_shap)
    ws6.cell(row=row, column=4, value=s8_tier)
    ws6.cell(row=row, column=5, value=s8_score)
    ws6.cell(row=row, column=6, value=s9_rank)
    row += 1

style_data_table(ws6, 3)
auto_width(ws6)

# ============================================================
# Save
# ============================================================
output_file = os.path.join(OUTPUT_DIR, f"BCP_CIRI_Report_{datetime.now().strftime('%Y%m%d')}.xlsx")
wb.save(output_file)
print(f"Report saved: {output_file}")

# Cleanup temp files
for f in ["_temp_shap.xlsx", "_temp_tier1.xlsx"]:
    fp = os.path.join(OUTPUT_DIR, f)
    if os.path.exists(fp):
        os.remove(fp)
