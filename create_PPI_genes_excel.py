import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 文件路径
ppi_file = "C:/Users/Jy-Mentor-7/Desktop/你信我 这不是重蹈覆辙/GSE61616_cluster_ssGSEA_PPI_results_v2/String_PPI_input_genes.txt"
mapping_file = "C:/Users/Jy-Mentor-7/Desktop/你信我 这不是重蹈覆辙/大创/大鼠 小鼠 人类映射库.txt"
output_file = "C:/Users/Jy-Mentor-7/Desktop/你信我 这不是重蹈覆辙/PPI_Input_Genes_Human.xlsx"

# 读取PPI输入基因（大鼠基因）
with open(ppi_file, 'r') as f:
    rat_genes = [line.strip().upper() for line in f if line.strip()]

print(f"PPI输入基因数（大鼠）: {len(rat_genes)}")

# 读取映射库，建立大鼠→人类映射
mapping_lines = open(mapping_file, 'r').readlines()
header_line = None
for i, line in enumerate(mapping_lines):
    if line.startswith('RAT_GENE_SYMBOL'):
        header_line = i
        break

# 解析映射文件
rat_to_human = {}
for line in mapping_lines[header_line+1:]:
    parts = line.strip().split('\t')
    if len(parts) >= 2:
        rat_gene = parts[0].strip().upper()
        human_ortholog = parts[1].strip().upper()
        
        if rat_gene and human_ortholog and human_ortholog != 'N/A':
            # 处理多个人类同源基因（用|分隔）
            human_genes = [g.strip() for g in human_ortholog.split('|') if g.strip()]
            if rat_gene not in rat_to_human:
                rat_to_human[rat_gene] = []
            rat_to_human[rat_gene].extend(human_genes)

# 大鼠基因映射到人类基因
results = []
mapped_count = 0
unmapped_genes = []

for rat_gene in rat_genes:
    if rat_gene in rat_to_human:
        human_genes = list(set(rat_to_human[rat_gene]))  # 去重
        for human_gene in human_genes:
            results.append({
                'Rat_Gene': rat_gene,
                'Human_Gene': human_gene,
                'Mapping_Status': 'Mapped'
            })
        mapped_count += 1
    else:
        results.append({
            'Rat_Gene': rat_gene,
            'Human_Gene': 'N/A',
            'Mapping_Status': 'Unmapped'
        })
        unmapped_genes.append(rat_gene)

# 创建DataFrame
df = pd.DataFrame(results)

# 去重（保留唯一的Human_Gene）
df_unique = df[df['Mapping_Status'] == 'Mapped'].drop_duplicates(subset=['Human_Gene'])

print(f"成功映射基因数: {mapped_count}")
print(f"未映射基因数: {len(unmapped_genes)}")
print(f"去重后人类基因数: {len(df_unique)}")

# 创建Excel
wb = Workbook()
ws = wb.active
ws.title = "PPI输入基因 (Human)"

# 样式定义
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")
title_font = Font(bold=True, size=14, color="1F4E78")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 标题
ws['A1'] = "PPI网络输入基因 (转换为人类基因符号)"
ws['A1'].font = title_font
ws.merge_cells('A1:C1')

# 统计信息
ws['A3'] = f"原始大鼠基因数: {len(rat_genes)}"
ws['A4'] = f"成功映射: {mapped_count}"
ws['A5'] = f"未映射: {len(unmapped_genes)}"
ws['A6'] = f"去重后人类基因数: {len(df_unique)}"

# 表头
headers = ['序号', '大鼠基因', '人类基因', '映射状态']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=8, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# 数据（只输出Mapped的去重结果）
for idx, (_, row) in enumerate(df_unique.iterrows(), 1):
    ws.cell(row=8+idx, column=1, value=idx).border = thin_border
    ws.cell(row=8+idx, column=2, value=row['Rat_Gene']).border = thin_border
    ws.cell(row=8+idx, column=3, value=row['Human_Gene']).border = thin_border
    ws.cell(row=8+idx, column=4, value=row['Mapping_Status']).border = thin_border

# 设置列宽
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 12

# 如果有未映射的基因，创建第二个sheet
if unmapped_genes:
    ws2 = wb.create_sheet("未映射基因")
    ws2['A1'] = "未映射的大鼠基因"
    ws2['A1'].font = title_font
    
    ws2['A3'] = "以下基因在映射库中未找到对应的人类同源基因:"
    
    for idx, gene in enumerate(unmapped_genes, 1):
        ws2.cell(row=3+idx, column=1, value=gene)
    
    ws2.column_dimensions['A'].width = 20

# 保存
wb.save(output_file)
print(f"\nExcel文件已保存: {output_file}")
print(f"包含 {len(df_unique)} 个人类基因，每行一个基因")
