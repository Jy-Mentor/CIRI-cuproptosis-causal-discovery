import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# 设置路径
work_dir = "C:/Users/Jy-Mentor-7/Desktop/你信我 这不是重蹈覆辙"
result_dir = os.path.join(work_dir, "String_Network_Connectivity_Results")
output_file = os.path.join(work_dir, "String_Network_Connectivity_Summary.xlsx")

# 创建workbook
wb = Workbook()
wb.remove(wb.active)

# 定义样式
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
title_font = Font(bold=True, size=14, color="1F4E78")
subtitle_font = Font(bold=True, size=11)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ==================== Sheet 1: 连通性诊断概述 ====================
ws_summary = wb.create_sheet("连通性诊断概述")

summary_data = [
    ["String PPI网络连通性诊断报告"],
    [""],
    ["基本网络统计"],
    ["总节点数", 142],
    ["总边数", 1944],
    [""],
    ["连通分量分析"],
    ["连通分量数", 9],
    ["最大连通分量节点数", 133],
    ["最大连通分量占比", "93.66%"],
    ["第二大连通分量节点数", 2],
    [""],
    ["路径长度分析（最大连通分量）"],
    ["最大连通分量节点数", 133],
    ["最大连通分量边数", 1942],
    ["平均最短路径长度", 2.518],
    ["网络直径", 7],
    [""],
    ["聚类系数"],
    ["全局聚类系数", 0.4974],
    ["平均局部聚类系数", 0.5615],
    [""],
    ["度分布统计"],
    ["平均度", 27.38],
    ["度中位数", 18],
    ["最大度", 134],
    ["最小度", 0],
    ["度标准差", 29.66],
    [""],
    ["孤立节点分析"],
    ["孤立节点数", 7],
    ["孤立节点列表", "BST1, CNR2, MGAT1, RENBP, SERPINB10, STK4, TCN2"],
    [""],
    ["网络鲁棒性分析"],
    ["平均边介数", 11.37],
    ["最大边介数", 131],
    ["高介数边数 (>95%分位数)", 98],
]

for row_idx, row_data in enumerate(summary_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.font = Font(bold=True, size=16, color="1F4E78")
            cell.alignment = Alignment(horizontal="center")
        elif value in ["基本网络统计", "连通分量分析", "路径长度分析（最大连通分量）", 
                       "聚类系数", "度分布统计", "孤立节点分析", "网络鲁棒性分析"]:
            cell.font = subtitle_font
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

ws_summary.column_dimensions['A'].width = 30
ws_summary.column_dimensions['B'].width = 20

# ==================== Sheet 2: 连通分量统计 ====================
ws_component = wb.create_sheet("连通分量统计")
component_df = pd.read_csv(os.path.join(result_dir, "component_statistics.txt"), sep="\t")

ws_component['A1'] = "连通分量统计"
ws_component['A1'].font = title_font

for r_idx, row in enumerate(dataframe_to_rows(component_df, index=False, header=True), 2):
    for c_idx, value in enumerate(row, 1):
        cell = ws_component.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        if r_idx == 2:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

ws_component.column_dimensions['A'].width = 15
ws_component.column_dimensions['B'].width = 15

# ==================== Sheet 3: 中心性指标汇总 ====================
ws_centrality = wb.create_sheet("中心性指标汇总")
centrality_df = pd.read_csv(os.path.join(result_dir, "centrality_measures.txt"), sep="\t")

ws_centrality['A1'] = "网络中心性指标汇总"
ws_centrality['A1'].font = title_font

for r_idx, row in enumerate(dataframe_to_rows(centrality_df, index=False, header=True), 2):
    for c_idx, value in enumerate(row, 1):
        cell = ws_centrality.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        if r_idx == 2:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_centrality.column_dimensions[col].width = 15

# ==================== Sheet 4: 孤立节点列表 ====================
ws_isolated = wb.create_sheet("孤立节点列表")
isolated_df = pd.read_csv(os.path.join(result_dir, "isolated_nodes.txt"), sep="\t")

ws_isolated['A1'] = "孤立节点列表"
ws_isolated['A1'].font = title_font

ws_isolated['A2'] = "以下节点在网络中没有连接边:"

for r_idx, row in enumerate(dataframe_to_rows(isolated_df, index=False, header=True), 3):
    for c_idx, value in enumerate(row, 1):
        cell = ws_isolated.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        if r_idx == 3:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

ws_isolated.column_dimensions['A'].width = 20

# ==================== Sheet 5: 连通分量成员分配 ====================
ws_membership = wb.create_sheet("连通分量成员分配")
membership_df = pd.read_csv(os.path.join(result_dir, "component_membership.txt"), sep="\t")

ws_membership['A1'] = "节点连通分量成员分配"
ws_membership['A1'].font = title_font

for r_idx, row in enumerate(dataframe_to_rows(membership_df, index=False, header=True), 2):
    for c_idx, value in enumerate(row, 1):
        cell = ws_membership.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        if r_idx == 2:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

ws_membership.column_dimensions['A'].width = 15
ws_membership.column_dimensions['B'].width = 15

# ==================== Sheet 6: 中心性指标Top 10 ====================
ws_top10 = wb.create_sheet("中心性指标Top10")

ws_top10['A1'] = "网络中心性指标 Top 10"
ws_top10['A1'].font = title_font

# 度中心性 Top 10
top_degree = ['AKT1', 'TNF', 'IL6', 'EGFR', 'HIF1A', 'NFKB1', 'STAT3', 'TGFB1', 'PPARG', 'CCL2']
degree_values = [134, 128, 126, 104, 96, 96, 96, 96, 94, 90]

ws_top10['A3'] = "度中心性 Top 10"
ws_top10['A3'].font = subtitle_font

headers = ['排名', '基因', '度值']
for col, header in enumerate(headers, 1):
    cell = ws_top10.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

for i in range(10):
    ws_top10.cell(row=5+i, column=1, value=i+1).border = thin_border
    ws_top10.cell(row=5+i, column=2, value=top_degree[i]).border = thin_border
    ws_top10.cell(row=5+i, column=3, value=degree_values[i]).border = thin_border

# 介数中心性 Top 10
betweenness_data = [
    ('AKT1', 0.1670), ('TNF', 0.0893), ('EGFR', 0.0770), ('PPARG', 0.0756),
    ('IL6', 0.0670), ('MAOB', 0.0632), ('STAT5A', 0.0581), ('PARP1', 0.0569),
    ('HIF1A', 0.0453), ('HMOX1', 0.0343)
]

ws_top10['A17'] = "介数中心性 Top 10"
ws_top10['A17'].font = subtitle_font

headers2 = ['排名', '基因', '介数中心性']
for col, header in enumerate(headers2, 1):
    cell = ws_top10.cell(row=18, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

for i, (gene, val) in enumerate(betweenness_data):
    ws_top10.cell(row=19+i, column=1, value=i+1).border = thin_border
    ws_top10.cell(row=19+i, column=2, value=gene).border = thin_border
    ws_top10.cell(row=19+i, column=3, value=round(val, 4)).border = thin_border

# 接近中心性 Top 10
closeness_data = [
    ('AKT1', 0.6316), ('TNF', 0.6139), ('IL6', 0.6055), ('EGFR', 0.5593),
    ('HIF1A', 0.5546), ('PPARG', 0.5546), ('STAT3', 0.5546), ('NFKB1', 0.5432),
    ('TGFB1', 0.5410), ('CCL2', 0.5344)
]

ws_top10['A31'] = "接近中心性 Top 10"
ws_top10['A31'].font = subtitle_font

headers3 = ['排名', '基因', '接近中心性']
for col, header in enumerate(headers3, 1):
    cell = ws_top10.cell(row=32, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

for i, (gene, val) in enumerate(closeness_data):
    ws_top10.cell(row=33+i, column=1, value=i+1).border = thin_border
    ws_top10.cell(row=33+i, column=2, value=gene).border = thin_border
    ws_top10.cell(row=33+i, column=3, value=round(val, 4)).border = thin_border

# 特征向量中心性 Top 10
eigen_data = [
    ('IL6', 1.0000), ('TNF', 0.9933), ('AKT1', 0.9813), ('NFKB1', 0.9020),
    ('STAT3', 0.8922), ('TGFB1', 0.8742), ('HIF1A', 0.8465), ('EGFR', 0.8420),
    ('CCL2', 0.8374), ('PPARG', 0.8059)
]

ws_top10['A45'] = "特征向量中心性 Top 10"
ws_top10['A45'].font = subtitle_font

headers4 = ['排名', '基因', '特征向量中心性']
for col, header in enumerate(headers4, 1):
    cell = ws_top10.cell(row=46, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

for i, (gene, val) in enumerate(eigen_data):
    ws_top10.cell(row=47+i, column=1, value=i+1).border = thin_border
    ws_top10.cell(row=47+i, column=2, value=gene).border = thin_border
    ws_top10.cell(row=47+i, column=3, value=round(val, 4)).border = thin_border

for col in ['A', 'B', 'C']:
    ws_top10.column_dimensions[col].width = 15

# 保存文件
wb.save(output_file)
print(f"String网络连通性诊断汇总Excel已生成: {output_file}")
print("包含以下sheet:")
for sheet in wb.sheetnames:
    print(f"  - {sheet}")
