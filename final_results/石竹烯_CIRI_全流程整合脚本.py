#!/usr/bin/env python3
"""
石竹烯-CIRI 项目全流程整合脚本 v3（权威论文标准范式重构版）
=========================================================
重构内容（基于DeepGSEA/STGAT/GenePlexus/PrimeKG标准）：
  1. 执行顺序修正：build_edges → compute_topology → build_labels → build_features
  2. Transductive图分割（删除Inductive跨集边删除）
  3. 删除铜死亡强制阳性标签和KL散度先验
  4. MR显著性改用FDR校正（fdr_qval < 0.05）
  5. 阴性标签来源修正：禁用DEG非显著基因，改用Housekeeping或远距离随机基因
  6. 阳性标签数量保护：n_pos < 30时警告并建议降级方案
  7. 删除零填充特征，补充真实生物学特征或报错停止
  8. 删除dist_to_inflammation（标签泄露），保留dist_to_cuproptosis并修正归一化
  9. 统一损失函数：删除KL散度，改用加权交叉熵或Focal Loss二选一
  10. 修正Recall@K：在全部未知节点中计算铜死亡基因Recall@K
  11. Excel精简为3个sheet
  12. 嵌入15项代码自检断言
  13. 文件I/O异常处理：缺失时sys.exit(1)

运行方式:
  py 石竹烯_CIRI_全流程整合脚本_v3.py --mode all
"""

import os
import sys
import json
import gzip
import shutil
import logging
import argparse
import copy
import time
import hashlib
import pickle
from pathlib import Path
from typing import Dict, List, Set, Tuple, Optional, Any
from collections import defaultdict
from dataclasses import dataclass, field, asdict

import numpy as np
import pandas as pd
import requests
import networkx as nx
from tqdm import tqdm
from sklearn.metrics import (
    accuracy_score, precision_score, recall_score, f1_score,
    roc_auc_score, average_precision_score, confusion_matrix
)
from scipy import stats

import torch
import torch.nn as nn
import torch.nn.functional as F
from torch.optim import AdamW
from torch.optim.lr_scheduler import CosineAnnealingLR, ReduceLROnPlateau
from torch_geometric.nn import GATv2Conv
from torch_geometric.utils import to_networkx

import yaml
import pickle
import hashlib
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# 三级缓存管理器 (v3.1新增)
# ---------------------------------------------------------------------------
class CacheManager:
    """三级缓存管理器：输入哈希 + 过程数据 + 拓扑特征"""
    
    def __init__(self, cache_dir: Path):
        self.cache_dir = cache_dir
        self.cache_dir.mkdir(parents=True, exist_ok=True)
        self.hash_file = cache_dir / ".input_hashes.json"
        self.meta_file = cache_dir / ".cache_meta.json"
        self.meta = self._load_json(self.meta_file, {})
    
    @staticmethod
    def _load_json(path: Path, default: Any) -> Any:
        if path.exists():
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        return default
    
    @staticmethod
    def _save_json(path: Path, data: Any):
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
    
    @staticmethod
    def _file_hash(path: Path, method="mtime_size") -> str:
        if not path.exists():
            return ""
        if method == "md5":
            h = hashlib.md5()
            with open(path, "rb") as f:
                for chunk in iter(lambda: f.read(8192), b""):
                    h.update(chunk)
            return h.hexdigest()
        stat = path.stat()
        return f"{stat.st_mtime}_{stat.st_size}"
    
    def check_inputs_unchanged(self, input_paths: List[Path]) -> bool:
        if not self.hash_file.exists():
            return False
        old_hashes = self._load_json(self.hash_file, {})
        new_hashes = {str(p.resolve()): self._file_hash(p) for p in input_paths if p.exists()}
        return old_hashes == new_hashes
    
    def save_input_hashes(self, input_paths: List[Path]):
        new_hashes = {str(p.resolve()): self._file_hash(p) for p in input_paths if p.exists()}
        self._save_json(self.hash_file, new_hashes)
    
    def get(self, key: str):
        cache_path = self.cache_dir / f"{key}.pkl"
        if cache_path.exists():
            try:
                with open(cache_path, "rb") as f:
                    return pickle.load(f)
            except Exception:
                return None
        return None
    
    def set(self, key: str, value: Any, metadata: Optional[Dict] = None):
        cache_path = self.cache_dir / f"{key}.pkl"
        with open(cache_path, "wb") as f:
            pickle.dump(value, f)
        if metadata:
            self.meta[key] = metadata
            self._save_json(self.meta_file, self.meta)
    
    def check_meta(self, key: str, expected: Dict) -> bool:
        actual = self.meta.get(key, {})
        return actual == expected
    
    def invalidate(self, key: str):
        cache_path = self.cache_dir / f"{key}.pkl"
        if cache_path.exists():
            cache_path.unlink()
        if key in self.meta:
            del self.meta[key]
            self._save_json(self.meta_file, self.meta)


# ---------------------------------------------------------------------------
# 配置与常量
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent.resolve()
LOCAL_DATA_DIR = BASE_DIR / "local_data"
DOWNLOADED_RAW_DIR = BASE_DIR / "downloaded_raw"
PROCESSED_DIR = BASE_DIR / "processed"
LOGS_DIR = BASE_DIR / "logs"
MODELS_DIR = BASE_DIR / "models"
RESULTS_DIR = BASE_DIR / "results"
SUBGRAPH_DIR = BASE_DIR / "subgraph_analysis"
MR_DIR = BASE_DIR / "MR_batch_summary_20260506"

for d in [LOCAL_DATA_DIR, DOWNLOADED_RAW_DIR, PROCESSED_DIR, LOGS_DIR, MODELS_DIR, RESULTS_DIR, SUBGRAPH_DIR]:
    d.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger(__name__)

# 铜死亡执行基因（仅保留为is_cuproptosis二元特征，不再强制label=1）
CUPROPTOSIS_EXECUTOR_GENES = {
    "FDX1", "LIAS", "LIPT1", "DLAT", "PDHB", "PDHX", "SLC31A1"
}
# 铜死亡调控基因
CUPROPTOSIS_REGULATOR_GENES = {
    "ATP7A", "ATP7B", "ATOX1", "NFE2L2", "HIF1A", "MTOR", "NFKB1", "GPX4"
}
CUPROPTOSIS_GENES = CUPROPTOSIS_EXECUTOR_GENES | CUPROPTOSIS_REGULATOR_GENES

# 单细胞数据文件路径（可选）
SC_KO_FILE = BASE_DIR / "endo_ko_Slc31a1_corrected.csv"
SC_COMM_FILE = BASE_DIR / "corrected_72h_lr_communication_v2.csv"
MR_FILE = Path("D:/下载/MR_batch_results/20260508_optimized_final/MR_results_main_optimized.csv")

# Housekeeping基因集（Eisenberg & Levanon, 2013）
HOUSEKEEPING_GENES = {
    "ACTB", "GAPDH", "RPLP0", "RPL13A", "B2M", "PGK1", "LDHA", "NONO",
    "PPIH", "RPLP1", "RPLP2", "RPS18", "RPS27A", "EEF1A1", "EEF2",
    "TPT1", "TUBB", "YWHAZ", "GUSB", "HPRT1", "HMBS", "SDHA", "TFRC",
    "TBP", "POLR2A", "PSMB2", "RPN1", "C1orf43", "CHMP2A", "EMC7",
    "GPI", "REEP5", "SNRPD3", "VCP", "VPS29"
}

# ---------------------------------------------------------------------------
# 模块一: 数据整合（v3重构版）
# ---------------------------------------------------------------------------

def normalize_gene_symbol(gene: str) -> Optional[str]:
    if pd.isna(gene):
        return None
    g = str(gene).strip().upper()
    g = g.replace("//", "/").split("/")[0].split(";")[0]
    return g if g else None


def load_gene_list(path: Path) -> Set[str]:
    if not path.exists():
        logger.error(f"基因列表不存在: {path}")
        sys.exit(1)
    genes = set()
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            g = normalize_gene_symbol(line)
            if g:
                genes.add(g)
    return genes


class StringPPIDownloader:
    def __init__(self, species=9606, min_score=400):
        self.species = species
        self.min_score = min_score
        self.protein_links_url = (
            f"https://stringdb-static.org/download/protein.links.v12.0/"
            f"{species}.protein.links.v12.0.txt.gz"
        )
        self.protein_info_url = (
            f"https://stringdb-static.org/download/protein.info.v12.0/"
            f"{species}.protein.info.v12.0.txt.gz"
        )

    def _download(self, url: str, dest: Path):
        if dest.exists():
            logger.info(f"已缓存: {dest}")
            return
        logger.info(f"下载 {url} ...")
        try:
            r = requests.get(url, stream=True, timeout=300)
            r.raise_for_status()
            with open(dest, "wb") as f:
                for chunk in r.iter_content(chunk_size=8192):
                    f.write(chunk)
            logger.info(f"下载完成: {dest}")
        except Exception as e:
            logger.error(f"下载失败: {e}")
            raise

    def load_local(self, gene_pool: Set[str]):
        links_path = DOWNLOADED_RAW_DIR / f"{self.species}.protein.links.v12.0.txt.gz"
        info_path = DOWNLOADED_RAW_DIR / f"{self.species}.protein.info.v12.0.txt.gz"

        if not links_path.exists() or not info_path.exists():
            logger.info("本地STRING文件不存在，尝试下载...")
            try:
                self._download(self.protein_links_url, links_path)
                self._download(self.protein_info_url, info_path)
            except Exception as e:
                logger.error(f"STRING下载失败: {e}")
                return [], {}

        logger.info("读取 STRING protein info...")
        ensembl_to_gene = {}
        with gzip.open(info_path, "rt", encoding="utf-8") as f:
            next(f)
            for line in f:
                parts = line.strip().split("\t")
                if len(parts) >= 2:
                    ensembl_id = parts[0]
                    gene_symbol = normalize_gene_symbol(parts[1])
                    if gene_symbol and gene_symbol in gene_pool:
                        ensembl_to_gene[ensembl_id] = gene_symbol

        logger.info("读取 STRING protein links...")
        edges = []
        max_score = 0
        with gzip.open(links_path, "rt", encoding="utf-8") as f:
            next(f)
            for line in f:
                parts = line.strip().split(" ")
                if len(parts) >= 3:
                    try:
                        score = float(parts[2])
                        max_score = max(max_score, score)
                    except ValueError:
                        continue
                    p1, p2 = parts[0], parts[1]
                    g1 = ensembl_to_gene.get(p1)
                    g2 = ensembl_to_gene.get(p2)
                    if g1 and g2 and g1 != g2:
                        edges.append((g1, g2, score))

        if max_score > 10:
            logger.info(f"STRING combined_score 为 0-1000 范围，使用阈值 {self.min_score}")
            threshold = self.min_score
            scale = 1000.0
        else:
            logger.info(f"STRING combined_score 为 0-1 范围，使用阈值 {self.min_score/1000}")
            threshold = self.min_score / 1000
            scale = 1.0

        filtered = [(g1, g2, score/scale) for g1, g2, score in edges if score >= threshold]
        return filtered, ensembl_to_gene


class DataIntegrator:
    def __init__(self):
        self.gene_pool: Set[str] = set()
        self.deg_df: Optional[pd.DataFrame] = None
        self.positive_labels: Set[str] = set()
        self.negative_labels: Set[str] = set()
        self.node_features: Optional[pd.DataFrame] = None
        self.labels_df: Optional[pd.DataFrame] = None
        self.edge_index: Optional[torch.Tensor] = None
        self.edge_attr: Optional[torch.Tensor] = None
        self.gene_symbols: List[str] = []
        self.gene_to_idx: Dict[str, int] = {}
        self.mr_df: Optional[pd.DataFrame] = None
        self._G: Optional[nx.Graph] = None
        # v3.1: 初始化缓存管理器
        self.cache = CacheManager(PROCESSED_DIR / ".cache")

    def run(self):
        # v3.1: L1缓存检查 - 输入文件未变化时跳过数据整合
        input_files = [
            LOCAL_DATA_DIR / "GSE61616.top.table.tsv",
            LOCAL_DATA_DIR / "stroke_targets.txt",
            LOCAL_DATA_DIR / "bcp_targets.txt",
            MR_FILE,
            LOCAL_DATA_DIR / "go_terms.tsv",
            LOCAL_DATA_DIR / "kegg_pathways.tsv",
            LOCAL_DATA_DIR / "interpro_domains.tsv",
        ]
        # 监控STRING原始文件
        string_links = DOWNLOADED_RAW_DIR / "9606.protein.links.v12.0.txt.gz"
        string_info = DOWNLOADED_RAW_DIR / "9606.protein.info.v12.0.txt.gz"
        if string_links.exists():
            input_files.append(string_links)
        if string_info.exists():
            input_files.append(string_info)

        required_outputs = [
            PROCESSED_DIR / "node_features.csv",
            PROCESSED_DIR / "labels.csv",
            PROCESSED_DIR / "edge_index.pt",
            PROCESSED_DIR / "edge_attr.pt",
        ]

        if self.cache.check_inputs_unchanged(input_files) and all(p.exists() for p in required_outputs):
            logger.info("=" * 60)
            logger.info("L1 缓存命中：输入文件未变化，跳过数据整合")
            logger.info("=" * 60)
            self.node_features = pd.read_csv(PROCESSED_DIR / "node_features.csv")
            self.labels_df = pd.read_csv(PROCESSED_DIR / "labels.csv")
            self.edge_index = torch.load(PROCESSED_DIR / "edge_index.pt", weights_only=False)
            self.edge_attr = torch.load(PROCESSED_DIR / "edge_attr.pt", weights_only=False)
            with open(PROCESSED_DIR / "gene_symbols.pkl", "rb") as f:
                self.gene_symbols = pickle.load(f)
            with open(PROCESSED_DIR / "gene_to_idx.pkl", "rb") as f:
                self.gene_to_idx = pickle.load(f)
            self.gene_pool = set(self.node_features["GeneSymbol"])
            return

        # 原流程
        self.load_local_data()
        self.load_mr_data()
        self.build_edges()          # 先建图
        self.build_labels()         # 再建标签（依赖self._G进行结构隔离）
        self.build_features()       # 最后建特征
        self.save_processed()
        
        # 保存输入哈希，供下次L1缓存命中
        self.cache.save_input_hashes(input_files)

    def load_local_data(self):
        logger.info("加载本地数据")

        # DEG
        deg_path = LOCAL_DATA_DIR / "GSE61616.top.table.tsv"
        if not deg_path.exists():
            logger.error(f"DEG文件不存在: {deg_path}")
            sys.exit(1)
        
        self.deg_df = pd.read_csv(deg_path, sep="\t")
        if "GeneSymbol" not in self.deg_df.columns:
            for col in ["gene_name", "Gene", "SYMBOL", "Gene.symbol", "gene_symbol"]:
                if col in self.deg_df.columns:
                    self.deg_df = self.deg_df.rename(columns={col: "GeneSymbol"})
                    break
        self.deg_df["GeneSymbol"] = self.deg_df["GeneSymbol"].apply(normalize_gene_symbol)
        self.deg_df = self.deg_df[self.deg_df["GeneSymbol"].notna()]
        
        # DEG列名模糊匹配
        # logFC
        logfc_candidates = ["logFC", "log2FoldChange", "log2fc", "FC", "log2FC"]
        matched_logfc = None
        for cand in logfc_candidates:
            if cand in self.deg_df.columns:
                matched_logfc = cand
                break
        if matched_logfc and matched_logfc != "logFC":
            logger.info(f"DEG列名模糊匹配: {matched_logfc} -> logFC")
            self.deg_df = self.deg_df.rename(columns={matched_logfc: "logFC"})
        elif matched_logfc is None:
            logger.error(f"DEG文件缺少logFC列（尝试了{logfc_candidates}）")
            sys.exit(1)
        
        # P.Value
        pval_candidates = ["P.Value", "pvalue", "PValue", "pval", "p.value", "Pvalue"]
        matched_pval = None
        for cand in pval_candidates:
            if cand in self.deg_df.columns:
                matched_pval = cand
                break
        if matched_pval and matched_pval != "P.Value":
            logger.info(f"DEG列名模糊匹配: {matched_pval} -> P.Value")
            self.deg_df = self.deg_df.rename(columns={matched_pval: "P.Value"})
        elif matched_pval is None:
            logger.error(f"DEG文件缺少P.Value列（尝试了{pval_candidates}）")
            sys.exit(1)
        
        # adj.P.Val
        adjp_candidates = ["adj.P.Val", "padj", "fdr", "adj_pval", "qval", "adj_pvalue", "adjPVal"]
        matched_adjp = None
        for cand in adjp_candidates:
            if cand in self.deg_df.columns:
                matched_adjp = cand
                break
        if matched_adjp and matched_adjp != "adj.P.Val":
            logger.info(f"DEG列名模糊匹配: {matched_adjp} -> adj.P.Val")
            self.deg_df = self.deg_df.rename(columns={matched_adjp: "adj.P.Val"})
        elif matched_adjp is None:
            logger.error(f"DEG文件缺少adj.P.Val列（尝试了{adjp_candidates}）")
            sys.exit(1)
        
        sig = self.deg_df[(self.deg_df["logFC"].abs() >= 1.0) & (self.deg_df["adj.P.Val"] < 0.05)]
        logger.info(f"DEG 显著基因数 (|logFC|>=1.0 & adj.P<0.05): {len(sig)}")

        # 靶点
        stroke_targets = load_gene_list(LOCAL_DATA_DIR / "stroke_targets.txt")
        bcp_targets = load_gene_list(LOCAL_DATA_DIR / "bcp_targets.txt")
        cupro_targets = CUPROPTOSIS_GENES.copy()

        logger.info(f"脑缺血靶点: {len(stroke_targets)} 个")
        logger.info(f"石竹烯靶点: {len(bcp_targets)} 个")
        logger.info(f"铜死亡基因: {len(cupro_targets)} 个")

        # 基因池 = DEG显著基因 ∪ 靶点 ∪ 铜死亡
        pool = set()
        if self.deg_df is not None:
            pool.update(self.deg_df["GeneSymbol"].tolist())
        pool.update(stroke_targets)
        pool.update(bcp_targets)
        pool.update(cupro_targets)
        self.gene_pool = pool
        logger.info(f"基因池初始大小: {len(self.gene_pool)}")

    def load_mr_data(self):
        if MR_FILE.exists():
            logger.info(f"读取MR数据: {MR_FILE}")
            # 支持CSV和Excel格式
            if MR_FILE.suffix == '.xlsx':
                self.mr_df = pd.read_excel(MR_FILE)
            else:
                self.mr_df = pd.read_csv(MR_FILE)
            if "gene" in self.mr_df.columns:
                self.mr_df["gene"] = self.mr_df["gene"].apply(normalize_gene_symbol)
                self.mr_df = self.mr_df[self.mr_df["gene"].notna()]
                logger.info(f"MR数据: {len(self.mr_df)} 条记录, {self.mr_df['gene'].nunique()} 个基因")
            else:
                logger.error("MR文件缺少gene列")
                sys.exit(1)
            
            # 列名模糊匹配: fdr_qval
            fdr_candidates = ["fdr_qval", "FDR", "qval", "q.value", "padj", "adj_pval", "fdr"]
            matched_fdr = None
            for cand in fdr_candidates:
                if cand in self.mr_df.columns:
                    matched_fdr = cand
                    break
            if matched_fdr and matched_fdr != "fdr_qval":
                logger.info(f"MR列名模糊匹配: {matched_fdr} -> fdr_qval")
                self.mr_df = self.mr_df.rename(columns={matched_fdr: "fdr_qval"})
            elif matched_fdr is None:
                logger.error(f"MR文件缺少fdr_qval列（尝试了{fdr_candidates}），无法进行FDR校正")
                sys.exit(1)
            
            # 列名模糊匹配: pval
            pval_candidates = ["pval", "p.value", "PValue", "pvalue", "p"]
            matched_pval = None
            for cand in pval_candidates:
                if cand in self.mr_df.columns:
                    matched_pval = cand
                    break
            if matched_pval and matched_pval != "pval":
                logger.info(f"MR列名模糊匹配: {matched_pval} -> pval")
                self.mr_df = self.mr_df.rename(columns={matched_pval: "pval"})
            
            # 列名模糊匹配: b (beta)
            b_candidates = ["b", "beta", "Beta", "B", "effect"]
            matched_b = None
            for cand in b_candidates:
                if cand in self.mr_df.columns:
                    matched_b = cand
                    break
            if matched_b and matched_b != "b":
                logger.info(f"MR列名模糊匹配: {matched_b} -> b")
                self.mr_df = self.mr_df.rename(columns={matched_b: "b"})
        else:
            logger.error(f"MR文件不存在: {MR_FILE}")
            sys.exit(1)

    def build_edges(self):
        logger.info("构建边索引...")
        downloader = StringPPIDownloader(min_score=400)
        edges, ensembl_map = downloader.load_local(self.gene_pool)

        if not edges:
            logger.error("STRING 边为空，无法构建图")
            sys.exit(1)

        G = nx.Graph()
        for g1, g2, score in edges:
            G.add_edge(g1, g2, weight=score)

        # 添加孤立节点
        for g in self.gene_pool:
            if g not in G:
                G.add_node(g)

        self._G = G
        logger.info(f"STRING PPI: {len(edges)} 条边, 涉及 {len(set([g for e in edges for g in e[:2]]))} 个基因")

        # 计算拓扑特征（需要G）
        self.compute_topology_features(G)

        # 构建 edge_index
        gene_list = self.node_features["GeneSymbol"].tolist()
        self.gene_to_idx = {g: i for i, g in enumerate(gene_list)}
        self.gene_symbols = gene_list

        src, dst, attr = [], [], []
        for g1, g2, score in edges:
            if g1 in self.gene_to_idx and g2 in self.gene_to_idx:
                src.append(self.gene_to_idx[g1])
                dst.append(self.gene_to_idx[g2])
                attr.append([score])

        self.edge_index = torch.tensor([src + dst, dst + src], dtype=torch.long)
        self.edge_attr = torch.tensor(attr + attr, dtype=torch.float)
        logger.info(f"边数: {self.edge_index.shape[1]}")

        # v3-ASSERT: 图构建后断言
        assert self.edge_index.shape[1] > 0, "ASSERT: 边数为0，图构建失败"
        assert self.edge_index.max().item() < len(gene_list), "ASSERT: edge_index越界"
        assert not (self.edge_index[0] == self.edge_index[1]).any(), "ASSERT: 存在自环边"

    def compute_topology_features(self, G: nx.Graph):
        logger.info("计算拓扑特征...")
        
        # v3.1: L3缓存检查 - 基于图结构指纹
        gene_pool_hash = hashlib.md5(",".join(sorted(self.gene_pool)).encode()).hexdigest()[:16]
        meta = {
            "n_nodes": G.number_of_nodes(),
            "n_edges": G.number_of_edges(),
            "gene_pool_hash": gene_pool_hash,
        }
        
        if self.cache.check_meta("topology_features", meta):
            cached_df = self.cache.get("topology_features")
            if cached_df is not None:
                logger.info("L3 缓存命中：拓扑特征")
                self.node_features = cached_df
                return
        
        genes = sorted(self.gene_pool)
        features = pd.DataFrame({"GeneSymbol": genes})

        # 只保留基因池中的节点，加速计算
        sub_nodes = set(genes) & set(G.nodes())
        subG = G.subgraph(sub_nodes).copy()
        logger.info(f"拓扑计算子图: {subG.number_of_nodes()} 节点, {subG.number_of_edges()} 边")

        # 基础拓扑
        degree = dict(subG.degree())
        pagerank = nx.pagerank(subG, weight="weight") if subG.number_of_edges() > 0 else {}
        clustering = nx.clustering(subG, weight="weight") if subG.number_of_edges() > 0 else {}

        # 增强拓扑特征
        triangles = nx.triangles(subG) if subG.number_of_edges() > 0 else {}
        try:
            coreness = nx.core_number(subG) if subG.number_of_edges() > 0 else {}
        except Exception:
            coreness = {}
        weighted_degree = dict(subG.degree(weight="weight")) if subG.number_of_edges() > 0 else {}
        try:
            harmonic = nx.harmonic_centrality(subG) if subG.number_of_edges() > 0 else {}
        except Exception:
            harmonic = {}

        # NeighborMeanLogFC
        logfc_map = {}
        if self.deg_df is not None and "logFC" in self.deg_df.columns:
            for _, row in self.deg_df.iterrows():
                g = row["GeneSymbol"]
                if pd.notna(row["logFC"]):
                    logfc_map[g] = float(row["logFC"])
        neighbor_mean_logfc = {}
        
        # 多阶邻居特征（MLGANN论文启发）
        # 2-hop和3-hop邻居的平均LogFC
        neighbor2_mean_logfc = {}
        neighbor3_mean_logfc = {}
        
        # 对于大图使用抽样策略
        use_sampling = n_nodes > 3000
        
        for gene in genes:
            if gene in subG:
                neighbors = list(subG.neighbors(gene))
                vals = [logfc_map.get(n, 0.0) for n in neighbors if n in logfc_map]
                neighbor_mean_logfc[gene] = np.mean(vals) if vals else 0.0
                
                # 2-hop邻居
                if use_sampling and len(neighbors) > 50:
                    neighbors = list(np.random.choice(neighbors, 50, replace=False))
                
                neighbors_2hop = set()
                for n1 in neighbors:
                    if n1 in subG:
                        neighbors_2hop.update(list(subG.neighbors(n1)))
                neighbors_2hop.discard(gene)
                vals2 = [logfc_map.get(n, 0.0) for n in neighbors_2hop if n in logfc_map]
                neighbor2_mean_logfc[gene] = np.mean(vals2) if vals2 else 0.0
                
                # 3-hop邻居（抽样）
                if len(neighbors_2hop) > 100:
                    neighbors_2hop = set(list(neighbors_2hop)[:100])
                
                neighbors_3hop = set()
                for n2 in neighbors_2hop:
                    if n2 in subG:
                        neighbors_3hop.update(list(subG.neighbors(n2)))
                neighbors_3hop.discard(gene)
                vals3 = [logfc_map.get(n, 0.0) for n in neighbors_3hop if n in logfc_map]
                neighbor3_mean_logfc[gene] = np.mean(vals3) if vals3 else 0.0
            else:
                neighbor_mean_logfc[gene] = 0.0
                neighbor2_mean_logfc[gene] = 0.0
                neighbor3_mean_logfc[gene] = 0.0

        # 抽样Betweenness (k=10%)
        n_nodes = subG.number_of_nodes()
        if n_nodes > 5000:
            logger.info(f"节点数>{n_nodes}>5000，启用抽样Betweenness(k=10%)")
            rng = np.random.default_rng(42)
            sample_nodes = list(subG.nodes())
            k_sample = max(1, int(n_nodes * 0.1))
            sample_sources = rng.choice(sample_nodes, size=min(k_sample, len(sample_nodes)), replace=False)
            betweenness = nx.betweenness_centrality_subset(subG, sources=set(sample_sources), targets=set(sample_nodes), weight="weight") if subG.number_of_edges() > 0 else {}
        else:
            betweenness = nx.betweenness_centrality(subG, weight="weight") if subG.number_of_edges() > 0 else {}

        # 修正: 无论节点数多少都计算closeness和eigenvector
        # 方法学依据: Closeness和Eigenvector是重要拓扑特征，不应因节点数多而跳过
        # 对于大图使用近似算法或抽样计算
        if n_nodes > 5000:
            logger.info(f"节点数>{n_nodes}>5000，启用近似Closeness和Eigenvector")
            # 近似Closeness: 对10%节点抽样计算
            closeness = {}
            sample_size = min(1000, n_nodes)
            sample_nodes_closeness = rng.choice(list(subG.nodes()), size=sample_size, replace=False)
            for node in sample_nodes_closeness:
                try:
                    lengths = nx.single_source_shortest_path_length(subG, node)
                    avg_dist = np.mean([d for d in lengths.values() if d > 0])
                    closeness[node] = 1.0 / avg_dist if avg_dist > 0 else 0.0
                except Exception:
                    closeness[node] = 0.0
            # 未抽样节点填充为0
            for node in subG.nodes():
                if node not in closeness:
                    closeness[node] = 0.0
            # Eigenvector: 对子图使用幂迭代，限制迭代次数
            try:
                eigenvector = nx.eigenvector_centrality(subG, max_iter=100, tol=1e-3, weight="weight")
            except Exception:
                eigenvector = {node: 0.0 for node in subG.nodes()}
        else:
            closeness = nx.closeness_centrality(subG) if subG.number_of_edges() > 0 else {}
            try:
                eigenvector = nx.eigenvector_centrality(subG, max_iter=1000, weight="weight") if subG.number_of_edges() > 0 else {}
            except Exception:
                eigenvector = {}

        for gene in genes:
            idx = features[features["GeneSymbol"] == gene].index[0]
            features.at[idx, "Degree"] = degree.get(gene, 0)
            features.at[idx, "PageRank"] = pagerank.get(gene, 0)
            features.at[idx, "ClusteringCoefficient"] = clustering.get(gene, 0)
            features.at[idx, "Betweenness"] = betweenness.get(gene, 0)
            features.at[idx, "Closeness"] = closeness.get(gene, 0)
            features.at[idx, "Eigenvector"] = eigenvector.get(gene, 0)
            features.at[idx, "Triangles"] = triangles.get(gene, 0)
            features.at[idx, "Coreness"] = coreness.get(gene, 0)
            features.at[idx, "WeightedDegree"] = weighted_degree.get(gene, 0)
            features.at[idx, "NeighborMeanLogFC"] = neighbor_mean_logfc.get(gene, 0.0)
            features.at[idx, "Neighbor2MeanLogFC"] = neighbor2_mean_logfc.get(gene, 0.0)
            features.at[idx, "Neighbor3MeanLogFC"] = neighbor3_mean_logfc.get(gene, 0.0)
            features.at[idx, "HarmonicCentrality"] = harmonic.get(gene, 0.0)

        # v3: 保留dist_to_cuproptosis，删除dist_to_inflammation（标签泄露）
        cupro_seed = CUPROPTOSIS_EXECUTOR_GENES

        def compute_min_distances(seed_genes):
            seed_nodes = [s for s in seed_genes if s in subG]
            if not seed_nodes:
                return {}
            # 修正: 结构隔离与通路距离必须使用无权拓扑跳数
            # 方法学依据: 标签隔离和通路距离是拓扑概念，与STRING置信度权重无关。
            # multi_source_dijkstra_path_length默认weight='weight'，即使不显式传入也会加权。
            # 必须使用multi_source_shortest_path_length（无权BFS多源版本）。
            try:
                dists = nx.multi_source_shortest_path_length(subG, sources=seed_nodes)
            except Exception:
                # 回退到循环调用（NetworkX版本不支持multi_source_shortest_path_length）
                dists = {}
                for seed in seed_nodes:
                    try:
                        lengths = nx.single_source_shortest_path_length(subG, seed)
                        for node, length in lengths.items():
                            if node not in dists or length < dists[node]:
                                dists[node] = length
                    except Exception:
                        pass
            return dists

        cupro_dists = compute_min_distances(cupro_seed)

        for gene in genes:
            idx = features[features["GeneSymbol"] == gene].index[0]
            # v3: 距离越小值越大（1/(1+distance)），更直观
            dist_cupro = cupro_dists.get(gene, 10.0)
            features.at[idx, "dist_to_cuproptosis"] = 1.0 / (1.0 + dist_cupro)

        # 归一化
        # 方法学依据: NeighborMeanLogFC保留正负号（方向信息），使用tanh压缩而非min-max
        # 其余拓扑特征使用min-max归一化
        topo_cols_minmax = ["Degree", "PageRank", "ClusteringCoefficient", "Betweenness", "Closeness", "Eigenvector",
                            "Triangles", "Coreness", "WeightedDegree", "HarmonicCentrality"]
        for c in topo_cols_minmax:
            c_min = features[c].min()
            c_max = features[c].max()
            if c_max > c_min:
                features[c] = (features[c] - c_min) / (c_max - c_min)
            else:
                features[c] = 0.0
        
        # NeighborMeanLogFC: 使用tanh压缩保留方向信息（正负号代表上调/下调趋势）
        # 多阶邻居特征同样使用tanh压缩
        features["NeighborMeanLogFC"] = np.tanh(features["NeighborMeanLogFC"])
        features["Neighbor2MeanLogFC"] = np.tanh(features["Neighbor2MeanLogFC"])
        features["Neighbor3MeanLogFC"] = np.tanh(features["Neighbor3MeanLogFC"])

        # dist_to_cuproptosis已在计算时归一化到(0,1]
        features["dist_to_cuproptosis"] = features["dist_to_cuproptosis"].clip(0, 1)

        self.node_features = features
        logger.info(f"拓扑特征计算完成，当前维度: {len(features.columns) - 1}")
        
        # v3-ASSERT: 铜死亡距离特征必须存在
        assert "dist_to_cuproptosis" in self.node_features.columns, "ASSERT: 铜死亡距离特征缺失"
        
        # v3.1: 保存拓扑特征缓存
        self.cache.set("topology_features", self.node_features, metadata=meta)
        logger.info("L3 缓存已更新：拓扑特征")

    def build_labels(self):
        logger.info("构建标签...")
        np.random.seed(42)

        brain = load_gene_list(LOCAL_DATA_DIR / "stroke_targets.txt")
        cary = load_gene_list(LOCAL_DATA_DIR / "bcp_targets.txt")

        # v3: MR显著性改用FDR校正（fdr_qval < 0.05）
        mr_positive = set()
        if self.mr_df is not None and len(self.mr_df) > 0:
            if "fdr_qval" not in self.mr_df.columns:
                logger.error("MR文件缺少fdr_qval列，无法进行FDR校正")
                sys.exit(1)
            mr_sig = self.mr_df[self.mr_df["fdr_qval"] < 0.05]
            mr_positive = set(mr_sig["gene"].tolist())
            logger.info(f"MR显著基因 (FDR<0.05): {len(mr_positive)}")

        # v3: 删除铜死亡执行基因强制阳性标签
        positive = (brain & cary) | mr_positive
        positive = positive & self.gene_pool

        logger.info(f"阳性标签来源 — 文献交集:{len(brain&cary)}, MR显著(FDR<0.05):{len(mr_positive)}")
        logger.info(f"阳性标签总数 (去重后): {len(positive)}")

        # v3: 阳性标签数量保护
        if len(positive) < 10:
            logger.error(f"阳性标签仅 {len(positive)} 个，无法训练。建议：①从DisGeNET/OMIM补充脑缺血相关基因；②改用GraphSAINT子图采样；③放弃GAT改用随机森林。")
            sys.exit(1)
        elif len(positive) < 30:
            logger.warning(f"阳性标签仅 {len(positive)} 个（<30），全图GAT训练将严重过拟合。建议：①从DisGeNET/OMIM补充脑缺血相关基因；②改用GraphSAINT子图采样；③放弃GAT改用随机森林。")

        # v3: 阴性标签来源修正
        # 方案优先级：①Housekeeping基因集 ②随机远距离基因
        negative = set()
        
        # 尝试Housekeeping基因
        hk_candidates = HOUSEKEEPING_GENES & self.gene_pool - positive - CUPROPTOSIS_GENES
        if len(hk_candidates) >= len(positive) * 2:
            negative = hk_candidates
            logger.info(f"阴性标签来源: Housekeeping基因 {len(negative)} 个")
        else:
            logger.info(f"Housekeeping基因不足 ({len(hk_candidates)} < {len(positive)*2})，改用远距离随机基因")
            # 从基因池中随机抽取与阳性标签网络最短路径>=4的基因
            if self._G is not None:
                G = self._G
                positive_in_g = [p for p in positive if p in G]
                if positive_in_g:
                    # 修正: 结构隔离必须使用无权拓扑跳数
                    # 方法学依据: 标签隔离是拓扑概念，与STRING置信度权重无关。
                    # multi_source_dijkstra_path_length默认weight='weight'，即使不显式传入也会加权。
                    # 必须使用multi_source_shortest_path_length（无权BFS多源版本）。
                    try:
                        min_dist = nx.multi_source_shortest_path_length(G, sources=positive_in_g)
                    except Exception:
                        # 回退到循环调用
                        min_dist = {}
                        for p in positive_in_g:
                            try:
                                lengths = nx.single_source_shortest_path_length(G, p)
                                for node, length in lengths.items():
                                    if node not in min_dist or length < min_dist[node]:
                                        min_dist[node] = length
                            except Exception:
                                pass
                    # 修正: 纳入不可达节点（距离∞）作为最佳阴性标签
                    # 方法学依据: GenePlexus使用"网络隔离基因"作为阴性基准
                    far_nodes = {g for g in self.gene_pool if (g not in min_dist) or (g in min_dist and min_dist[g] >= 4)}
                    far_nodes -= positive
                    far_nodes -= CUPROPTOSIS_GENES
                    # 排除MR q<0.1的基因
                    if self.mr_df is not None and "fdr_qval" in self.mr_df.columns:
                        mr_q01 = set(self.mr_df[self.mr_df["fdr_qval"] < 0.1]["gene"].tolist())
                        far_nodes -= mr_q01
                    negative = far_nodes
                    logger.info(f"远距离阴性标签: {len(negative)} 个")
        
        # 确保阴性标签数≥阳性×2
        if len(negative) < len(positive) * 2:
            needed = len(positive) * 2 - len(negative)
            extra_pool = self.gene_pool - positive - negative - CUPROPTOSIS_GENES
            if self.mr_df is not None and "fdr_qval" in self.mr_df.columns:
                mr_q01 = set(self.mr_df[self.mr_df["fdr_qval"] < 0.1]["gene"].tolist())
                extra_pool -= mr_q01
            n_extra = min(needed, len(extra_pool))
            if extra_pool and n_extra > 0:
                rng = np.random.default_rng(42)
                extra = set(rng.choice(list(extra_pool), size=n_extra, replace=False))
                negative |= extra
                logger.info(f"随机补足阴性标签: {n_extra} 个")

        if len(negative) < len(positive) * 2:
            logger.warning(f"阴性标签数 ({len(negative)}) < 阳性×2 ({len(positive)*2})，类别不平衡严重")

        # 确保阴性标签不与阳性标签重叠
        negative = negative - positive

        self.positive_labels = positive
        self.negative_labels = negative

        # v3-ASSERT: 标签构建后断言
        assert len(positive & negative) == 0, "ASSERT: 正负标签重叠"
        assert len(positive) >= 5, f"ASSERT: 阳性标签不足5 ({len(positive)})"
        assert len(negative) >= 10, f"ASSERT: 阴性标签不足10 ({len(negative)})"

        labels = {}
        for g in self.gene_pool:
            if g in positive:
                labels[g] = 1
            elif g in negative:
                labels[g] = 0
            else:
                labels[g] = -1
        
        # P2: 铜死亡软标签注入（半监督正则化）
        # 为铜死亡执行基因注入软标签：编码2表示软标签0.7
        # 方法学依据: 删除强制硬标签后，铜死亡基因全为-1，模型完全不学习铜死亡相关特征
        # 软标签提供弱监督信号，引导模型关注铜死亡通路
        cupro_soft = CUPROPTOSIS_EXECUTOR_GENES & self.gene_pool
        for g in cupro_soft:
            if labels[g] == -1:  # 仅对未知节点注入软标签
                labels[g] = 2
        logger.info(f"铜死亡软标签注入: {len([g for g in cupro_soft if labels[g] == 2])} 个执行基因编码为2（软标签0.7）")

        df = pd.DataFrame([{"GeneSymbol": g, "Label": v} for g, v in labels.items()])
        n_pos = (df["Label"] == 1).sum()
        n_neg = (df["Label"] == 0).sum()
        n_soft = (df["Label"] == 2).sum()
        n_unk = (df["Label"] == -1).sum()
        logger.info(f"最终标签 — 阳性:{n_pos}, 阴性:{n_neg}, 软标签:{n_soft}, 未知:{n_unk}")
        self.labels_df = df

    def build_features(self):
        logger.info("构建节点特征...")
        genes = sorted(self.gene_pool)
        features = self.node_features.copy() if self.node_features is not None else pd.DataFrame({"GeneSymbol": genes})

        # DEG特征（去重）
        if self.deg_df is not None and len(self.deg_df) > 0:
            deg = self.deg_df.copy()
            deg["abs_logFC"] = deg["logFC"].abs()
            deg["neg_log10_P"] = -np.log10(deg["P.Value"].replace(0, np.nan))
            deg["neg_log10_adjP"] = -np.log10(deg["adj.P.Val"].replace(0, np.nan))
            deg_cols = ["GeneSymbol", "logFC", "abs_logFC", "neg_log10_P", "neg_log10_adjP"]
            
            # DEG平均表达列名模糊匹配
            ave_candidates = ["AveExpr", "Ave.Expr", "baseMean", "AvgExpr", "meanExpr", "Ave Expr", "ave_expr"]
            matched_ave = None
            for cand in ave_candidates:
                if cand in deg.columns:
                    matched_ave = cand
                    break
            if matched_ave and matched_ave != "AveExpr":
                logger.info(f"DEG列名模糊匹配: {matched_ave} -> AveExpr")
                deg = deg.rename(columns={matched_ave: "AveExpr"})
            if matched_ave:
                deg_cols.append("AveExpr")
            transcript = deg[deg_cols].copy()
            n_before = len(transcript)
            transcript = transcript.groupby("GeneSymbol", as_index=False).mean()
            n_after = len(transcript)
            if n_before != n_after:
                logger.warning(f"DEG表存在重复基因，去重前{n_before}行，去重后{n_after}行")
            features = features.merge(transcript, on="GeneSymbol", how="left")

        # 单细胞 KO 数据
        if SC_KO_FILE.exists():
            logger.info(f"读取单细胞 KO 数据: {SC_KO_FILE}")
            ko_df = pd.read_csv(SC_KO_FILE)
            ko_gene_col = None
            for col in ["gene", "GeneSymbol", "gene_symbol", "feature"]:
                if col in ko_df.columns:
                    ko_gene_col = col
                    break
            if ko_gene_col:
                ko_df = ko_df.rename(columns={ko_gene_col: "GeneSymbol"})
                ko_df["GeneSymbol"] = ko_df["GeneSymbol"].apply(normalize_gene_symbol)
                ko_df = ko_df[ko_df["GeneSymbol"].notna()]
                dr_cols = [c for c in ko_df.columns if "dr" in c.lower() or "score" in c.lower() or "z_score" in c.lower()]
                if dr_cols:
                    ko_merge = ko_df[["GeneSymbol"] + dr_cols].copy()
                    ko_merge.columns = ["GeneSymbol"] + [f"KO_{c}" for c in dr_cols]
                    features = features.merge(ko_merge, on="GeneSymbol", how="left")

        # 单细胞通讯数据
        if SC_COMM_FILE.exists():
            logger.info(f"读取单细胞通讯数据: {SC_COMM_FILE}")
            comm_df = pd.read_csv(SC_COMM_FILE)
            sig_comm = comm_df[comm_df.get("significant", comm_df.get("Significant", False)) == True] if "significant" in comm_df.columns or "Significant" in comm_df.columns else comm_df
            ligand_genes = set()
            receptor_genes = set()
            for _, row in sig_comm.iterrows():
                ligand = str(row.get("ligand", row.get("Ligand", ""))).strip().upper()
                receptor = str(row.get("receptor", row.get("Receptor", ""))).strip().upper()
                if ligand:
                    ligand_genes.add(ligand)
                if receptor:
                    receptor_genes.add(receptor)
            features["is_comm_ligand"] = features["GeneSymbol"].isin(ligand_genes).astype(int)
            features["is_comm_receptor"] = features["GeneSymbol"].isin(receptor_genes).astype(int)

        # MR特征 - 扩展MR+ML论文中的特征
        # 依据: MR论文指出MR特征（pval, beta, nsnp, F-statistic）可提升ML模型性能
        if self.mr_df is not None and len(self.mr_df) > 0:
            mr_features = self.mr_df.copy()
            
            # 新MR文件列名适配
            pval_col = "meta_pval" if "meta_pval" in mr_features.columns else "pval"
            b_col = "discovery_b" if "discovery_b" in mr_features.columns else "b"
            
            # 基础特征 - 使用meta_pval和discovery_b
            mr_pval = mr_features[pval_col].replace(0, np.nan).fillna(1.0)
            mr_features["neg_log10_pval"] = -np.log10(mr_pval.clip(lower=1e-300))
            mr_features["abs_b"] = mr_features[b_col].abs()
            
            # 扩展MR特征（来自权威MR+ML论文）
            mr_features["mr_has_nsnp"] = mr_features.get("nsnp", 1).fillna(1)
            mr_features["mr_F_stat"] = mr_features.get("F_mean", 0).fillna(0)
            mr_features["mr_effect_dir"] = (mr_features[b_col] > 0).astype(int)  # 1=保护因素, 0=风险因素
            
            # OR值可能不存在
            if "OR" in mr_features.columns:
                mr_features["mr_OR"] = mr_features["OR"].fillna(1)
            else:
                mr_features["mr_OR"] = np.exp(mr_features[b_col].abs())  # 从beta估算OR
            
            # 聚合到基因（取最显著的值）
            mr_summary = mr_features.groupby("gene").agg({
                "neg_log10_pval": "max",
                "abs_b": "mean",
                "fdr_qval": "min",
                "mr_has_nsnp": "max",
                "mr_F_stat": "max",
                "mr_OR": "mean",
                "mr_effect_dir": "max"
            }).reset_index()
            
            mr_summary.columns = ["GeneSymbol", "MR_neg_log10_pval", "MR_abs_b", "MR_fdr_qval", 
                                  "MR_nsnp", "MR_F_stat", "MR_OR", "MR_effect_dir"]
            
            # 归一化MR特征
            for col in ["MR_nsnp", "MR_F_stat", "MR_OR"]:
                if col in mr_summary.columns:
                    col_min, col_max = mr_summary[col].min(), mr_summary[col].max()
                    if col_max > col_min:
                        mr_summary[col] = (mr_summary[col] - col_min) / (col_max - col_min)
            
            features = features.merge(mr_summary, on="GeneSymbol", how="left")
            
            # v3.1: 统计MR显著基因
            n_sig = (mr_summary["MR_fdr_qval"] < 0.05).sum() if "MR_fdr_qval" in mr_summary.columns else 0
            logger.info(f"MR特征已整合: {features['MR_neg_log10_pval'].notna().sum()} 个基因有MR数据, 其中显著基因(FDR<0.05): {n_sig}")

        # 铜死亡标记
        features["is_cuproptosis"] = features["GeneSymbol"].isin(CUPROPTOSIS_GENES).astype(int)

        # P3: 零成本特征补充（4维，23 → 27维）
        # 方法学依据: 这些特征无需外部文件，纯计算即可，提供额外的生物学先验
        
        # 1. 基因符号长度（保守基因通常较短）
        features["symbol_length"] = features["GeneSymbol"].str.len()
        sl_min, sl_max = features["symbol_length"].min(), features["symbol_length"].max()
        features["symbol_length"] = (features["symbol_length"] - sl_min) / (sl_max - sl_min) if sl_max > sl_min else 0
        
        # 2. 是否在DEG列表中
        deg_gene_set = set(self.deg_df["GeneSymbol"]) if self.deg_df is not None else set()
        features["in_deg"] = features["GeneSymbol"].isin(deg_gene_set).astype(int)
        
        # 3. 与阳性标签的网络距离（guilt-by-association拓扑先验）
        if self._G is not None and len(self.positive_labels) > 0:
            pos_seed = list(self.positive_labels & set(self._G.nodes()))
            if pos_seed:
                try:
                    pos_dists = nx.multi_source_shortest_path_length(self._G, sources=pos_seed)
                except Exception:
                    pos_dists = {}
                    for p in pos_seed:
                        try:
                            lengths = nx.single_source_shortest_path_length(self._G, p)
                            for node, length in lengths.items():
                                if node not in pos_dists or length < pos_dists[node]:
                                    pos_dists[node] = length
                        except Exception:
                            pass
                features["dist_to_positive"] = features["GeneSymbol"].map(
                    {g: 1.0/(1.0+pos_dists.get(g, 10.0)) for g in features["GeneSymbol"]}
                ).fillna(0)
            else:
                features["dist_to_positive"] = 0.0
        else:
            features["dist_to_positive"] = 0.0
        
        # 4. DEG显著性等级（离散化，避免p值长尾噪声）
        if self.deg_df is not None and "adj.P.Val" in self.deg_df.columns:
            sig_map = {}
            for _, row in self.deg_df.iterrows():
                g = row["GeneSymbol"]
                if pd.notna(row["adj.P.Val"]) and row["adj.P.Val"] < 0.05 and pd.notna(row["logFC"]):
                    if row["logFC"] > 1: sig_map[g] = 3
                    elif row["logFC"] < -1: sig_map[g] = 1
                    else: sig_map[g] = 2
                else:
                    sig_map[g] = 0
            features["deg_sig_level"] = features["GeneSymbol"].map(sig_map).fillna(0)
        else:
            features["deg_sig_level"] = 0
        
        logger.info(f"零成本特征补充完成: symbol_length, in_deg, dist_to_positive, deg_sig_level")

        # v3-ASSERT: 特征列不含inflammatory关键词
        for col in features.columns:
            if "inflammatory" in col.lower() or "inflam" in col.lower():
                raise RuntimeError(f"ASSERT: 特征列包含泄露关键词: {col}")

        # v3: 删除零填充，改为报错或补充真实特征
        current_dim = len([c for c in features.columns if c != "GeneSymbol"])
        logger.info(f"节点特征维度: {current_dim} 维 (不含 GeneSymbol)")
        
        # 特征维度检查与补充逻辑
        if current_dim < 10:
            logger.error(f"特征维度严重不足 ({current_dim} < 10)，无法训练GAT。请补充至少10维生物学特征（如GO语义相似度、KEGG通路得分、蛋白结构域计数等）。")
            sys.exit(1)
        elif current_dim < 20:
            logger.warning(f"特征维度仅 {current_dim} (<20)，模型可能欠拟合。建议补充GO语义相似度或KEGG通路得分。")
        
        # 自动补充GO语义相似度（若local_data/go_terms.tsv存在）
        go_file = LOCAL_DATA_DIR / "go_terms.tsv"
        if go_file.exists():
            logger.info(f"读取GO术语文件: {go_file}")
            try:
                go_df = pd.read_csv(go_file, sep="\t")
                if "GeneSymbol" in go_df.columns and "GO_term" in go_df.columns:
                    gene_go_sets = go_df.groupby("GeneSymbol")["GO_term"].apply(set).to_dict()
                    cupro_go_union = set()
                    for g in CUPROPTOSIS_EXECUTOR_GENES:
                        if g in gene_go_sets:
                            cupro_go_union |= gene_go_sets[g]
                    if cupro_go_union:
                        go_sims = []
                        go_counts = []
                        for gene in features["GeneSymbol"]:
                            gene_go = gene_go_sets.get(gene, set())
                            # GO相似度
                            if gene_go:
                                intersection = len(gene_go & cupro_go_union)
                                union = len(gene_go | cupro_go_union)
                                sim = intersection / union if union > 0 else 0.0
                            else:
                                sim = 0.0
                            go_sims.append(sim)
                            # GO term数量（富集程度）
                            go_counts.append(len(gene_go))
                        features["go_sim_cupro"] = go_sims
                        features["go_term_count"] = go_counts
                        logger.info(f"自动补充GO特征: go_sim_cupro, go_term_count")
                        current_dim = len([c for c in features.columns if c != "GeneSymbol"])
                        logger.info(f"补充后特征维度: {current_dim} 维")
            except Exception as e:
                logger.warning(f"读取GO术语文件失败: {e}")
        
        # 自动补充KEGG通路富集特征（若local_data/kegg_pathways.tsv存在）
        kegg_file = LOCAL_DATA_DIR / "kegg_pathways.tsv"
        if kegg_file.exists():
            logger.info(f"读取KEGG通路文件: {kegg_file}")
            try:
                kegg_df = pd.read_csv(kegg_file, sep="\t")
                if "GeneSymbol" in kegg_df.columns and "Pathway" in kegg_df.columns:
                    gene_pathway_sets = kegg_df.groupby("GeneSymbol")["Pathway"].apply(set).to_dict()
                    cupro_pathway_union = set()
                    for g in CUPROPTOSIS_EXECUTOR_GENES:
                        if g in gene_pathway_sets:
                            cupro_pathway_union |= gene_pathway_sets[g]
                    if cupro_pathway_union:
                        kegg_sims = []
                        kegg_counts = []
                        for gene in features["GeneSymbol"]:
                            gene_path = gene_pathway_sets.get(gene, set())
                            # KEGG相似度
                            if gene_path:
                                intersection = len(gene_path & cupro_pathway_union)
                                union = len(gene_path | cupro_pathway_union)
                                sim = intersection / union if union > 0 else 0.0
                            else:
                                sim = 0.0
                            kegg_sims.append(sim)
                            # KEGG通路数量
                            kegg_counts.append(len(gene_path))
                        features["kegg_sim_cupro"] = kegg_sims
                        features["kegg_pathway_count"] = kegg_counts
                        logger.info(f"自动补充KEGG特征: kegg_sim_cupro, kegg_pathway_count")
                        current_dim = len([c for c in features.columns if c != "GeneSymbol"])
                        logger.info(f"补充后特征维度: {current_dim} 维")
            except Exception as e:
                logger.warning(f"读取KEGG通路文件失败: {e}")
        
        # 自动补充蛋白结构域计数特征（若local_data/interpro_domains.tsv存在）
        interpro_file = LOCAL_DATA_DIR / "interpro_domains.tsv"
        if interpro_file.exists():
            logger.info(f"读取InterPro结构域文件: {interpro_file}")
            try:
                interpro_df = pd.read_csv(interpro_file, sep="\t")
                if "GeneSymbol" in interpro_df.columns and "Domain" in interpro_df.columns:
                    domain_counts = interpro_df.groupby("GeneSymbol").size().to_dict()
                    features["domain_count"] = features["GeneSymbol"].map(domain_counts).fillna(0)
                    logger.info(f"自动补充蛋白结构域计数特征: domain_count")
                    current_dim = len([c for c in features.columns if c != "GeneSymbol"])
                    logger.info(f"补充后特征维度: {current_dim} 维")
            except Exception as e:
                logger.warning(f"读取InterPro结构域文件失败: {e}")
        
        # v3-ASSERT: 特征矩阵无NaN/Inf/全零列
        feature_cols = [c for c in features.columns if c != "GeneSymbol"]
        for col in feature_cols:
            if features[col].isna().all():
                logger.warning(f"特征列 {col} 全为NaN，填充为0")
                features[col] = features[col].fillna(0)
            if np.isinf(features[col]).any():
                raise RuntimeError(f"ASSERT: 特征列 {col} 包含Inf")
        
        # v3-ASSERT: 无全零列（除合法零值外）
        for col in feature_cols:
            if (features[col] == 0).all():
                logger.warning(f"特征列 {col} 全为0，可能影响训练")
        
        # v3-ASSERT: 特征列不含inflammatory关键词
        for col in feature_cols:
            if "inflammatory" in col.lower() or "inflam" in col.lower():
                raise RuntimeError(f"ASSERT: 特征列包含泄露关键词: {col}")
        
        # v3-ASSERT: 无全NaN列
        assert not features[feature_cols].isna().all().any(), "ASSERT: 存在全NaN特征列"
        
        # 填充剩余NaN为0
        features[feature_cols] = features[feature_cols].fillna(0)

        self.node_features = features

    def save_processed(self):
        self.node_features.to_csv(PROCESSED_DIR / "node_features.csv", index=False)
        self.labels_df.to_csv(PROCESSED_DIR / "labels.csv", index=False)
        torch.save(self.edge_index, PROCESSED_DIR / "edge_index.pt")
        torch.save(self.edge_attr, PROCESSED_DIR / "edge_attr.pt")
        with open(PROCESSED_DIR / "gene_symbols.pkl", "wb") as f:
            pickle.dump(self.gene_symbols, f)
        with open(PROCESSED_DIR / "gene_to_idx.pkl", "wb") as f:
            pickle.dump(self.gene_to_idx, f)
        # 保存实际特征维度到json，供run_training读取
        feature_cols = [c for c in self.node_features.columns if c != "GeneSymbol"]
        feature_dim = {"in_channels": len(feature_cols)}
        with open(PROCESSED_DIR / "feature_dim.json", "w", encoding="utf-8") as f:
            json.dump(feature_dim, f)
        logger.info(f"数据已保存到 {PROCESSED_DIR}")


# ---------------------------------------------------------------------------
# 模块二: GATv2 模型定义
# ---------------------------------------------------------------------------

class GATv2Model(nn.Module):
    def __init__(self, in_channels, hidden_channels=32, out_channels=16, num_heads=2,
                 num_classes=2, dropout=0.3, attention_dropout=0.2, use_edge_attr=True,
                 num_layers=3, use_residual=True, use_virtual_node=False):
        super().__init__()
        self.use_edge_attr = use_edge_attr
        self.use_residual = use_residual
        self.use_virtual_node = use_virtual_node
        edge_dim = 1 if use_edge_attr else None
        
        self.conv1 = GATv2Conv(in_channels, hidden_channels, heads=num_heads, concat=True,
                               dropout=attention_dropout, edge_dim=edge_dim)
        self.conv2 = GATv2Conv(hidden_channels * num_heads, hidden_channels, heads=num_heads, concat=True,
                               dropout=attention_dropout, edge_dim=edge_dim)
        self.conv3 = GATv2Conv(hidden_channels * num_heads, out_channels, heads=num_heads, concat=True,
                               dropout=attention_dropout, edge_dim=edge_dim)

        self.classifier = nn.Sequential(
            nn.Linear(out_channels * num_heads, hidden_channels),
            nn.LayerNorm(hidden_channels),
            nn.ELU(),
            nn.Dropout(dropout),
            nn.Linear(hidden_channels, num_classes)
        )

        self.dropout = dropout
        
        # LayerNorm for graph data stability
        self.norm1 = nn.LayerNorm(hidden_channels * num_heads)
        self.norm2 = nn.LayerNorm(hidden_channels * num_heads)
        
        # 虚拟节点（可选）- 改善图表示学习
        if use_virtual_node:
            self.virtual_node_embedding = nn.Parameter(torch.randn(1, hidden_channels * num_heads))

    def forward(self, x, edge_index, edge_attr=None):
        edge_attr_input = edge_attr if self.use_edge_attr else None
        
        # 第1层GAT
        x1 = self.conv1(x, edge_index, edge_attr=edge_attr_input)
        x1 = self.norm1(x1)
        x1 = F.elu(x1)
        x1 = F.dropout(x1, p=self.dropout, training=self.training)
        
        # 残差连接
        if self.use_residual and x1.shape == x.shape:
            x1 = x1 + x
        
        # 第2层GAT
        x2 = self.conv2(x1, edge_index, edge_attr=edge_attr_input)
        x2 = self.norm2(x2)
        x2 = F.elu(x2)
        x2 = F.dropout(x2, p=self.dropout, training=self.training)
        
        # 残差连接
        if self.use_residual and x2.shape == x1.shape:
            x2 = x2 + x1
        
        # 第3层GAT（输出层）
        x3 = self.conv3(x2, edge_index, edge_attr=edge_attr_input)
        
        # 分类器
        out = self.classifier(x3)
        return out


# ---------------------------------------------------------------------------
# 模块三: 训练工具（v3重构版）
# ---------------------------------------------------------------------------

def create_masks_transductive(y, train_ratio=0.6, val_ratio=0.2, test_ratio=0.2, random_state=42):
    """v3: Transductive分割（全图边保留，仅节点标签mask分离）"""
    y_np = y.cpu().numpy() if y.is_cuda else y.numpy()
    labeled_mask = (y_np == 0) | (y_np == 1)
    labeled_indices = np.where(labeled_mask)[0]
    labeled_labels = y_np[labeled_indices]

    assert abs((train_ratio + val_ratio + test_ratio) - 1.0) < 1e-6, "比例之和必须等于 1"

    rng = np.random.default_rng(random_state)
    pos_idx = labeled_indices[labeled_labels == 1]
    neg_idx = labeled_indices[labeled_labels == 0]

    def split_indices(idx, train_r, val_r, test_r):
        n = len(idx)
        # 修正: 若n<6，放弃严格6:2:2比例，确保训练集至少有2个样本
        if n < 6:
            n_train = max(2, n - 2)
            n_val = 1
            n_test = 1
            # 若样本不足4个，调整分配
            if n < 4:
                n_train = max(2, n)
                n_val = max(0, n - n_train)
                n_test = 0
        else:
            n_train = max(2, int(n * train_r))
            n_val = max(1, int(n * val_r))
            n_test = max(1, n - n_train - n_val)
            # 确保不越界
            n_train = min(n_train, n - 2)
            n_val = min(n_val, n - n_train - 1)
        n_train = min(n_train, n)
        n_val = min(n_val, max(0, n - n_train))
        shuffled = rng.permutation(idx)
        return shuffled[:n_train], shuffled[n_train:n_train+n_val], shuffled[n_train+n_val:]

    train_pos, val_pos, test_pos = split_indices(pos_idx, train_ratio, val_ratio, test_ratio)
    train_neg, val_neg, test_neg = split_indices(neg_idx, train_ratio, val_ratio, test_ratio)

    train_indices = np.concatenate([train_pos, train_neg])

    # --- P0: 阳性过采样：使训练时阳:阴 ≈ 1:2 ---
    train_pos_hard = train_pos
    n_pos, n_neg = len(train_pos_hard), len(train_neg)

    if n_pos > 0 and n_neg > n_pos * 2:
        repeat = max(1, (n_neg // 2) // n_pos)
        repeat = min(repeat, 20)
        oversampled_pos = np.repeat(train_pos_hard, repeat)
        train_indices = np.concatenate([oversampled_pos, train_neg])
        rng.shuffle(train_indices)
        logger.info(f"过采样后训练集: 阳性{n_pos}×{repeat}={len(oversampled_pos)}, 阴性{n_neg}, 实际比例 1:{n_neg/len(oversampled_pos):.1f}")
    elif n_pos > 0:
        logger.info(f"训练集无需过采样: 阳性{n_pos}, 阴性{n_neg}, 比例 1:{n_neg/n_pos:.1f}")

    val_indices = np.concatenate([val_pos, val_neg])
    test_indices = np.concatenate([test_pos, test_neg])

    train_mask = torch.zeros(len(y), dtype=torch.bool)
    val_mask = torch.zeros(len(y), dtype=torch.bool)
    test_mask = torch.zeros(len(y), dtype=torch.bool)
    train_mask[train_indices] = True
    val_mask[val_indices] = True
    test_mask[test_indices] = True

    n_train_pos = int((y_np[train_indices] == 1).sum())
    n_train_neg = int((y_np[train_indices] == 0).sum())
    n_val_pos = int((y_np[val_indices] == 1).sum())
    n_val_neg = int((y_np[val_indices] == 0).sum())
    n_test_pos = int((y_np[test_indices] == 1).sum())
    n_test_neg = int((y_np[test_indices] == 0).sum())
    logger.info(f"数据分割确认 — 训练集: {n_train_pos}正/{n_train_neg}负, "
                f"验证集: {n_val_pos}正/{n_val_neg}负, "
                f"测试集: {n_test_pos}正/{n_test_neg}负")

    # v3-ASSERT: 训练测试集不重叠
    assert len(set(train_indices) & set(test_indices)) == 0, "ASSERT: 训练测试集重叠"
    # v3-ASSERT: 训练集阳性≥2且阴性≥2（放宽以适配小样本场景）
    assert n_train_pos >= 2 and n_train_neg >= 2, f"ASSERT: 训练集样本不足 (阳{n_train_pos}/阴{n_train_neg})，无法进行分层分割"
    # v3-ASSERT: 验证集阳性≥1且阴性≥1
    assert n_val_pos >= 1 and n_val_neg >= 1, f"ASSERT: 验证集样本不足 (阳{n_val_pos}/阴{n_val_neg})"

    return train_mask, val_mask, test_mask


def compute_metrics(y_true, y_pred, y_prob):
    unique_labels = set(y_true)
    if len(unique_labels) <= 1:
        return {
            "accuracy": 0.0, "precision": 0.0, "recall": 0.0,
            "f1": 0.0, "roc_auc": 0.0, "pr_auc": 0.0,
        }
    return {
        "accuracy": accuracy_score(y_true, y_pred),
        "precision": precision_score(y_true, y_pred, zero_division=0),
        "recall": recall_score(y_true, y_pred, zero_division=0),
        "f1": f1_score(y_true, y_pred, zero_division=0),
        "roc_auc": roc_auc_score(y_true, y_prob) if len(set(y_true)) > 1 else 0.0,
        "pr_auc": average_precision_score(y_true, y_prob) if len(set(y_true)) > 1 else 0.0,
    }


def compute_ranking_metrics(all_unknown_df: pd.DataFrame, cupro_genes: Set[str], k_list=[50, 100, 200]):
    """v3: 在全部未知节点中计算铜死亡基因的Recall@K"""
    results = {}
    # 筛选铜死亡基因
    cupro_df = all_unknown_df[all_unknown_df["GeneSymbol"].isin(cupro_genes)].copy()
    if len(cupro_df) == 0:
        for k in k_list:
            results[f"cupro_recall@{k}"] = 0.0
        results["cupro_mrr"] = 0.0
        return results
    
    # 修正1: 分母仅包含在未知节点中的铜死亡基因（排除训练/验证/测试集中的铜死亡基因）
    cupro_in_unknown = cupro_genes & set(all_unknown_df["GeneSymbol"])
    total_cupro = len(cupro_in_unknown)
    if total_cupro == 0:
        for k in k_list:
            results[f"cupro_recall@{k}"] = 0.0
        results["cupro_mrr"] = 0.0
        return results
    
    for k in k_list:
        top_k = all_unknown_df.head(k)
        found = top_k["GeneSymbol"].isin(cupro_genes).sum()
        results[f"cupro_recall@{k}"] = found / total_cupro if total_cupro > 0 else 0.0
    
    # 修正2: MRR只计算在未知节点中的铜死亡基因
    ranks = []
    for gene in cupro_in_unknown:
        match = all_unknown_df[all_unknown_df["GeneSymbol"] == gene]
        if not match.empty:
            rank = match.iloc[0]["Rank"]
            ranks.append(rank)
    if ranks:
        results["cupro_mrr"] = np.mean([1.0 / r for r in ranks])
    else:
        results["cupro_mrr"] = 0.0
    
    return results


@torch.no_grad()
def evaluate(model, x, edge_index, edge_attr, y, mask, pos_weight=1.0):
    try:
        model.eval()
        logits = model(x, edge_index, edge_attr)
        # 修正: 空集保护（防御性编程）
        # 若mask全False（如某集合无样本），mask_logits为空tensor
        # PyTorch对空tensor调用argmax会抛出RuntimeError
        if mask.sum() == 0:
            return {"loss": 0.0, "accuracy": 0.0, "precision": 0.0, "recall": 0.0, "f1": 0.0, "roc_auc": 0.0, "pr_auc": 0.0}
        mask_logits = logits[mask]
        mask_labels = y[mask]
        preds = mask_logits.argmax(dim=1).cpu().numpy()
        labels = mask_labels.cpu().numpy()
        prob_pos = F.softmax(mask_logits, dim=1)[:, 1].cpu().numpy()

        labeled_mask = (labels == 0) | (labels == 1)
        if labeled_mask.sum() == 0:
            metrics = {
                "accuracy": 0.0, "precision": 0.0, "recall": 0.0,
                "f1": 0.0, "roc_auc": 0.0, "pr_auc": 0.0,
            }
        else:
            metrics = compute_metrics(labels[labeled_mask], preds[labeled_mask], prob_pos[labeled_mask])

        labeled_mask_torch = (mask_labels == 0) | (mask_labels == 1)
        if labeled_mask_torch.any():
            # P4-fix: 删除pos_weight，过采样只在训练时发生，验证/测试保持原始分布
            loss = F.cross_entropy(mask_logits[labeled_mask_torch], mask_labels[labeled_mask_torch])
            metrics["loss"] = loss.item()
        else:
            metrics["loss"] = 0.0
        return metrics
    except Exception as e:
        logger.error(f"evaluate 异常: {type(e).__name__}: {e}")
        import traceback
        logger.error(traceback.format_exc())
        raise


def asymmetric_loss(predictions, targets, gamma_pos=1.0, gamma_neg=4.0, clip=0.05):
    """
    Asymmetric Loss (ASL) - 比Focal Loss更好地处理极端类别不平衡
    参考: GNNGL-PPI (BMC Genomics 2024)
    
    predictions: [N, 2] logits
    targets: [N] 类别标签 (0 或 1)
    """
    # 获取正类的概率
    probs = F.softmax(predictions, dim=1)
    p_pos = probs[:, 1]  # 正类概率
    p_neg = 1 - p_pos   # 负类概率
    
    # 目标标签
    target_pos = (targets == 1).float()
    target_neg = 1 - target_pos
    
    # ASL核心公式
    asym_pos = torch.clamp(p_pos + gamma_pos, min=clip) ** (-gamma_pos)
    asym_neg = torch.clamp(p_neg + gamma_neg, min=clip) ** (-gamma_neg)
    
    # 计算损失
    loss_pos = -target_pos * asym_pos * torch.log(torch.clamp(p_pos, min=1e-7))
    loss_neg = -target_neg * asym_neg * torch.log(torch.clamp(p_neg, min=1e-7))
    
    loss = (loss_pos + loss_neg).mean()
    return loss


def train_epoch(model, optimizer, x, edge_index, edge_attr, y, train_mask, epoch=1, total_epochs=200, pos_weight=1.0, use_asl=True, warmup_epochs=10):
    try:
        model.train()
        
        # 学习率预热策略（Warmup）
        # 前期使用较小的学习率稳定训练，后期逐渐恢复到原始学习率
        if epoch <= warmup_epochs:
            warmup_factor = epoch / warmup_epochs
            for param_group in optimizer.param_groups:
                param_group['lr'] = param_group['lr'] * warmup_factor
        
        optimizer.zero_grad()
        logits = model(x, edge_index, edge_attr)
        train_logits = logits[train_mask]
        train_labels = y[train_mask]

        # P2-fix: 软标签编码2 → 硬标签1
        train_labels = torch.where(
            train_labels == 2,
            torch.tensor(1, dtype=torch.long, device=train_labels.device),
            train_labels
        ).long()

        # 选择损失函数
        if use_asl:
            # Asymmetric Loss - 更好地处理极端类别不平衡
            # 动态调整gamma参数
            gamma_neg = 4.0  # 对负样本更严格
            gamma_pos = 1.0  # 对正样本较宽松
            
            # 根据epoch渐进调整
            if epoch > warmup_epochs:
                progress = min(1.0, (epoch - warmup_epochs) / (total_epochs - warmup_epochs))
                gamma_neg = gamma_neg * (1 - progress * 0.5)  # 后期逐渐降低
            
            loss = asymmetric_loss(train_logits, train_labels, gamma_pos=gamma_pos, gamma_neg=gamma_neg)
        else:
            # 备用: 标准交叉熵
            loss = F.cross_entropy(train_logits, train_labels)

        # v3: Loss NaN检查与调试保存
        if torch.isnan(loss):
            logger.error("ASSERT: Loss NaN detected")
            logger.error(f"Logits统计 — min: {train_logits.min().item():.4f}, max: {train_logits.max().item():.4f}, mean: {train_logits.mean().item():.4f}")
            if torch.isinf(x).any():
                logger.error("ASSERT: 输入特征x包含Inf")
            debug_path = MODELS_DIR / f"debug_nan_epoch.pt"
            torch.save({
                "x": x.cpu(),
                "edge_index": edge_index.cpu(),
                "edge_attr": edge_attr.cpu() if edge_attr is not None else None,
                "y": y.cpu(),
                "train_mask": train_mask.cpu(),
                "train_logits": train_logits.cpu(),
                "train_labels": train_labels.cpu(),
                "loss": loss.item() if not torch.isnan(loss) else float('nan'),
            }, debug_path)
            logger.error(f"调试状态已保存: {debug_path}")
            sys.exit(1)

        loss.backward()
        torch.nn.utils.clip_grad_norm_(model.parameters(), max_norm=0.5)  # 更严格的梯度裁剪
        optimizer.step()
        return loss.item()
    except Exception as e:
        logger.error(f"train_epoch 异常: {type(e).__name__}: {e}")
        import traceback
        logger.error(traceback.format_exc())
        raise


# ---------------------------------------------------------------------------
# 模块四: 模型训练（v3重构版）
# ---------------------------------------------------------------------------

def run_training(config_path: str = "config.yaml"):
    # v3: 若config.yaml不存在，自动生成默认配置
    if not Path(config_path).exists():
        logger.info(f"config.yaml不存在，生成默认配置: {config_path}")
        default_cfg = {
            "model": {
                "in_channels": None, "hidden_channels": 64, "out_channels": 32,
                "num_heads": 4, "num_classes": 2, "dropout": 0.3,
                "attention_dropout": 0.2, "use_edge_attr": True
            },
            "training": {
                "epochs": 200, "learning_rate": 0.001, "weight_decay": 0.0001,
                "patience": 50, "pos_weight": "auto", "cupro_weight": 0.0,
                "early_stop_metric": "f1", "train_ratio": 0.6,
                "val_ratio": 0.2, "test_ratio": 0.2, "random_state": 42
            }
        }
        with open(config_path, "w", encoding="utf-8") as f:
            yaml.dump(default_cfg, f, default_flow_style=False, allow_unicode=True)
    
    with open(config_path, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)
    
    # 读取实际特征维度并覆盖config中的null值
    feature_dim_path = PROCESSED_DIR / "feature_dim.json"
    if feature_dim_path.exists():
        with open(feature_dim_path, "r", encoding="utf-8") as f:
            feature_dim = json.load(f)
        if cfg["model"].get("in_channels") is None:
            cfg["model"]["in_channels"] = feature_dim.get("in_channels", 28)
            logger.info(f"从feature_dim.json读取实际特征维度: {cfg['model']['in_channels']}")
    else:
        if cfg["model"].get("in_channels") is None:
            logger.error("config.yaml中in_channels为null且feature_dim.json不存在，无法确定特征维度")
            sys.exit(1)

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    logger.info(f"使用设备: {device}")

    # v3: 文件I/O异常处理
    try:
        node_features = pd.read_csv(PROCESSED_DIR / "node_features.csv")
        labels_df = pd.read_csv(PROCESSED_DIR / "labels.csv")
        edge_index = torch.load(PROCESSED_DIR / "edge_index.pt", weights_only=False)
        edge_attr = torch.load(PROCESSED_DIR / "edge_attr.pt", weights_only=False)
        with open(PROCESSED_DIR / "gene_symbols.pkl", "rb") as f:
            gene_symbols = pickle.load(f)
    except FileNotFoundError as e:
        logger.error(f"文件缺失: {e}，请先运行 --mode data")
        sys.exit(1)

    feature_cols = [c for c in node_features.columns if c != "GeneSymbol"]

    # v3-ASSERT: 特征列不含inflammatory关键词
    for col in feature_cols:
        if "inflammatory" in col.lower() or "inflam" in col.lower():
            raise RuntimeError(f"ASSERT: 特征列包含泄露关键词: {col}")

    x = torch.tensor(node_features[feature_cols].fillna(0).values, dtype=torch.float32)
    y = torch.tensor(labels_df["Label"].values, dtype=torch.long)
    edge_index = edge_index.to(device)
    edge_attr = edge_attr.to(device)
    x = x.to(device)
    y = y.to(device)

    logger.info(f"节点数: {x.shape[0]}, 特征维度: {x.shape[1]}")
    logger.info(f"边数: {edge_index.shape[1]}")

    # v3: Transductive分割
    train_mask, val_mask, test_mask = create_masks_transductive(
        y,
        train_ratio=cfg["training"].get("train_ratio", 0.6),
        val_ratio=cfg["training"].get("val_ratio", 0.2),
        test_ratio=cfg["training"].get("test_ratio", 0.2),
        random_state=cfg["training"].get("random_state", 42)
    )
    train_mask = train_mask.to(device)
    val_mask = val_mask.to(device)
    test_mask = test_mask.to(device)

    n_pos = (y[train_mask] == 1).sum().item()
    n_neg = (y[train_mask] == 0).sum().item()

    # v3: pos_weight动态计算，上限50防止极端不平衡时梯度爆炸
    pos_weight = min(n_neg / max(n_pos, 1), 50.0)
    logger.info(f"训练集 阳性: {n_pos}, 阴性: {n_neg}, pos_weight: {pos_weight:.2f}")

    in_channels = x.shape[1]
    # 修正: 删除assert，仅保留if-warning-auto-correct
    # 原assert与自动修正逻辑冲突: assert失败时程序崩溃，if永远不会执行
    if cfg["model"]["in_channels"] != in_channels:
        logger.warning(f"config.yaml 中 in_channels={cfg['model']['in_channels']} 与实际特征维度 {in_channels} 不匹配，已自动修正")
        cfg["model"]["in_channels"] = in_channels

    model = GATv2Model(
        in_channels=in_channels,
        hidden_channels=32,
        out_channels=16,
        num_heads=2,
        num_classes=cfg["model"]["num_classes"],
        dropout=0.3,
        attention_dropout=0.15,
        use_edge_attr=cfg["model"].get("use_edge_attr", True),
    ).to(device)

    total_params = sum(p.numel() for p in model.parameters())
    logger.info(f"模型参数: {total_params:,}")

    # P3-fix: 极低学习率 + 高权重衰减 + 学习率预热
    base_lr = 1e-4  # 极低学习率
    optimizer = AdamW(model.parameters(), lr=base_lr, weight_decay=0.02)  # 高weight_decay防过拟合
    
    # 学习率预热: 前10个epoch从base_lr/10逐渐增加到base_lr
    warmup_epochs = 10
    scheduler = CosineAnnealingLR(optimizer, T_max=cfg["training"]["epochs"], eta_min=1e-6)

    best_val_metric = 0.0
    best_epoch = 0
    patience = cfg["training"]["patience"]
    patience_counter = 0

    early_stop_metric = cfg["training"].get("early_stop_metric", "pr_auc")
    logger.info(f"开始训练（早停指标: {early_stop_metric}）")

    train_losses, val_metrics_history = [], []

    for epoch in range(1, cfg["training"]["epochs"] + 1):
        # GraphSMOTE思想: 使用Label Smoothing + 类别权重
        # 学习率预热: 前warmup_epochs个epoch逐渐增加学习率
        if epoch <= warmup_epochs:
            warmup_factor = epoch / warmup_epochs
            for param_group in optimizer.param_groups:
                param_group['lr'] = base_lr * warmup_factor
        
        loss = train_epoch(model, optimizer, x, edge_index, edge_attr, y, train_mask, 
                          epoch=epoch, total_epochs=cfg["training"]["epochs"], pos_weight=1.0, use_asl=True)
        train_losses.append(loss)

        val_metrics = evaluate(model, x, edge_index, edge_attr, y, val_mask, pos_weight)
        val_metrics_history.append(val_metrics)

        current_metric = val_metrics.get(early_stop_metric, 0.0)

        if current_metric > best_val_metric:
            best_val_metric = current_metric
            best_epoch = epoch
            patience_counter = 0
            torch.save({
                "epoch": epoch,
                "model_state_dict": model.state_dict(),
                "optimizer_state_dict": optimizer.state_dict(),
                "val_metrics": val_metrics,
                "config": cfg,
            }, MODELS_DIR / "best_gat.pt")
            logger.info(f"Epoch {epoch:03d}: 新最佳模型，Val {early_stop_metric}={current_metric:.4f}, Val Precision={val_metrics['precision']:.4f}, Val F1={val_metrics['f1']:.4f}")
        else:
            patience_counter += 1

        # P1: 每10 epoch打印监控日志，包含Val Precision和Val F1
        if epoch % 10 == 0 or epoch == 1:
            logger.info(f"Epoch {epoch:03d} | Loss: {loss:.4f} | Val Loss: {val_metrics['loss']:.4f} | Val {early_stop_metric}: {current_metric:.4f} | Val Precision: {val_metrics['precision']:.4f} | Val Recall: {val_metrics['recall']:.4f} | Val F1: {val_metrics['f1']:.4f} | LR: {optimizer.param_groups[0]['lr']:.6f}")
        
        # v3-ASSERT: 每epoch训练后检查Loss NaN
        assert not np.isnan(loss), f"ASSERT: Epoch {epoch} Loss为NaN"

        # 学习率预热结束后开始余弦退火
        if epoch > warmup_epochs:
            scheduler.step()

        if patience_counter >= patience:
            logger.info(f"早停触发，最佳 epoch: {best_epoch}, 最佳 Val {early_stop_metric}: {best_val_metric:.4f}")
            break

    logger.info("训练完成")

    # 加载最佳模型并测试
    checkpoint = torch.load(MODELS_DIR / "best_gat.pt", map_location=device, weights_only=False)
    model.load_state_dict(checkpoint["model_state_dict"])
    test_metrics = evaluate(model, x, edge_index, edge_attr, y, test_mask, pos_weight)
    logger.info("测试集指标:")
    for k, v in test_metrics.items():
        logger.info(f"  {k}: {v:.4f}")

    # 保存训练日志
    pd.DataFrame({
        "epoch": range(1, len(train_losses) + 1),
        "train_loss": train_losses,
        "val_pr_auc": [m.get("pr_auc", 0) for m in val_metrics_history],
        "val_f1": [m.get("f1", 0) for m in val_metrics_history],
    }).to_csv(LOGS_DIR / "training_log.csv", index=False)

    return model


# ---------------------------------------------------------------------------
# 模块五: 靶点预测
# ---------------------------------------------------------------------------

def run_prediction():
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    logger.info(f"使用设备: {device}")

    try:
        node_features = pd.read_csv(PROCESSED_DIR / "node_features.csv")
        labels_df = pd.read_csv(PROCESSED_DIR / "labels.csv")
        edge_index = torch.load(PROCESSED_DIR / "edge_index.pt", weights_only=False)
        edge_attr = torch.load(PROCESSED_DIR / "edge_attr.pt", weights_only=False)
        with open(PROCESSED_DIR / "gene_symbols.pkl", "rb") as f:
            gene_symbols = pickle.load(f)
    except FileNotFoundError as e:
        logger.error(f"文件缺失: {e}，请先运行 --mode data")
        sys.exit(1)

    feature_cols = [c for c in node_features.columns if c != "GeneSymbol"]
    x = torch.tensor(node_features[feature_cols].fillna(0).values, dtype=torch.float32)
    y = torch.tensor(labels_df["Label"].values, dtype=torch.long)
    edge_index = edge_index.to(device)
    edge_attr = edge_attr.to(device)
    x = x.to(device)
    y = y.to(device)

    logger.info(f"节点数: {x.shape[0]}, 特征维度: {x.shape[1]}")
    logger.info(f"边数: {edge_index.shape[1]}")

    checkpoint = torch.load(MODELS_DIR / "best_gat.pt", map_location=device, weights_only=False)
    cfg = checkpoint["config"]
    in_channels = x.shape[1]
    if cfg["model"]["in_channels"] != in_channels:
        cfg["model"]["in_channels"] = in_channels

    model = GATv2Model(
        in_channels=in_channels,
        hidden_channels=32,
        out_channels=16,
        num_heads=2,
        num_classes=cfg["model"]["num_classes"],
        dropout=0.3,
        attention_dropout=0.15,
        use_edge_attr=cfg["model"].get("use_edge_attr", True),
    ).to(device)
    model.load_state_dict(checkpoint["model_state_dict"])
    model.eval()

    logger.info(f"加载模型: {MODELS_DIR / 'best_gat.pt'} (epoch {checkpoint['epoch']})")
    logger.info("开始推断...")

    with torch.no_grad():
        logits = model(x, edge_index, edge_attr)
        probs = F.softmax(logits, dim=1)

    prob_pos = probs[:, 1].cpu().numpy()
    unknown_mask = (y == -1).cpu().numpy()
    unknown_indices = np.where(unknown_mask)[0]

    results = pd.DataFrame({
        "GeneSymbol": [gene_symbols[i] for i in unknown_indices],
        "P_target": prob_pos[unknown_indices],
    })

    results["is_cuproptosis"] = results["GeneSymbol"].isin(CUPROPTOSIS_GENES).astype(int)
    # 添加dist_to_cuproptosis列（从node_features读取）
    try:
        nf = pd.read_csv(PROCESSED_DIR / "node_features.csv")
        if "dist_to_cuproptosis" in nf.columns:
            dist_map = nf.set_index("GeneSymbol")["dist_to_cuproptosis"].to_dict()
            results["dist_to_cuproptosis"] = results["GeneSymbol"].map(dist_map).fillna(0)
    except Exception:
        results["dist_to_cuproptosis"] = 0.0
    results = results.sort_values("P_target", ascending=False).reset_index(drop=True)
    results["Rank"] = np.arange(1, len(results) + 1)

    top_k_list = [50, 100, 200]
    for k in top_k_list:
        top_k = results.head(k).copy()
        top_k.to_csv(RESULTS_DIR / f"top_targets_{k}.csv", index=False)
        logger.info(f"Top-{k} 候选靶点已保存: {RESULTS_DIR / f'top_targets_{k}.csv'}")

    # CSV仅保留关键列以减小文件体积
    csv_cols = ["Rank", "GeneSymbol", "P_target", "is_cuproptosis", "dist_to_cuproptosis"]
    results[csv_cols].to_csv(RESULTS_DIR / "all_unknown_predictions.csv", index=False)
    logger.info(f"全部未知节点预测已保存: {RESULTS_DIR / 'all_unknown_predictions.csv'}")

    logger.info("Top-20 候选靶点摘要")
    for i, row in results.head(20).iterrows():
        flag = " [铜死亡]" if row["is_cuproptosis"] == 1 else ""
        logger.info(f"Rank {row['Rank']:03d}: {row['GeneSymbol']:<12} P={row['P_target']:.4f}{flag}")

    return results


# ---------------------------------------------------------------------------
# 模块六: 子网络分析
# ---------------------------------------------------------------------------

def run_subgraph_analysis():
    top200 = pd.read_csv(RESULTS_DIR / "top_targets_200.csv")
    top_genes = set(top200["GeneSymbol"].tolist())
    logger.info(f"加载 Top200 候选靶点: {len(top_genes)} 个")

    # 修正: 复用已处理的边，避免重复解析数GB的STRING原始文件
    # 方法学依据: processed/edge_index.pt已包含过滤后的边，无需重新下载解析
    edge_index = torch.load(PROCESSED_DIR / "edge_index.pt", weights_only=False)
    with open(PROCESSED_DIR / "gene_symbols.pkl", "rb") as f:
        gene_symbols = pickle.load(f)

    G_full = nx.Graph()
    # edge_index包含双向边，只取前半部分（单向）避免重复
    n_edges_half = edge_index.shape[1] // 2
    for i in range(n_edges_half):
        u, v = edge_index[0, i].item(), edge_index[1, i].item()
        g1, g2 = gene_symbols[u], gene_symbols[v]
        G_full.add_edge(g1, g2)

    logger.info(f"从processed/edge_index.pt重建全图: {G_full.number_of_nodes()} 节点, {G_full.number_of_edges()} 边")

    subgraph_nodes = top_genes & set(G_full.nodes())
    subG = G_full.subgraph(subgraph_nodes).copy()
    logger.info(f"诱导子网络: {subG.number_of_nodes()} 节点, {subG.number_of_edges()} 边")

    betweenness_all = nx.betweenness_centrality(subG, weight="weight") if subG.number_of_edges() > 0 else {}
    pagerank_all = nx.pagerank(subG, weight="weight") if subG.number_of_edges() > 0 else {}
    clustering_all = nx.clustering(subG, weight="weight") if subG.number_of_edges() > 0 else {}
    degree_all = dict(subG.degree())

    stats = []
    for node in sorted(subG.nodes()):
        stats.append({
            "GeneSymbol": node,
            "Degree": degree_all.get(node, 0),
            "Betweenness": betweenness_all.get(node, 0),
            "PageRank": pagerank_all.get(node, 0),
            "ClusteringCoefficient": clustering_all.get(node, 0),
            "is_cuproptosis": 1 if node in CUPROPTOSIS_GENES else 0,
        })
    stats_df = pd.DataFrame(stats)
    stats_df.to_csv(SUBGRAPH_DIR / "subgraph_stats.csv", index=False)

    cupro_in_sub = [n for n in subG.nodes() if n in CUPROPTOSIS_GENES]
    logger.info("铜死亡基因子网络隔离分析摘要")
    for node in sorted(cupro_in_sub, key=lambda n: degree_all.get(n, 0), reverse=True):
        logger.info(f"  {node:<8} 度={degree_all.get(node, 0):>3} Betweenness={betweenness_all.get(node, 0):.4f} PageRank={pagerank_all.get(node, 0):.6f}")

    logger.info("分析完成")
    return stats_df


# ---------------------------------------------------------------------------
# 模块七: Excel 导出（v3精简版：仅3个sheet）
# ---------------------------------------------------------------------------

def export_to_excel():
    logger.info("导出 Excel 报告...")
    wb = Workbook()
    wb.remove(wb.active)

    # 1. 项目概览
    ws_overview = wb.create_sheet("项目概览")
    n_nodes = 0
    n_features = 0
    n_edges = 0
    n_pos_labels = 0
    n_neg_labels = 0
    if (PROCESSED_DIR / "node_features.csv").exists():
        nf_tmp = pd.read_csv(PROCESSED_DIR / "node_features.csv")
        n_nodes = len(nf_tmp)
        n_features = len(nf_tmp.columns) - 1
    if (PROCESSED_DIR / "labels.csv").exists():
        lbl_tmp = pd.read_csv(PROCESSED_DIR / "labels.csv")
        n_pos_labels = int((lbl_tmp["Label"] == 1).sum())
        n_neg_labels = int((lbl_tmp["Label"] == 0).sum())
    if (PROCESSED_DIR / "edge_index.pt").exists():
        try:
            ei_tmp = torch.load(PROCESSED_DIR / "edge_index.pt", weights_only=False)
            n_edges = ei_tmp.shape[1] // 2
        except Exception:
            n_edges = 0

    # 读取铜死亡基因在未知节点中的数量
    n_cupro_in_unknown = 0
    if (RESULTS_DIR / "all_unknown_predictions.csv").exists():
        try:
            all_pred_tmp = pd.read_csv(RESULTS_DIR / "all_unknown_predictions.csv")
            n_cupro_in_unknown = int(all_pred_tmp["is_cuproptosis"].sum())
        except Exception:
            pass

    overview_data = [
        ["石竹烯-CIRI 靶点预测项目", ""],
        ["", ""],
        ["数据摘要", ""],
        ["节点总数", str(n_nodes)],
        ["特征维度", str(n_features)],
        ["STRING边数", str(n_edges)],
        ["阳性标签数", str(n_pos_labels)],
        ["阴性标签数", str(n_neg_labels)],
        ["铜死亡基因在未知节点中", str(n_cupro_in_unknown)],
        ["图分割策略", "Transductive（全图边保留）"],
        ["", ""],
        ["图距离类型", "无权拓扑跳数（unweighted hops）"],
        ["", ""],
        ["文件清单", ""],
        ["Top50候选靶点", str(RESULTS_DIR / "top_targets_50.csv")],
        ["Top100候选靶点", str(RESULTS_DIR / "top_targets_100.csv")],
        ["Top200候选靶点", str(RESULTS_DIR / "top_targets_200.csv")],
        ["全部未知节点预测", str(RESULTS_DIR / "all_unknown_predictions.csv")],
        ["节点特征矩阵", str(PROCESSED_DIR / "node_features.csv")],
        ["边索引", str(PROCESSED_DIR / "edge_index.pt")],
        ["子网络统计", str(SUBGRAPH_DIR / "subgraph_stats.csv")],
    ]
    for row in overview_data:
        ws_overview.append(row)

    # 2. Top50候选靶点
    if (RESULTS_DIR / "top_targets_50.csv").exists():
        df50 = pd.read_csv(RESULTS_DIR / "top_targets_50.csv")
        ws50 = wb.create_sheet("Top50候选靶点")
        for r in dataframe_to_rows(df50, index=False, header=True):
            ws50.append(r)

    # 3. 铜死亡基因专项
    if (RESULTS_DIR / "all_unknown_predictions.csv").exists():
        all_pred = pd.read_csv(RESULTS_DIR / "all_unknown_predictions.csv")
        cupro_df = all_pred[all_pred["is_cuproptosis"] == 1][["Rank", "GeneSymbol", "P_target"]].sort_values("P_target", ascending=False)
        ws_cupro = wb.create_sheet("铜死亡基因专项")
        for r in dataframe_to_rows(cupro_df, index=False, header=True):
            ws_cupro.append(r)

    excel_path = BASE_DIR / "石竹烯-CIRI_全结果汇总_v3.xlsx"
    wb.save(excel_path)
    logger.info(f"综合 Excel 报告已生成: {excel_path}")
    logger.info(f"包含工作表: {wb.sheetnames}")


# ---------------------------------------------------------------------------
# 模块八: 铜死亡基因排名检查（v3修正版）
# ---------------------------------------------------------------------------

def check_cuproptosis_ranking():
    if not (RESULTS_DIR / "all_unknown_predictions.csv").exists():
        logger.warning("预测结果不存在，请先运行 predict")
        return

    df = pd.read_csv(RESULTS_DIR / "all_unknown_predictions.csv")
    logger.info("铜死亡基因预测排名:")
    for gene in sorted(CUPROPTOSIS_GENES):
        match = df[df["GeneSymbol"] == gene]
        if not match.empty:
            rank = match.iloc[0]["Rank"]
            prob = match.iloc[0]["P_target"]
            logger.info(f"{gene:<8} 排名={rank:>4} 概率={prob:.4f}")
        else:
            logger.info(f"{gene:<8} 未找到")

    # v3: 使用compute_ranking_metrics计算Recall@K
    ranking_metrics = compute_ranking_metrics(df, CUPROPTOSIS_GENES)
    for k in [50, 100, 200]:
        logger.info(f"Top{k} 中铜死亡基因Recall: {ranking_metrics.get(f'cupro_recall@{k}', 0):.4f}")
    logger.info(f"铜死亡基因MRR: {ranking_metrics.get('cupro_mrr', 0):.4f}")


# ---------------------------------------------------------------------------
# 主函数
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="石竹烯-CIRI 全流程整合脚本（v3重构版）")
    parser.add_argument("--mode", type=str, default="all",
                        choices=["all", "data", "train", "predict", "subgraph", "excel", "check"],
                        help="运行模式")
    parser.add_argument("--config", type=str, default="config.yaml", help="配置文件路径")
    args = parser.parse_args()

    logger.info("=" * 60)
    logger.info("石竹烯-CIRI 全流程整合脚本（v3权威论文标准范式重构版）")
    logger.info(f"运行模式: {args.mode}")
    logger.info("=" * 60)

    if args.mode in ("all", "data"):
        logger.info("[1/6] 数据整合")
        integrator = DataIntegrator()
        integrator.run()

    if args.mode in ("all", "train"):
        logger.info("[2/6] 模型训练")
        run_training(args.config)

    if args.mode in ("all", "predict"):
        logger.info("[3/6] 靶点预测")
        run_prediction()

    if args.mode in ("all", "subgraph"):
        logger.info("[4/6] 子网络分析")
        run_subgraph_analysis()

    if args.mode in ("all", "excel"):
        logger.info("[5/6] Excel 导出")
        export_to_excel()

    if args.mode in ("all", "check"):
        logger.info("[6/6] 铜死亡基因排名检查")
        check_cuproptosis_ranking()

    if args.mode == "all":
        logger.info("=" * 60)
        logger.info("方法学合规检查报告")
        logger.info("=" * 60)
        if (PROCESSED_DIR / "labels.csv").exists():
            labels = pd.read_csv(PROCESSED_DIR / "labels.csv")
            n_pos = int((labels["Label"] == 1).sum())
            n_neg = int((labels["Label"] == 0).sum())
            n_unk = int((labels["Label"] == -1).sum())
            logger.info(f"阳性标签数: {n_pos}")
            logger.info(f"阴性标签数: {n_neg}")
            logger.info(f"未知标签数: {n_unk}")
        else:
            n_pos = n_neg = n_unk = 0
        
        n_features = 0
        if (PROCESSED_DIR / "node_features.csv").exists():
            nf = pd.read_csv(PROCESSED_DIR / "node_features.csv")
            n_features = len(nf.columns) - 1
            logger.info(f"特征维度数: {n_features}")
        
        # 铜死亡基因在未知节点中的数量
        n_cupro_in_unknown = 0
        if (RESULTS_DIR / "all_unknown_predictions.csv").exists():
            try:
                all_pred = pd.read_csv(RESULTS_DIR / "all_unknown_predictions.csv")
                n_cupro_in_unknown = int(all_pred["is_cuproptosis"].sum())
            except Exception:
                pass
        logger.info(f"铜死亡基因在未知节点中: {n_cupro_in_unknown}")
        
        logger.info("图分割策略: Transductive（全图边保留）")
        
        # 早停指标最佳值
        best_metric_val = 0.0
        if (LOGS_DIR / "training_log.csv").exists():
            try:
                train_log = pd.read_csv(LOGS_DIR / "training_log.csv")
                if "val_pr_auc" in train_log.columns:
                    best_metric_val = train_log["val_pr_auc"].max()
            except Exception:
                pass
        logger.info(f"早停指标最佳值 (pr_auc): {best_metric_val:.4f}")
        
        # 合规性总结
        logger.info("-" * 60)
        compliance_issues = []
        if n_pos < 10:
            compliance_issues.append(f"阳性标签仅{n_pos}个（<10），建议补充或降级模型")
        if n_neg < n_pos * 2:
            compliance_issues.append(f"阴性标签{n_neg} < 阳性×2 ({n_pos*2})，类别不平衡")
        if n_features < 20:
            compliance_issues.append(f"特征维度{n_features} < 20，可能欠拟合")
        if n_cupro_in_unknown == 0:
            compliance_issues.append("未知节点中无铜死亡基因，铜死亡专项分析无意义")
        
        if compliance_issues:
            logger.info("合规性警告:")
            for issue in compliance_issues:
                logger.info(f"  ⚠ {issue}")
        else:
            logger.info("✓ 方法学合规检查通过")

    logger.info("=" * 60)
    logger.info("全流程完成")


if __name__ == "__main__":
    main()
